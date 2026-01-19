"""
Year-by-Year Return Analysis for All 8 Trading Strategies
Generates an Excel file with formulas for yearly profit distribution (2020-2025)
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

# Base path
BASE_PATH = "/Users/sureshpatil/Desktop/Portfolio Creation"

def parse_csv_trades(filepath):
    """Parse trade data from MT4/MT5 CSV export files."""
    trades = []
    
    try:
        # Read file with flexible parsing
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        # Find the header row
        header_idx = None
        for i, line in enumerate(lines):
            if '#,Time,Type' in line or 'Time,Type' in line:
                header_idx = i
                break
        
        if header_idx is None:
            return trades
        
        # Read data starting from header
        df = pd.read_csv(filepath, skiprows=header_idx, encoding='utf-8', 
                         on_bad_lines='skip')
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Check for required columns
        time_col = None
        profit_col = None
        type_col = None
        
        for col in df.columns:
            if 'Time' in col:
                time_col = col
            if 'Profit' in col:
                profit_col = col
            if 'Type' in col:
                type_col = col
        
        if time_col is None or profit_col is None:
            return trades
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                time_val = str(row[time_col]).strip()
                profit_val = row[profit_col]
                
                # Skip rows without valid profit
                if pd.isna(profit_val) or str(profit_val).strip() == '':
                    continue
                
                # Parse profit value (handle comma-formatted numbers)
                if isinstance(profit_val, str):
                    profit_val = profit_val.replace(',', '').replace('"', '')
                    profit_val = float(profit_val)
                else:
                    profit_val = float(profit_val)
                
                # Parse date (format: YYYY.MM.DD HH:MM or similar)
                date_str = time_val.split(' ')[0]  # Get just the date part
                
                # Try different date formats
                for fmt in ['%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y']:
                    try:
                        trade_date = datetime.strptime(date_str, fmt)
                        break
                    except:
                        continue
                else:
                    continue
                
                year = trade_date.year
                
                # Only include years 2020-2025
                if 2020 <= year <= 2025:
                    trades.append({
                        'date': trade_date,
                        'year': year,
                        'profit': profit_val
                    })
                    
            except Exception as e:
                continue
        
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
    
    return trades


def parse_xlsx_trades(filepath):
    """Parse trade data from Excel files with MT4/MT5 strategy tester format."""
    trades = []
    
    try:
        xls = pd.ExcelFile(filepath)
        
        for sheet_name in xls.sheet_names:
            try:
                # Read entire sheet first to find the header row
                df_raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                
                # Find the row containing 'Time' header
                header_row = None
                for idx, row in df_raw.iterrows():
                    row_str = ' '.join([str(v).lower() for v in row.values if pd.notna(v)])
                    if 'time' in row_str and ('profit' in row_str or 'balance' in row_str):
                        header_row = idx
                        break
                
                if header_row is None:
                    continue
                
                # Re-read with correct header
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
                
                # Clean column names
                df.columns = [str(c).strip() for c in df.columns]
                
                # Find time and profit columns
                time_col = None
                profit_col = None
                
                for col in df.columns:
                    col_lower = col.lower()
                    if 'time' in col_lower:
                        time_col = col
                    if 'profit' in col_lower:
                        profit_col = col
                
                if time_col is None or profit_col is None:
                    continue
                
                for idx, row in df.iterrows():
                    try:
                        time_val = row[time_col]
                        profit_val = row[profit_col]
                        
                        # Skip empty profits or header values
                        if pd.isna(profit_val):
                            continue
                        
                        profit_str = str(profit_val).strip()
                        if profit_str == '' or profit_str.lower() == 'profit':
                            continue
                        
                        # Parse profit (handle various formats)
                        if isinstance(profit_val, str):
                            # Remove currency symbols, spaces, and commas
                            profit_val = profit_val.replace(',', '').replace('"', '')
                            profit_val = profit_val.replace(' ', '').replace('$', '')
                            if profit_val == '':
                                continue
                            # Check if it's a valid number
                            try:
                                profit_val = float(profit_val)
                            except ValueError:
                                continue
                        else:
                            try:
                                profit_val = float(profit_val)
                            except (ValueError, TypeError):
                                continue
                        
                        # Skip zero profits (often just position opens)
                        if profit_val == 0:
                            continue
                        
                        # Skip initial deposits (typically large round numbers like 100000, 10000, 2000)
                        # These appear as the first "profit" but are actually deposits
                        abs_profit = abs(profit_val)
                        if abs_profit in [100000, 100000.00, 10000, 10000.00, 2000, 2000.00, 1000, 1000.00]:
                            continue
                        
                        # Parse date
                        trade_date = None
                        if isinstance(time_val, datetime):
                            trade_date = time_val
                        elif isinstance(time_val, pd.Timestamp):
                            trade_date = time_val.to_pydatetime()
                        else:
                            time_str = str(time_val).strip()
                            if time_str.lower() == 'time' or time_str == 'nan':
                                continue
                            
                            # Extract date part
                            date_part = time_str.split(' ')[0]
                            
                            for fmt in ['%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y', '%d-%m-%Y']:
                                try:
                                    trade_date = datetime.strptime(date_part, fmt)
                                    break
                                except:
                                    continue
                        
                        if trade_date is None:
                            continue
                        
                        year = trade_date.year
                        
                        if 2020 <= year <= 2025:
                            trades.append({
                                'date': trade_date,
                                'year': year,
                                'profit': profit_val
                            })
                            
                    except Exception as e:
                        continue
                        
            except Exception as e:
                continue
                
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
    
    return trades


def get_strategy_data():
    """Collect trade data from all 8 strategies."""
    
    strategies = {}
    
    # 1. AURUM Strategy
    print("Processing AURUM Strategy...")
    aurum_trades = []
    aurum_files = [
        os.path.join(BASE_PATH, "AURUM/Gold /Gold - Indivisual TP.xlsx"),
        os.path.join(BASE_PATH, "AURUM/USDJPY/USDJPY - AVG TP.xlsx")
    ]
    for f in aurum_files:
        if os.path.exists(f):
            aurum_trades.extend(parse_xlsx_trades(f))
    strategies['AURUM'] = aurum_trades
    
    # 2. Falcon Strategy
    print("Processing Falcon Strategy...")
    falcon_trades = []
    falcon_csv = os.path.join(BASE_PATH, "Falcon/V5.csv")
    if os.path.exists(falcon_csv):
        falcon_trades.extend(parse_csv_trades(falcon_csv))
    strategies['Falcon'] = falcon_trades
    
    # 3. Gold Dip Strategy
    print("Processing Gold Dip Strategy...")
    gold_dip_trades = []
    gold_dip_folders = ['EURUSD', 'AUDUSD', 'GBPUSD', 'EURAUD', 'EURJPY', 'EURCHF', 'AUDJPY', 'USDCAD']
    for folder in gold_dip_folders:
        csv_path = os.path.join(BASE_PATH, f"Gold Dip/{folder}/{folder}.csv")
        if os.path.exists(csv_path):
            gold_dip_trades.extend(parse_csv_trades(csv_path))
    strategies['Gold Dip'] = gold_dip_trades
    
    # 4. Pair Trading EA Strategy
    print("Processing Pair Trading EA Strategy...")
    pair_trading_trades = []
    pair_folders = ['AUDUSD-AUDCAD', 'EURUSD-GBPUSD', 'EURUSD_AUDUSD', 'EURGBP-GBPCHF', 'USDCAD_AUDCHF']
    for folder in pair_folders:
        folder_path = os.path.join(BASE_PATH, f"Pair Trading EA/{folder}")
        if os.path.exists(folder_path):
            for f in os.listdir(folder_path):
                if f.endswith('.xlsx'):
                    pair_trading_trades.extend(parse_xlsx_trades(os.path.join(folder_path, f)))
    strategies['Pair Trading EA'] = pair_trading_trades
    
    # 5. Reversal Strategy
    print("Processing Reversal Strategy...")
    reversal_trades = []
    reversal_xlsx = os.path.join(BASE_PATH, "Reversal Strategy/All Pairs - 1 Day.xlsx")
    if os.path.exists(reversal_xlsx):
        reversal_trades.extend(parse_xlsx_trades(reversal_xlsx))
    strategies['Reversal Strategy'] = reversal_trades
    
    # 6. RSI 6 Trades Strategy
    print("Processing RSI 6 Trades Strategy...")
    rsi6_trades = []
    rsi6_folders = ['GBPAUD', 'EURAUD', 'GBPUSD', 'EURGBP', 'GBPCAD', 'EURCAD', 'EURUSD', 
                    'USDCHF', 'NZDCHF', 'USDJPY', 'CADCHF', 'GBPCHF', 'EURCHF', 'USDCAD', 'AUDUSD', 'AUDNZD']
    for folder in rsi6_folders:
        xlsx_path = os.path.join(BASE_PATH, f"RSI 6 trades/{folder}/{folder}.xlsx")
        if os.path.exists(xlsx_path):
            rsi6_trades.extend(parse_xlsx_trades(xlsx_path))
    strategies['RSI 6 Trades'] = rsi6_trades
    
    # 7. RSI Correlation Strategy
    print("Processing RSI Correlation Strategy...")
    rsi_corr_trades = []
    rsi_corr_folders = ['AUDUSD_GBPNZD', 'EURAUD_CADCHF', 'EURGBP_GBPCHF', 
                        'GBPUSD_USDCAD', 'GBPUSD_USDCHF', 'USDCAD_AUDCHF']
    for folder in rsi_corr_folders:
        folder_path = os.path.join(BASE_PATH, f"RSI corelation/{folder}")
        if os.path.exists(folder_path):
            for f in os.listdir(folder_path):
                if f.endswith('.xlsx'):
                    rsi_corr_trades.extend(parse_xlsx_trades(os.path.join(folder_path, f)))
    strategies['RSI Correlation'] = rsi_corr_trades
    
    # 8. 7th Strategy (RSI Pyramiding)
    print("Processing 7th Strategy (RSI Pyramiding)...")
    strategy7_trades = []
    strategy7_files = [
        os.path.join(BASE_PATH, "7th strategy/XAUUSD 20-25.csv"),
        os.path.join(BASE_PATH, "7th strategy/XAGUSD 20-25.csv")
    ]
    for f in strategy7_files:
        if os.path.exists(f):
            strategy7_trades.extend(parse_csv_trades(f))
    strategies['7th Strategy (RSI Pyramiding)'] = strategy7_trades
    
    return strategies


def calculate_yearly_returns(strategies):
    """Calculate year-by-year returns for each strategy."""
    
    years = [2020, 2021, 2022, 2023, 2024, 2025]
    
    results = {}
    
    for strategy_name, trades in strategies.items():
        if not trades:
            results[strategy_name] = {year: 0.0 for year in years}
            continue
        
        yearly_profits = {year: 0.0 for year in years}
        
        for trade in trades:
            year = trade['year']
            if year in yearly_profits:
                yearly_profits[year] += trade['profit']
        
        results[strategy_name] = yearly_profits
    
    return results


def create_excel_with_formulas(results, output_path):
    """Create Excel file with formulas for yearly returns analysis."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Yearly Returns Analysis"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=11)
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    years = [2020, 2021, 2022, 2023, 2024, 2025]
    strategies = list(results.keys())
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "Year-by-Year Return Distribution Analysis - All 8 Trading Strategies"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Generated date
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Section 1: Raw Profit Data (for formulas to reference)
    ws['A4'] = "SECTION 1: YEARLY PROFIT DATA ($)"
    ws['A4'].font = Font(bold=True, size=14)
    
    # Headers for profit data
    headers = ['Strategy'] + [str(year) for year in years] + ['Total Profit', 'Avg Yearly']
    start_row = 6
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows with raw profit values
    for row_idx, strategy in enumerate(strategies, start_row + 1):
        # Strategy name
        ws.cell(row=row_idx, column=1, value=strategy).border = border
        
        # Yearly profits
        for col_idx, year in enumerate(years, 2):
            profit = results[strategy][year]
            cell = ws.cell(row=row_idx, column=col_idx, value=round(profit, 2))
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            
            # Conditional formatting
            if profit > 0:
                cell.fill = positive_fill
            elif profit < 0:
                cell.fill = negative_fill
        
        # Total Profit formula
        total_col = len(years) + 2
        total_cell = ws.cell(row=row_idx, column=total_col)
        total_cell.value = f"=SUM(B{row_idx}:G{row_idx})"
        total_cell.number_format = '#,##0.00'
        total_cell.border = border
        total_cell.font = Font(bold=True)
        
        # Average Yearly formula
        avg_col = len(years) + 3
        avg_cell = ws.cell(row=row_idx, column=avg_col)
        avg_cell.value = f"=AVERAGE(B{row_idx}:G{row_idx})"
        avg_cell.number_format = '#,##0.00'
        avg_cell.border = border
    
    # Total row for all strategies
    total_row = start_row + len(strategies) + 1
    ws.cell(row=total_row, column=1, value="PORTFOLIO TOTAL").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = border
    
    for col_idx, year in enumerate(years, 2):
        cell = ws.cell(row=total_row, column=col_idx)
        col_letter = get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}{start_row + 1}:{col_letter}{total_row - 1})"
        cell.number_format = '#,##0.00'
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
    
    # Total of totals
    total_col = len(years) + 2
    total_total = ws.cell(row=total_row, column=total_col)
    total_total.value = f"=SUM(H{start_row + 1}:H{total_row - 1})"
    total_total.number_format = '#,##0.00'
    total_total.fill = total_fill
    total_total.font = total_font
    total_total.border = border
    
    # Average of averages
    avg_col = len(years) + 3
    avg_total = ws.cell(row=total_row, column=avg_col)
    avg_total.value = f"=AVERAGE(I{start_row + 1}:I{total_row - 1})"
    avg_total.number_format = '#,##0.00'
    avg_total.fill = total_fill
    avg_total.font = total_font
    avg_total.border = border
    
    # Section 2: Percentage Distribution
    section2_start = total_row + 3
    ws.cell(row=section2_start, column=1, value="SECTION 2: YEARLY PROFIT AS % OF TOTAL PROFIT").font = Font(bold=True, size=14)
    
    header_row2 = section2_start + 2
    for col, header in enumerate(headers[:-1], 1):  # Exclude 'Avg Yearly' for this section
        cell = ws.cell(row=header_row2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Add "% of Total" header
    cell = ws.cell(row=header_row2, column=len(headers) - 1, value="Total %")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    
    # Percentage data rows
    for row_idx, strategy in enumerate(strategies, header_row2 + 1):
        data_source_row = start_row + 1 + (row_idx - header_row2 - 1)
        
        ws.cell(row=row_idx, column=1, value=strategy).border = border
        
        for col_idx, year in enumerate(years, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_letter = get_column_letter(col_idx)
            # Formula: year profit / total profit (with error handling for zero division)
            cell.value = f"=IF(H{data_source_row}=0,0,{col_letter}{data_source_row}/H{data_source_row})"
            cell.number_format = '0.00%'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
        
        # Total percentage (should always be 100% or 0%)
        total_pct_cell = ws.cell(row=row_idx, column=len(years) + 2)
        total_pct_cell.value = f"=SUM(B{row_idx}:G{row_idx})"
        total_pct_cell.number_format = '0.00%'
        total_pct_cell.border = border
        total_pct_cell.font = Font(bold=True)
    
    # Section 3: Year-over-Year Growth
    section3_start = header_row2 + len(strategies) + 3
    ws.cell(row=section3_start, column=1, value="SECTION 3: YEAR-OVER-YEAR PROFIT COMPARISON").font = Font(bold=True, size=14)
    
    header_row3 = section3_start + 2
    yoy_headers = ['Strategy', '2020→2021', '2021→2022', '2022→2023', '2023→2024', '2024→2025']
    for col, header in enumerate(yoy_headers, 1):
        cell = ws.cell(row=header_row3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # YoY change data rows
    for row_idx, strategy in enumerate(strategies, header_row3 + 1):
        data_source_row = start_row + 1 + (row_idx - header_row3 - 1)
        
        ws.cell(row=row_idx, column=1, value=strategy).border = border
        
        # YoY change formulas
        year_cols = ['B', 'C', 'D', 'E', 'F', 'G']
        for col_idx in range(2, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            prev_col = year_cols[col_idx - 2]
            curr_col = year_cols[col_idx - 1]
            cell.value = f"={curr_col}{data_source_row}-{prev_col}{data_source_row}"
            cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
    
    # Section 4: Portfolio Contribution Analysis
    section4_start = header_row3 + len(strategies) + 3
    ws.cell(row=section4_start, column=1, value="SECTION 4: STRATEGY CONTRIBUTION TO YEARLY PORTFOLIO PROFIT (%)").font = Font(bold=True, size=14)
    
    header_row4 = section4_start + 2
    for col, header in enumerate(headers[:len(years)+1], 1):
        cell = ws.cell(row=header_row4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Contribution percentage rows
    for row_idx, strategy in enumerate(strategies, header_row4 + 1):
        data_source_row = start_row + 1 + (row_idx - header_row4 - 1)
        
        ws.cell(row=row_idx, column=1, value=strategy).border = border
        
        for col_idx, year in enumerate(years, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_letter = get_column_letter(col_idx)
            # Formula: strategy yearly profit / portfolio yearly profit (with error handling)
            cell.value = f"=IF({col_letter}{total_row}=0,0,{col_letter}{data_source_row}/{col_letter}{total_row})"
            cell.number_format = '0.00%'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
    
    # Total row (should be 100%)
    total_row4 = header_row4 + len(strategies) + 1
    ws.cell(row=total_row4, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row4, column=1).fill = total_fill
    ws.cell(row=total_row4, column=1).border = border
    
    for col_idx in range(2, len(years) + 2):
        cell = ws.cell(row=total_row4, column=col_idx)
        col_letter = get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}{header_row4 + 1}:{col_letter}{total_row4 - 1})"
        cell.number_format = '0.00%'
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
    
    # Section 5: Summary Statistics
    section5_start = total_row4 + 3
    ws.cell(row=section5_start, column=1, value="SECTION 5: SUMMARY STATISTICS").font = Font(bold=True, size=14)
    
    stats_start = section5_start + 2
    stats_headers = ['Metric', 'Value', 'Formula Used']
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=stats_start, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    stats = [
        ('Best Single Year (All Strategies)', f"=MAX(B{total_row}:G{total_row})", 'MAX of portfolio yearly totals'),
        ('Worst Single Year (All Strategies)', f"=MIN(B{total_row}:G{total_row})", 'MIN of portfolio yearly totals'),
        ('Average Yearly Return (Portfolio)', f"=AVERAGE(B{total_row}:G{total_row})", 'AVERAGE of portfolio yearly totals'),
        ('Total Portfolio Profit', f"=H{total_row}", 'Sum of all yearly profits'),
        ('Years with Positive Returns', f"=COUNTIF(B{total_row}:G{total_row},\">0\")", 'Count of years > 0'),
        ('Years with Negative Returns', f"=COUNTIF(B{total_row}:G{total_row},\"<0\")", 'Count of years < 0'),
        ('Best Performing Strategy', f"=INDEX(A{start_row+1}:A{total_row-1},MATCH(MAX(H{start_row+1}:H{total_row-1}),H{start_row+1}:H{total_row-1},0))", 'Strategy with highest total'),
        ('Worst Performing Strategy', f"=INDEX(A{start_row+1}:A{total_row-1},MATCH(MIN(H{start_row+1}:H{total_row-1}),H{start_row+1}:H{total_row-1},0))", 'Strategy with lowest total'),
    ]
    
    for row_offset, (metric, formula, description) in enumerate(stats):
        row = stats_start + row_offset + 1
        ws.cell(row=row, column=1, value=metric).border = border
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = '#,##0.00' if 'Best' not in metric and 'Worst' not in metric.split()[0] else '@'
        cell.border = border
        ws.cell(row=row, column=3, value=description).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 32
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['C'].width = 14  # Adjust for formula column in stats
    
    # Save workbook
    wb.save(output_path)
    print(f"\nExcel file saved to: {output_path}")


def main():
    print("=" * 60)
    print("Year-by-Year Return Analysis for All 8 Trading Strategies")
    print("=" * 60 + "\n")
    
    # Collect trade data from all strategies
    print("Collecting trade data from all strategies...\n")
    strategies = get_strategy_data()
    
    # Show trade counts
    print("\n" + "-" * 40)
    print("Trade Data Summary:")
    print("-" * 40)
    for name, trades in strategies.items():
        print(f"  {name}: {len(trades)} trades found")
    
    # Calculate yearly returns
    print("\nCalculating year-by-year returns...")
    results = calculate_yearly_returns(strategies)
    
    # Display summary
    print("\n" + "=" * 80)
    print("YEARLY PROFIT SUMMARY (2020-2025)")
    print("=" * 80)
    print(f"\n{'Strategy':<30} {'2020':>10} {'2021':>10} {'2022':>10} {'2023':>10} {'2024':>10} {'2025':>10} {'TOTAL':>12}")
    print("-" * 112)
    
    portfolio_totals = {year: 0.0 for year in [2020, 2021, 2022, 2023, 2024, 2025]}
    
    for strategy, yearly_profits in results.items():
        row = f"{strategy:<30}"
        total = 0
        for year in [2020, 2021, 2022, 2023, 2024, 2025]:
            profit = yearly_profits[year]
            row += f" {profit:>10.2f}"
            total += profit
            portfolio_totals[year] += profit
        row += f" {total:>12.2f}"
        print(row)
    
    print("-" * 112)
    total_row = f"{'PORTFOLIO TOTAL':<30}"
    grand_total = 0
    for year in [2020, 2021, 2022, 2023, 2024, 2025]:
        total_row += f" {portfolio_totals[year]:>10.2f}"
        grand_total += portfolio_totals[year]
    total_row += f" {grand_total:>12.2f}"
    print(total_row)
    
    # Create Excel file with formulas
    output_path = os.path.join(BASE_PATH, "Yearly_Returns_Analysis.xlsx")
    create_excel_with_formulas(results, output_path)
    
    print("\n" + "=" * 60)
    print("Analysis Complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
