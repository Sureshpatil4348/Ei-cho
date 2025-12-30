"""
Comprehensive Correlation Analysis
Creates a detailed correlation analysis showing:
1. Correlation between pairs within each strategy
2. Correlation between different strategies
3. Visual correlation matrices and heatmaps
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os
from datetime import datetime

# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

def load_csv_equity_curve(file_path, pair_name):
    """Load equity curve from CSV file"""
    try:
        df = pd.read_csv(file_path, skiprows=2)
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Filter only close trades with balance changes
        df = df[df['Type'].str.contains('close', case=False, na=False)]
        df = df[['Time', 'Balance']].copy()
        
        # Clean balance column (remove commas and quotes)
        df['Balance'] = df['Balance'].astype(str).str.replace(',', '').str.replace('"', '')
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
        
        # Parse time
        df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
        
        # Drop rows with NaN
        df = df.dropna()
        
        # Set time as index
        df = df.set_index('Time')
        df = df.sort_index()
        
        # Rename column
        df.columns = [pair_name]
        
        return df
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None

def load_excel_equity_curve(file_path, pair_name):
    """Load equity curve from Excel file"""
    try:
        df = pd.read_excel(file_path, skiprows=2)
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Filter only close trades with balance changes
        df = df[df['Type'].str.contains('close', case=False, na=False)]
        df = df[['Time', 'Balance']].copy()
        
        # Clean balance column
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
        
        # Parse time
        df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
        
        # Drop rows with NaN
        df = df.dropna()
        
        # Set time as index
        df = df.set_index('Time')
        df = df.sort_index()
        
        # Rename column
        df.columns = [pair_name]
        
        return df
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None

# ============================================================================
# STRATEGY DATA PATHS
# ============================================================================

STRATEGY_PATHS = {
    '7th_Strategy': [
        ('/Users/sureshpatil/Desktop/Portfolio Creation/7th strategy/XAUUSD 20-25.csv', 'XAUUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/7th strategy/XAGUSD 20-25.csv', 'XAGUSD'),
    ],
    'Falcon': [
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Falcon/V5.csv', 'EURUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Falcon/v5-v2 - Tp 60,SL 60 all day.csv', 'EURUSD_V2'),
    ],
    'Gold_Dip': [
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/EURUSD/EURUSD.csv', 'EURUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/GBPUSD/GBPUSD.csv', 'GBPUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/AUDUSD/AUDUSD.csv', 'AUDUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/USDCAD/USDCAD.csv', 'USDCAD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/EURJPY/EURJPY.csv', 'EURJPY'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/AUDJPY/AUDJPY.csv', 'AUDJPY'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/EURAUD/EURAUD.csv', 'EURAUD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/EURCHF/EURCHF.csv', 'EURCHF'),
    ],
    'RSI_6_Trades': [
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/EURUSD/EURUSD.xlsx', 'EURUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/GBPUSD/GBPUSD.xlsx', 'GBPUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/AUDUSD/AUDUSD.xlsx', 'AUDUSD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/USDCAD/USDCAD.xlsx', 'USDCAD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/USDJPY/USDJPY.xlsx', 'USDJPY'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/USDCHF/USDCHF.xlsx', 'USDCHF'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/EURAUD/EURAUD.xlsx', 'EURAUD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/EURCAD/EURCAD.xlsx', 'EURCAD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/EURCHF/EURCHF.xlsx', 'EURCHF'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/EURGBP/EURGBP.xlsx', 'EURGBP'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/GBPAUD/GBPAUD.xlsx', 'GBPAUD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/GBPCAD/GBPCAD.xlsx', 'GBPCAD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/GBPCHF/GBPCHF.xlsx', 'GBPCHF'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/AUDNZD/AUDNZD.xlsx', 'AUDNZD'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/NZDCHF/NZDCHF.xlsx', 'NZDCHF'),
        ('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/CADCHF/CADCHF.xlsx', 'CADCHF'),
    ],
}

# ============================================================================
# CORRELATION CALCULATION
# ============================================================================

def calculate_daily_returns(equity_df):
    """Calculate daily returns from equity curve"""
    # Remove duplicate timestamps (keep last)
    equity_df = equity_df[~equity_df.index.duplicated(keep='last')]
    
    # Resample to daily frequency
    daily_equity = equity_df.resample('D').last().ffill()
    
    # Calculate returns
    returns = daily_equity.pct_change().fillna(0)
    
    return returns

def load_strategy_data(strategy_name, paths):
    """Load all pairs for a strategy"""
    print(f"\nLoading {strategy_name}...")
    
    equity_curves = []
    pair_names = []
    
    for file_path, pair_name in paths:
        if not os.path.exists(file_path):
            print(f"  Warning: File not found: {file_path}")
            continue
            
        # Load based on file extension
        if file_path.endswith('.xlsx'):
            equity_df = load_excel_equity_curve(file_path, pair_name)
        else:
            equity_df = load_csv_equity_curve(file_path, pair_name)
        
        if equity_df is not None and len(equity_df) > 0:
            print(f"  ✓ Loaded {pair_name}: {len(equity_df)} data points")
            equity_curves.append(equity_df)
            pair_names.append(pair_name)
        else:
            print(f"  ✗ Failed to load {pair_name}")
    
    if not equity_curves:
        return None, None
    
    # Calculate returns for each equity curve separately to avoid index issues
    returns_list = []
    for equity_df in equity_curves:
        # Remove duplicate timestamps
        equity_df = equity_df[~equity_df.index.duplicated(keep='last')]
        # Resample to daily
        daily_equity = equity_df.resample('D').last().ffill()
        # Calculate returns
        ret = daily_equity.pct_change().fillna(0)
        returns_list.append(ret)
    
    # Merge returns with outer join
    returns = pd.concat(returns_list, axis=1, join='outer')
    returns = returns.fillna(0)
    
    return returns, pair_names

def calculate_correlation_matrix(returns_df):
    """Calculate correlation matrix from returns"""
    return returns_df.corr()

# ============================================================================
# EXCEL STYLING
# ============================================================================

def style_header(ws, row, start_col, end_col):
    """Apply header styling"""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_strategy_header(ws, row, start_col, end_col, color="4472C4"):
    """Apply strategy section header styling"""
    strategy_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    strategy_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = strategy_fill
        cell.font = strategy_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def add_border(ws, start_row, end_row, start_col, end_col):
    """Add border to a range"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border

def apply_correlation_color_scale(ws, start_row, end_row, start_col, end_col):
    """Apply color scale to correlation values"""
    # Create color scale rule: Red (negative) -> White (0) -> Green (positive)
    rule = ColorScaleRule(
        start_type='num', start_value=-1, start_color='FF6B6B',
        mid_type='num', mid_value=0, mid_color='FFFFFF',
        end_type='num', end_value=1, end_color='4ECB71'
    )
    
    # Apply to range
    range_string = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    ws.conditional_formatting.add(range_string, rule)

# ============================================================================
# CREATE CORRELATION SHEETS
# ============================================================================

def create_within_strategy_correlation_sheet(wb):
    """Create sheet showing correlation within each strategy"""
    ws = wb.create_sheet("Within_Strategy_Correlations")
    
    row = 1
    
    # Main title
    ws.cell(row=row, column=1, value="CORRELATION ANALYSIS: PAIRS WITHIN EACH STRATEGY")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value="How correlated are the pairs within each strategy? (Based on daily returns)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    row += 2
    
    strategy_colors = ["4472C4", "ED7D31", "70AD47", "5B9BD5"]
    
    for idx, (strategy_name, paths) in enumerate(STRATEGY_PATHS.items()):
        # Load strategy data
        returns, pair_names = load_strategy_data(strategy_name, paths)
        
        if returns is None or len(pair_names) < 2:
            print(f"Skipping {strategy_name} - insufficient data")
            continue
        
        # Calculate correlation
        corr_matrix = calculate_correlation_matrix(returns)
        
        # Strategy header
        ws.cell(row=row, column=1, value=f"STRATEGY: {strategy_name}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(pair_names)+1)
        style_strategy_header(ws, row, 1, len(pair_names)+1, strategy_colors[idx % len(strategy_colors)])
        row += 1
        
        # Correlation matrix headers
        ws.cell(row=row, column=1, value="Pair")
        for col_idx, pair in enumerate(pair_names, 2):
            ws.cell(row=row, column=col_idx, value=pair)
        style_header(ws, row, 1, len(pair_names)+1)
        row += 1
        
        start_data_row = row
        
        # Correlation matrix data
        for i, pair in enumerate(pair_names):
            ws.cell(row=row, column=1, value=pair)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for j, _ in enumerate(pair_names):
                corr_val = corr_matrix.iloc[i, j]
                cell = ws.cell(row=row, column=j+2, value=round(corr_val, 4))
                cell.alignment = Alignment(horizontal='center')
                
                # Diagonal cells (self-correlation)
                if i == j:
                    cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            
            row += 1
        
        # Apply color scale to correlation values
        apply_correlation_color_scale(ws, start_data_row, row-1, 2, len(pair_names)+1)
        add_border(ws, start_data_row-1, row-1, 1, len(pair_names)+1)
        
        # Statistics
        row += 1
        ws.cell(row=row, column=1, value="Statistics:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1
        
        # Get upper triangle of correlation matrix (excluding diagonal)
        mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
        correlations = corr_matrix.where(mask).stack().values
        
        ws.cell(row=row, column=1, value="Average Correlation:")
        ws.cell(row=row, column=2, value=round(np.mean(correlations), 4))
        row += 1
        
        ws.cell(row=row, column=1, value="Min Correlation:")
        ws.cell(row=row, column=2, value=round(np.min(correlations), 4))
        row += 1
        
        ws.cell(row=row, column=1, value="Max Correlation:")
        ws.cell(row=row, column=2, value=round(np.max(correlations), 4))
        row += 1
        
        ws.cell(row=row, column=1, value="Std Dev:")
        ws.cell(row=row, column=2, value=round(np.std(correlations), 4))
        row += 1
        
        # Diversification benefit
        avg_corr = np.mean(correlations)
        n_pairs = len(pair_names)
        diversification_benefit = 1 - avg_corr
        
        ws.cell(row=row, column=1, value="Diversification Benefit:")
        ws.cell(row=row, column=2, value=round(diversification_benefit * 100, 2))
        ws.cell(row=row, column=3, value="%")
        ws.cell(row=row, column=1).font = Font(bold=True, color="006600")
        ws.cell(row=row, column=2).font = Font(bold=True, color="006600")
        row += 1
        
        ws.cell(row=row, column=1, value="Interpretation:")
        ws.cell(row=row, column=1).font = Font(italic=True)
        
        if avg_corr < 0.3:
            interpretation = "Excellent diversification - pairs are weakly correlated"
            color = "006600"
        elif avg_corr < 0.5:
            interpretation = "Good diversification - moderate correlation"
            color = "0066CC"
        elif avg_corr < 0.7:
            interpretation = "Moderate diversification - pairs are correlated"
            color = "FF8800"
        else:
            interpretation = "Low diversification - pairs are highly correlated"
            color = "CC0000"
        
        ws.cell(row=row, column=2, value=interpretation)
        ws.cell(row=row, column=2).font = Font(color=color)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 3
    
    # Adjust column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 25
    
    return ws

def create_between_strategy_correlation_sheet(wb):
    """Create sheet showing correlation between strategies"""
    ws = wb.create_sheet("Between_Strategy_Correlations")
    
    row = 1
    
    # Main title
    ws.cell(row=row, column=1, value="CORRELATION ANALYSIS: BETWEEN STRATEGIES")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value="How correlated are the different strategies? (Based on combined daily returns)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    row += 2
    
    # Load data for each strategy (combined)
    strategy_returns = {}
    strategy_names = []
    
    for strategy_name, paths in STRATEGY_PATHS.items():
        returns, pair_names = load_strategy_data(strategy_name, paths)
        
        if returns is None:
            continue
        
        # Calculate combined strategy return (equal weighted)
        combined_return = returns.mean(axis=1)
        strategy_returns[strategy_name] = combined_return
        strategy_names.append(strategy_name)
    
    if len(strategy_names) < 2:
        ws.cell(row=row, column=1, value="Error: Insufficient strategy data for correlation analysis")
        return ws
    
    # Combine all strategy returns
    strategy_df = pd.DataFrame(strategy_returns)
    
    # Calculate correlation matrix
    corr_matrix = strategy_df.corr()
    
    # Section header
    ws.cell(row=row, column=1, value="STRATEGY CORRELATION MATRIX")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(strategy_names)+1)
    style_strategy_header(ws, row, 1, len(strategy_names)+1, "1F4E79")
    row += 1
    
    # Matrix headers
    ws.cell(row=row, column=1, value="Strategy")
    for col_idx, strategy in enumerate(strategy_names, 2):
        ws.cell(row=row, column=col_idx, value=strategy)
    style_header(ws, row, 1, len(strategy_names)+1)
    row += 1
    
    start_data_row = row
    
    # Matrix data
    for i, strategy in enumerate(strategy_names):
        ws.cell(row=row, column=1, value=strategy)
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for j, _ in enumerate(strategy_names):
            corr_val = corr_matrix.iloc[i, j]
            cell = ws.cell(row=row, column=j+2, value=round(corr_val, 4))
            cell.alignment = Alignment(horizontal='center')
            
            # Diagonal cells
            if i == j:
                cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        
        row += 1
    
    # Apply color scale
    apply_correlation_color_scale(ws, start_data_row, row-1, 2, len(strategy_names)+1)
    add_border(ws, start_data_row-1, row-1, 1, len(strategy_names)+1)
    
    row += 2
    
    # Statistics
    ws.cell(row=row, column=1, value="PORTFOLIO DIVERSIFICATION ANALYSIS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_strategy_header(ws, row, 1, 6, "70AD47")
    row += 1
    
    # Get upper triangle (excluding diagonal)
    mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
    correlations = corr_matrix.where(mask).stack().values
    
    ws.cell(row=row, column=1, value="Average Strategy Correlation:")
    ws.cell(row=row, column=2, value=round(np.mean(correlations), 4))
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1, value="Min Correlation:")
    ws.cell(row=row, column=2, value=round(np.min(correlations), 4))
    row += 1
    
    ws.cell(row=row, column=1, value="Max Correlation:")
    ws.cell(row=row, column=2, value=round(np.max(correlations), 4))
    row += 1
    
    ws.cell(row=row, column=1, value="Std Dev:")
    ws.cell(row=row, column=2, value=round(np.std(correlations), 4))
    row += 2
    
    # Portfolio diversification benefit
    avg_corr = np.mean(correlations)
    n_strategies = len(strategy_names)
    
    # Portfolio variance reduction
    variance_reduction = (1 - avg_corr) * (n_strategies - 1) / n_strategies * 100
    
    ws.cell(row=row, column=1, value="Portfolio Variance Reduction:")
    ws.cell(row=row, column=2, value=round(variance_reduction, 2))
    ws.cell(row=row, column=3, value="%")
    ws.cell(row=row, column=1).font = Font(bold=True, color="006600", size=11)
    ws.cell(row=row, column=2).font = Font(bold=True, color="006600", size=11)
    row += 1
    
    ws.cell(row=row, column=1, value="Diversification Benefit:")
    ws.cell(row=row, column=2, value=round((1 - avg_corr) * 100, 2))
    ws.cell(row=row, column=3, value="%")
    ws.cell(row=row, column=1).font = Font(bold=True, color="006600", size=11)
    ws.cell(row=row, column=2).font = Font(bold=True, color="006600", size=11)
    row += 2
    
    # Interpretation
    ws.cell(row=row, column=1, value="INTERPRETATION & RECOMMENDATIONS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_strategy_header(ws, row, 1, 6, "C00000")
    row += 1
    
    ws.cell(row=row, column=1, value="Strategy Correlation:")
    
    if avg_corr < 0.3:
        interpretation = "Excellent - Strategies are weakly correlated, providing strong diversification"
        color = "006600"
        recommendation = "Deploy all strategies with confidence. Low correlation maximizes risk-adjusted returns."
    elif avg_corr < 0.5:
        interpretation = "Good - Moderate correlation provides decent diversification"
        color = "0066CC"
        recommendation = "Good portfolio mix. Consider equal or inverse volatility weighting."
    elif avg_corr < 0.7:
        interpretation = "Fair - Strategies are somewhat correlated"
        color = "FF8800"
        recommendation = "Use risk parity allocation. Monitor overlapping exposures."
    else:
        interpretation = "Poor - Strategies are highly correlated"
        color = "CC0000"
        recommendation = "Consider removing redundant strategies or reducing overall exposure."
    
    ws.cell(row=row, column=2, value=interpretation)
    ws.cell(row=row, column=2).font = Font(color=color, bold=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    
    ws.cell(row=row, column=1, value="Recommendation:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=recommendation)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 3
    
    # Pairwise correlations detailed analysis
    ws.cell(row=row, column=1, value="PAIRWISE STRATEGY CORRELATIONS (Detailed)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    style_strategy_header(ws, row, 1, 5, "4472C4")
    row += 1
    
    headers = ['Strategy 1', 'Strategy 2', 'Correlation', 'Strength', 'Notes']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_header(ws, row, 1, 5)
    row += 1
    
    start_pair_row = row
    
    # List all pairwise correlations
    pairs = []
    for i in range(len(strategy_names)):
        for j in range(i+1, len(strategy_names)):
            corr = corr_matrix.iloc[i, j]
            pairs.append((strategy_names[i], strategy_names[j], corr))
    
    # Sort by absolute correlation (descending)
    pairs.sort(key=lambda x: abs(x[2]), reverse=True)
    
    for strat1, strat2, corr in pairs:
        ws.cell(row=row, column=1, value=strat1)
        ws.cell(row=row, column=2, value=strat2)
        ws.cell(row=row, column=3, value=round(corr, 4))
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        abs_corr = abs(corr)
        if abs_corr < 0.3:
            strength = "Weak"
            color_str = "006600"
            notes = "Excellent for diversification"
        elif abs_corr < 0.5:
            strength = "Moderate"
            color_str = "0066CC"
            notes = "Good diversification"
        elif abs_corr < 0.7:
            strength = "Strong"
            color_str = "FF8800"
            notes = "Some overlap"
        else:
            strength = "Very Strong"
            color_str = "CC0000"
            notes = "Highly redundant"
        
        ws.cell(row=row, column=4, value=strength)
        ws.cell(row=row, column=4).font = Font(color=color_str, bold=True)
        ws.cell(row=row, column=5, value=notes)
        
        row += 1
    
    add_border(ws, start_pair_row-1, row-1, 1, 5)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    return ws

def create_summary_sheet(wb):
    """Create executive summary sheet"""
    ws = wb.create_sheet("Executive_Summary", 0)  # Insert at beginning
    
    row = 1
    
    # Main title
    ws.cell(row=row, column=1, value="PORTFOLIO CORRELATION ANALYSIS - EXECUTIVE SUMMARY")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(bold=True, size=18, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 3
    
    # Overview
    ws.cell(row=row, column=1, value="OVERVIEW")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_strategy_header(ws, row, 1, 8, "1F4E79")
    row += 1
    
    ws.cell(row=row, column=1, value="This analysis examines correlations at two levels:")
    row += 1
    ws.cell(row=row, column=1, value="1. Within-Strategy: How correlated are pairs within each strategy?")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="   → Low correlation = better diversification within strategy")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 2
    
    ws.cell(row=row, column=1, value="2. Between-Strategy: How correlated are different strategies to each other?")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="   → Low correlation = better diversification across strategies")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 3
    
    # Key metrics
    ws.cell(row=row, column=1, value="KEY INSIGHTS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_strategy_header(ws, row, 1, 8, "70AD47")
    row += 1
    
    # Calculate overall metrics
    all_strategy_returns = {}
    within_strategy_stats = []
    
    for strategy_name, paths in STRATEGY_PATHS.items():
        returns, pair_names = load_strategy_data(strategy_name, paths)
        
        if returns is None or len(pair_names) < 2:
            continue
        
        # Within-strategy correlation
        corr_matrix = calculate_correlation_matrix(returns)
        mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
        correlations = corr_matrix.where(mask).stack().values
        avg_corr = np.mean(correlations)
        
        within_strategy_stats.append({
            'name': strategy_name,
            'pairs': len(pair_names),
            'avg_corr': avg_corr
        })
        
        # Store combined returns for between-strategy analysis
        combined_return = returns.mean(axis=1)
        all_strategy_returns[strategy_name] = combined_return
    
    # Display within-strategy stats
    ws.cell(row=row, column=1, value="Within-Strategy Correlations:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    headers = ['Strategy', 'Pairs', 'Avg Correlation', 'Diversification']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_header(ws, row, 1, 4)
    row += 1
    
    for stat in within_strategy_stats:
        ws.cell(row=row, column=1, value=stat['name'])
        ws.cell(row=row, column=2, value=stat['pairs'])
        ws.cell(row=row, column=3, value=round(stat['avg_corr'], 4))
        
        diversification = (1 - stat['avg_corr']) * 100
        ws.cell(row=row, column=4, value=f"{round(diversification, 2)}%")
        
        row += 1
    
    row += 2
    
    # Between-strategy correlation
    if len(all_strategy_returns) >= 2:
        strategy_df = pd.DataFrame(all_strategy_returns)
        between_corr_matrix = strategy_df.corr()
        mask = np.triu(np.ones_like(between_corr_matrix, dtype=bool), k=1)
        between_correlations = between_corr_matrix.where(mask).stack().values
        avg_between_corr = np.mean(between_correlations)
        
        ws.cell(row=row, column=1, value="Between-Strategy Correlation:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value="Average Correlation:")
        ws.cell(row=row, column=2, value=round(avg_between_corr, 4))
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        ws.cell(row=row, column=1, value="Portfolio Diversification Benefit:")
        ws.cell(row=row, column=2, value=f"{round((1-avg_between_corr)*100, 2)}%")
        ws.cell(row=row, column=1).font = Font(bold=True, color="006600")
        ws.cell(row=row, column=2).font = Font(bold=True, color="006600", size=12)
        row += 3
    
    # Recommendations
    ws.cell(row=row, column=1, value="RECOMMENDATIONS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_strategy_header(ws, row, 1, 8, "C00000")
    row += 1
    
    recommendations = [
        "✓ Review the 'Within_Strategy_Correlations' sheet to understand pair diversification",
        "✓ Review the 'Between_Strategy_Correlations' sheet for overall portfolio diversification",
        "✓ Lower correlation values indicate better diversification",
        "✓ Use this analysis alongside the capital allocation recommendations",
        "✓ Consider reducing exposure to highly correlated pairs or strategies",
        "✓ Aim for average correlations below 0.5 for good diversification",
    ]
    
    for rec in recommendations:
        ws.cell(row=row, column=1, value=rec)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    print("=" * 80)
    print("COMPREHENSIVE PORTFOLIO CORRELATION ANALYSIS")
    print("=" * 80)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    print("\nCreating correlation analysis sheets...")
    
    print("\n1. Creating Within-Strategy Correlations sheet...")
    create_within_strategy_correlation_sheet(wb)
    
    print("\n2. Creating Between-Strategy Correlations sheet...")
    create_between_strategy_correlation_sheet(wb)
    
    print("\n3. Creating Executive Summary...")
    create_summary_sheet(wb)
    
    # Save workbook
    output_path = '/Users/sureshpatil/Desktop/Portfolio Creation/Portfolio_Correlation_Analysis.xlsx'
    wb.save(output_path)
    
    print("\n" + "=" * 80)
    print(f"✓ Analysis complete! Saved to: {output_path}")
    print("=" * 80)
    print("\nSheets created:")
    print("  1. Executive_Summary - Key insights and recommendations")
    print("  2. Within_Strategy_Correlations - Pair correlations within each strategy")
    print("  3. Between_Strategy_Correlations - Correlations between different strategies")
    print("\nColor Scale Guide:")
    print("  🟢 Green = Positive correlation (0 to +1)")
    print("  ⚪ White = No correlation (0)")
    print("  🔴 Red = Negative correlation (-1 to 0)")

if __name__ == "__main__":
    main()
