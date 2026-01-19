"""
================================================================================
COMPREHENSIVE PORTFOLIO ANALYZER
================================================================================

A unified program that performs complete portfolio analysis:
1. Strategy Statistics with MT5 Sharpe Ratios
2. Pair and Strategy Capital Distribution
3. Final Portfolio Allocation (with user-selected methods)
4. Correlation Analysis (within and between strategies)

OUTPUTS:
- Portfolio_Analysis_Sheets.xlsx (4 sheets with all allocations)
- Portfolio_Correlation_Analysis.xlsx (3 sheets with correlation analysis)

USAGE:
    python portfolio_analyzer.py

TO ADD A NEW STRATEGY:
1. Add the strategy file path in STRATEGY_FILES config
2. Add pair data paths in STRATEGY_PATHS config
3. Add scaling factor in SCALING_FACTORS if tested at higher capital
4. Add pair allocation method in STRATEGY_PAIR_METHODS config

================================================================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os
import re
from datetime import datetime

# ============================================================================
# CONFIGURATION - EDIT THIS SECTION TO ADD NEW STRATEGIES
# ============================================================================

BASE_PATH = '/Users/sureshpatil/Desktop/Portfolio Creation'

# Strategy pair statistics CSV files (add new strategies here)
STRATEGY_FILES = {
    '7th_Strategy': '7th strategy/7th_Strategy_pair_statistics.csv',
    'Falcon': 'Falcon/Falcon_pair_statistics.csv',
    'Gold_Dip': 'Gold Dip/Gold_Dip_pair_statistics.csv',
    'RSI_Correlation': 'RSI corelation/RSI_Correlation_pair_statistics.csv',
    'RSI_6_Trades': 'RSI 6 trades/RSI_6_Trades_pair_statistics.csv',
    'AURUM': 'AURUM/AURUM_pair_statistics.csv',
    'PairTradingEA': 'Pair Trading EA/PairTradingEA_pair_statistics.csv',
    'Reversal_Strategy': 'Reversal Strategy/Reversal_Strategy_pair_statistics.csv',
}

# Scaling factors for strategies tested at higher capital (normalized to $1,000)
SCALING_FACTORS = {
    'Gold_Dip': 10,        # Tested at $10,000 instead of $1,000
    'RSI_Correlation': 100  # Tested at $100,000 instead of $1,000
}

# Sharpe Ratio caps for specific pairs (to avoid unrealistic values)
SHARPE_CAPS = {
    ('AURUM', 'XAUUSD_Grid'): 2.0,
}

# User's selected pair allocation methods for each strategy
STRATEGY_PAIR_METHODS = {
    '7th_Strategy': 'Sharpe_Weighted',
    'Falcon': 'Risk_Parity',
    'Gold_Dip': 'Risk_Parity',
    'RSI_Correlation': 'Inverse_Volatility',
    'RSI_6_Trades': 'Risk_Parity',
    'AURUM': 'Risk_Parity',
    'PairTradingEA': 'Max_Sharpe',
    'Reversal_Strategy': 'Risk_Parity'
}

# Strategy-level allocation method
STRATEGY_ALLOCATION_METHOD = 'Equal_Weight'

# Risk-free rate for Sharpe calculation
RISK_FREE_RATE = 0.0

# Assumed correlation between strategies (typical for forex)
STRATEGY_CORRELATION = 0.3

# Equity curve file paths for correlation analysis
STRATEGY_EQUITY_PATHS = {
    '7th_Strategy': [
        ('7th strategy/XAUUSD 20-25.csv', 'XAUUSD'),
        ('7th strategy/XAGUSD 20-25.csv', 'XAGUSD'),
    ],
    'Falcon': [
        ('Falcon/V5.csv', 'EURUSD'),
        ('Falcon/v5-v2 - Tp 60,SL 60 all day.csv', 'EURUSD_V2'),
    ],
    'Gold_Dip': [
        ('Gold Dip/EURUSD/EURUSD.csv', 'EURUSD'),
        ('Gold Dip/GBPUSD/GBPUSD.csv', 'GBPUSD'),
        ('Gold Dip/AUDUSD/AUDUSD.csv', 'AUDUSD'),
        ('Gold Dip/USDCAD/USDCAD.csv', 'USDCAD'),
        ('Gold Dip/EURJPY/EURJPY.csv', 'EURJPY'),
        ('Gold Dip/AUDJPY/AUDJPY.csv', 'AUDJPY'),
        ('Gold Dip/EURAUD/EURAUD.csv', 'EURAUD'),
        ('Gold Dip/EURCHF/EURCHF.csv', 'EURCHF'),
    ],
    'RSI_6_Trades': [
        ('RSI 6 trades/EURUSD/EURUSD.xlsx', 'EURUSD'),
        ('RSI 6 trades/GBPUSD/GBPUSD.xlsx', 'GBPUSD'),
        ('RSI 6 trades/AUDUSD/AUDUSD.xlsx', 'AUDUSD'),
        ('RSI 6 trades/USDCAD/USDCAD.xlsx', 'USDCAD'),
        ('RSI 6 trades/USDJPY/USDJPY.xlsx', 'USDJPY'),
        ('RSI 6 trades/USDCHF/USDCHF.xlsx', 'USDCHF'),
        ('RSI 6 trades/EURAUD/EURAUD.xlsx', 'EURAUD'),
        ('RSI 6 trades/EURCAD/EURCAD.xlsx', 'EURCAD'),
        ('RSI 6 trades/EURCHF/EURCHF.xlsx', 'EURCHF'),
        ('RSI 6 trades/EURGBP/EURGBP.xlsx', 'EURGBP'),
        ('RSI 6 trades/GBPAUD/GBPAUD.xlsx', 'GBPAUD'),
        ('RSI 6 trades/GBPCAD/GBPCAD.xlsx', 'GBPCAD'),
        ('RSI 6 trades/GBPCHF/GBPCHF.xlsx', 'GBPCHF'),
        ('RSI 6 trades/AUDNZD/AUDNZD.xlsx', 'AUDNZD'),
        ('RSI 6 trades/NZDCHF/NZDCHF.xlsx', 'NZDCHF'),
        ('RSI 6 trades/CADCHF/CADCHF.xlsx', 'CADCHF'),
    ],
    'AURUM': [
        ('AURUM/Gold /Gold - Indivisual TP.xlsx', 'XAUUSD'),
        ('AURUM/USDJPY/USDJPY - AVG TP.xlsx', 'USDJPY'),
    ],
    'PairTradingEA': [
        ('Pair Trading EA/EURUSD-GBPUSD/EURUSD-GBPUSD.xlsx', 'EURUSD_GBPUSD'),
        ('Pair Trading EA/EURUSD_AUDUSD/EURUSD-AUDUSD.xlsx', 'EURUSD_AUDUSD'),
        ('Pair Trading EA/EURGBP-GBPCHF/EURGBP-GBPCHF.xlsx', 'EURGBP_GBPCHF'),
        ('Pair Trading EA/AUDUSD-AUDCAD/AUDUSD-AUDCAD.xlsx', 'AUDUSD_AUDCAD'),
        ('Pair Trading EA/USDCAD_AUDCHF/USDCAD-AUDCHF.xlsx', 'USDCAD_AUDCHF'),
    ],
    'RSI_Correlation': [
        ('RSI corelation/AUDUSD_GBPNZD/AUDUSD_GBPNZD.xlsx', 'AUDUSD_GBPNZD'),
        ('RSI corelation/EURAUD_CADCHF/EURAUD-CADCHF.xlsx', 'EURAUD_CADCHF'),
        ('RSI corelation/EURGBP_GBPCHF/EURGBP_GBPCHF.xlsx', 'EURGBP_GBPCHF'),
        ('RSI corelation/GBPUSD_USDCAD/GBPUSD-USDCAD.xlsx', 'GBPUSD_USDCAD'),
        ('RSI corelation/GBPUSD_USDCHF/GBPUSD-USDCHF.xlsx', 'GBPUSD_USDCHF'),
        ('RSI corelation/USDCAD_AUDCHF/USDCAD-AUDCHF.xlsx', 'USDCAD_AUDCHF'),
    ],
    'Reversal_Strategy': [
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'USDCAD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'AUDNZD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'NZDUSD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'AUDJPY'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'GBPNZD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'USDCHF'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'GBPAUD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'AUDCAD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'GBPJPY'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'EURCAD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'EURAUD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'EURNZD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'EURCHF'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'CADCHF'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'AUDUSD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'GBPUSD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'GBPCAD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'CHFJPY'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'NZDCAD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'EURUSD'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'AUDCHF'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'NZDCHF'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'GBPCHF'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'NZDJPY'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'EURGBP'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'USDJPY'),
        ('Reversal Strategy/All Pairs - 1 Day.xlsx', 'CADJPY'),
    ],
}

# Strategy colors for Excel styling
STRATEGY_COLORS = ["4472C4", "ED7D31", "70AD47", "9E480E", "5B9BD5", "7030A0", "C00000", "FFC000"]

# Strategy display names mapping (internal name -> display name)
STRATEGY_DISPLAY_NAMES = {
    'AURUM': 'Black Dragon',
    'Falcon': 'Silver Falcon',
    'Gold_Dip': 'Iron Bear',
    'PairTradingEA': 'Twin Fox',
    'RSI_6_Trades': 'Red Kraken',
    'RSI_Correlation': 'Night Wolf',
    '7th_Strategy': 'Golden Stallion',
    'Reversal_Strategy': 'Shadow Owl'
}

def get_strategy_display_name(internal_name):
    """Get display name for a strategy, fallback to internal name if not found."""
    return STRATEGY_DISPLAY_NAMES.get(internal_name, internal_name)


# ============================================================================
# UTILITY FUNCTIONS - FILE LOADING
# ============================================================================

def read_html_file(filepath):
    """Read HTML file with various encodings"""
    try:
        with open(filepath, 'rb') as f:
            content = f.read()
        for enc in ['utf-16', 'utf-16-le', 'utf-16-be', 'utf-8', 'latin-1']:
            try:
                decoded = content.decode(enc)
                if '<' in decoded:
                    return decoded
            except:
                continue
    except:
        pass
    return None


def extract_mt5_sharpe(content):
    """Extract Sharpe Ratio directly from MT5 HTML report"""
    match = re.search(r'Sharpe\s*Ratio.*?<b>([+-]?\d+\.?\d*)</b>', content, re.IGNORECASE | re.DOTALL)
    return float(match.group(1)) if match else None


def extract_trades_from_excel(filepath):
    """Extract individual trade profits from Excel file"""
    try:
        df = pd.read_excel(filepath, skiprows=2)
        if 'Profit' in df.columns:
            profits = []
            for val in df['Profit'].dropna():
                try:
                    if isinstance(val, str):
                        val = float(val.replace(' ', '').replace(',', ''))
                    if val != 0:
                        profits.append(val)
                except:
                    pass
            return profits
    except:
        pass
    return []


def extract_trades_from_csv(filepath):
    """Extract individual trade profits from CSV file"""
    try:
        df = pd.read_csv(filepath, skiprows=2, header=None)
        profits = []
        for col in df.columns:
            try:
                vals = df[col].dropna()
                numeric_vals = pd.to_numeric(vals, errors='coerce').dropna()
                if len(numeric_vals) > 10:
                    if numeric_vals.min() < 0 and numeric_vals.max() > 0:
                        profits = numeric_vals.tolist()
                        break
            except:
                pass
        return profits
    except:
        pass
    return []


def extract_trades_from_reversal_excel(filepath):
    """Extract individual trade profits from Reversal Strategy Excel file"""
    try:
        df = pd.read_excel(filepath, skiprows=2)
        if 'Profit' in df.columns and 'Symbol' in df.columns:
            profits = []
            for idx, row in df.iterrows():
                if pd.isna(row['Symbol']):
                    continue
                val = row['Profit']
                if pd.isna(val):
                    continue
                try:
                    if isinstance(val, str):
                        val = float(val.replace(' ', '').replace(',', ''))
                    if val != 0 and abs(val) < 10000:
                        profits.append(val)
                except:
                    pass
            if profits:
                return profits
    except Exception as e:
        print(f"Error extracting Reversal Strategy trades: {e}")
    return []


def load_csv_equity_curve(file_path, pair_name):
    """Load equity curve from CSV file"""
    try:
        df = pd.read_csv(file_path, skiprows=2)
        df.columns = df.columns.str.strip()
        df = df[df['Type'].str.contains('close', case=False, na=False)]
        df = df[['Time', 'Balance']].copy()
        df['Balance'] = df['Balance'].astype(str).str.replace(',', '').str.replace('"', '')
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
        df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
        df = df.dropna()
        df = df.set_index('Time')
        df = df.sort_index()
        df.columns = [pair_name]
        return df
    except Exception as e:
        return None


def load_excel_equity_curve(file_path, pair_name):
    """Load equity curve from Excel file"""
    try:
        df = pd.read_excel(file_path, skiprows=2)
        df.columns = df.columns.str.strip()
        df = df[df['Type'].str.contains('close', case=False, na=False)]
        df = df[['Time', 'Balance']].copy()
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
        df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
        df = df.dropna()
        df = df.set_index('Time')
        df = df.sort_index()
        df.columns = [pair_name]
        return df
    except Exception as e:
        return None


def load_pairtrading_equity_curve(file_path, pair_name):
    """Load equity curve from Pair Trading EA Excel file"""
    try:
        df = pd.read_excel(file_path, skiprows=2)
        df.columns = df.columns.str.strip()
        df = df[df['Balance'].notna()]
        df = df[['Time', 'Balance']].copy()
        df['Balance'] = df['Balance'].astype(str).str.replace(' ', '').str.replace(',', '')
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
        df['Time'] = df['Time'].astype(str).str.strip()
        df['Time'] = pd.to_datetime(df['Time'], format='%Y.%m.%d %H:%M:%S', errors='coerce')
        df = df.dropna()
        df = df[~df.duplicated(subset=['Time'], keep='last')]
        df = df.set_index('Time')
        df = df.sort_index()
        df.columns = [pair_name]
        return df
    except Exception as e:
        return None


def load_reversal_strategy_equity_curve(file_path, pair_name):
    """Load equity curve from Reversal Strategy Excel file"""
    try:
        df = pd.read_excel(file_path, sheet_name='1day time frame 2020-25', skiprows=2)
        df.columns = df.columns.str.strip()
        if 'Symbol' in df.columns:
            df = df[df['Symbol'] == pair_name].copy()
        df = df[df['Balance'].notna() & df['Time'].notna()]
        df = df[['Time', 'Balance']].copy()
        df['Balance'] = df['Balance'].astype(str).str.replace(' ', '').str.replace(',', '')
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce')
        df['Time'] = pd.to_datetime(df['Time'], errors='coerce')
        df = df.dropna()
        df = df[~df.duplicated(subset=['Time'], keep='last')]
        df = df.set_index('Time')
        df = df.sort_index()
        df.columns = [pair_name]
        return df
    except Exception as e:
        return None


# ============================================================================
# SHARPE RATIO CALCULATIONS
# ============================================================================

def calculate_sharpe_from_trades(profits, trading_days=1825):
    """
    Calculate Sharpe Ratio from individual trade profits
    Formula: (Mean Trade Profit / Std Dev of Trade Profits) × √(Trades per Year)
    """
    if len(profits) < 2:
        return 0
    profits = np.array(profits)
    mean_profit = np.mean(profits)
    std_profit = np.std(profits, ddof=1)
    if std_profit == 0:
        return 0
    trades_per_year = len(profits) / (trading_days / 365) if trading_days > 0 else len(profits)
    sharpe = (mean_profit / std_profit) * np.sqrt(trades_per_year)
    return sharpe


def get_mt5_sharpe_for_strategy(strategy_name, pair_name):
    """Get MT5 Sharpe Ratio for a specific strategy and pair"""
    
    if strategy_name == 'PairTradingEA':
        folder_variants = [pair_name.replace('_', '-'), pair_name]
        for folder in folder_variants:
            html_path = os.path.join(BASE_PATH, 'Pair Trading EA', folder, f'{folder}.html')
            if os.path.exists(html_path):
                content = read_html_file(html_path)
                if content:
                    sharpe = extract_mt5_sharpe(content)
                    if sharpe:
                        return sharpe
    
    elif strategy_name == 'RSI_Correlation':
        folder_variants = [pair_name.replace('_', '-'), pair_name]
        for folder in folder_variants:
            html_path = os.path.join(BASE_PATH, 'RSI corelation', folder, f'{folder}.html')
            if os.path.exists(html_path):
                content = read_html_file(html_path)
                if content:
                    sharpe = extract_mt5_sharpe(content)
                    if sharpe:
                        return sharpe
    
    elif strategy_name == 'RSI_6_Trades':
        xlsx_path = os.path.join(BASE_PATH, 'RSI 6 trades', pair_name, f'{pair_name}.xlsx')
        if os.path.exists(xlsx_path):
            profits = extract_trades_from_excel(xlsx_path)
            if profits:
                return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'Gold_Dip':
        csv_path = os.path.join(BASE_PATH, 'Gold Dip', pair_name, f'{pair_name}.csv')
        if os.path.exists(csv_path):
            profits = extract_trades_from_csv(csv_path)
            if profits:
                return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'AURUM':
        folder_map = {'XAUUSD_Grid': 'Gold ', 'USDJPY_Grid': 'USDJPY'}
        if pair_name in folder_map:
            folder = folder_map[pair_name]
            folder_path = os.path.join(BASE_PATH, 'AURUM', folder)
            if os.path.exists(folder_path):
                for file in os.listdir(folder_path):
                    if file.endswith('.xlsx'):
                        xlsx_path = os.path.join(folder_path, file)
                        profits = extract_trades_from_excel(xlsx_path)
                        if profits:
                            return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == '7th_Strategy':
        file_map = {'XAUUSD': 'XAUUSD 20-25.csv', 'XAGUSD': 'XAGUSD 20-25.csv'}
        if pair_name in file_map:
            csv_path = os.path.join(BASE_PATH, '7th strategy', file_map[pair_name])
            if os.path.exists(csv_path):
                profits = extract_trades_from_csv(csv_path)
                if profits:
                    return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'Falcon':
        file_map = {'V5': 'V5.csv', 'v5-v2 - Tp 60,SL 60 all day': 'v5-v2 - Tp 60,SL 60 all day.csv'}
        if pair_name in file_map:
            csv_path = os.path.join(BASE_PATH, 'Falcon', file_map[pair_name])
            if os.path.exists(csv_path):
                profits = extract_trades_from_csv(csv_path)
                if profits:
                    return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'Reversal_Strategy':
        xlsx_path = os.path.join(BASE_PATH, 'Reversal Strategy', 'All Pairs - 1 Day.xlsx')
        if os.path.exists(xlsx_path):
            profits = extract_trades_from_reversal_excel(xlsx_path)
            if profits:
                return calculate_sharpe_from_trades(profits)
    
    return None


# ============================================================================
# ALLOCATION METHODS
# ============================================================================

def calculate_equal_weight(n_assets):
    """Equal weight allocation"""
    return np.ones(n_assets) / n_assets * 100


def calculate_inverse_volatility_weight(max_drawdowns):
    """Inverse volatility weighting (lower DD = higher weight)"""
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    inv_vol = 1 / dd
    return inv_vol / inv_vol.sum() * 100


def calculate_sharpe_weight(sharpe_ratios):
    """Sharpe ratio weighted allocation"""
    sharpe = np.array(sharpe_ratios)
    sharpe = np.maximum(sharpe, 0.001)
    return sharpe / sharpe.sum() * 100


def calculate_risk_parity_weight(sharpe_ratios, max_drawdowns):
    """Risk parity weighting (Sharpe/DD ratio)"""
    sharpe = np.array(sharpe_ratios)
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    risk_adj = sharpe / dd
    risk_adj = np.maximum(risk_adj, 0.001)
    return risk_adj / risk_adj.sum() * 100


def calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns):
    """Max Sharpe optimization weighting"""
    sharpe = np.array(sharpe_ratios)
    returns = np.array(returns)
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    score = sharpe * returns / dd
    score = np.maximum(score, 0.001)
    return score / score.sum() * 100


def get_pair_weights(df, method_name):
    """Get weights for pairs based on allocation method"""
    n_pairs = len(df)
    sharpe_ratios = df['Sharpe_Ratio'].values
    max_drawdowns = df['Max_Drawdown'].values
    returns = df['Correct_Return_Percent'].values
    
    methods = {
        'Equal_Weight': lambda: calculate_equal_weight(n_pairs),
        'Inverse_Volatility': lambda: calculate_inverse_volatility_weight(max_drawdowns),
        'Sharpe_Weighted': lambda: calculate_sharpe_weight(sharpe_ratios),
        'Risk_Parity': lambda: calculate_risk_parity_weight(sharpe_ratios, max_drawdowns),
        'Max_Sharpe': lambda: calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns),
    }
    return methods.get(method_name, methods['Equal_Weight'])()


def estimate_correlation_matrix(n, base_correlation=0.3):
    """Estimate correlation matrix with assumed correlation"""
    corr = np.full((n, n), base_correlation)
    np.fill_diagonal(corr, 1.0)
    return corr


def calculate_portfolio_sharpe(weights, sharpe_ratios, volatilities=None, correlation=0.3):
    """
    Calculate Portfolio Sharpe Ratio using: S_p = (w^T × μ) / sqrt(w^T × Σ × w)
    """
    weights = np.array(weights) / 100
    sharpe = np.array(sharpe_ratios)
    n = len(weights)
    
    if volatilities is None or len(volatilities) != n:
        return np.sum(weights * sharpe)
    
    vol = np.array(volatilities)
    vol = np.where(vol <= 0, 1, vol)
    mu = sharpe * vol
    portfolio_return = np.sum(weights * mu)
    corr_matrix = estimate_correlation_matrix(n, correlation)
    cov_matrix = np.outer(vol, vol) * corr_matrix
    portfolio_variance = weights @ cov_matrix @ weights
    portfolio_std = np.sqrt(portfolio_variance)
    
    if portfolio_std > 0:
        portfolio_sharpe = portfolio_return / portfolio_std
    else:
        portfolio_sharpe = 0
    
    return portfolio_sharpe


def calculate_portfolio_xirr(weights, xirr_values):
    """Calculate weighted portfolio XIRR"""
    weights = np.array(weights) / 100
    xirr = np.array(xirr_values)
    return np.sum(weights * xirr)


# ============================================================================
# EXCEL STYLING UTILITIES
# ============================================================================

def style_header(ws, row, start_col, end_col):
    """Apply header styling"""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def style_subheader(ws, row, start_col, end_col):
    """Apply subheader styling"""
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def style_strategy_header(ws, row, start_col, end_col, color="4472C4"):
    """Apply strategy section header styling"""
    strategy_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    strategy_font = Font(bold=True, color="FFFFFF", size=12)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = strategy_fill
        cell.font = strategy_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def style_result_row(ws, row, start_col, end_col, color="E2EFDA"):
    """Apply result row styling"""
    result_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    result_font = Font(bold=True, size=10)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = result_fill
        cell.font = result_font


def add_border(ws, start_row, end_row, start_col, end_col):
    """Add border to a range"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def apply_correlation_color_scale(ws, start_row, end_row, start_col, end_col):
    """Apply color scale to correlation values"""
    rule = ColorScaleRule(
        start_type='num', start_value=-1, start_color='FF6B6B',
        mid_type='num', mid_value=0, mid_color='FFFFFF',
        end_type='num', end_value=1, end_color='4ECB71'
    )
    range_string = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    ws.conditional_formatting.add(range_string, rule)


# ============================================================================
# DATA LOADING AND PROCESSING
# ============================================================================

def recalculate_metrics(df, strategy_name=None):
    """Recalculate all metrics based on correct Initial Balance = Max Drawdown × 2"""
    df = df.copy()
    
    scale = SCALING_FACTORS.get(strategy_name, 1)
    if scale > 1:
        df['Total_Profit'] = df['Total_Profit'] / scale
        df['Max_Drawdown'] = df['Max_Drawdown'] / scale
        if 'Initial_Balance' in df.columns:
            df['Initial_Balance'] = df['Initial_Balance'] / scale
        if 'Final_Balance' in df.columns:
            df['Final_Balance'] = df['Final_Balance'] / scale
    
    # Update Sharpe Ratio to MT5 standard for each pair
    for idx, row in df.iterrows():
        pair_name = row['Currency_Pair']
        mt5_sharpe = get_mt5_sharpe_for_strategy(strategy_name, pair_name)
        if mt5_sharpe is not None:
            cap_key = (strategy_name, pair_name)
            if cap_key in SHARPE_CAPS:
                mt5_sharpe = min(mt5_sharpe, SHARPE_CAPS[cap_key])
            df.at[idx, 'Sharpe_Ratio'] = round(mt5_sharpe, 2)
    
    df['Correct_Initial_Balance'] = df['Max_Drawdown'] * 2
    df['Correct_Final_Balance'] = df['Correct_Initial_Balance'] + df['Total_Profit']
    df['Correct_Return_Percent'] = (df['Total_Profit'] / df['Correct_Initial_Balance']) * 100
    df['Trading_Years'] = df['Trading_Period_Days'] / 365
    df['Correct_XIRR'] = df.apply(
        lambda row: ((row['Correct_Final_Balance'] / row['Correct_Initial_Balance']) ** 
                    (1 / max(row['Trading_Years'], 0.1)) - 1) * 100 
        if row['Correct_Initial_Balance'] > 0 else 0, 
        axis=1
    )
    df['Scale_Factor'] = scale
    return df


def load_all_strategies():
    """Load all strategy data from CSV files"""
    strategies_data = {}
    
    print("=" * 80)
    print("LOADING STRATEGY DATA")
    print("=" * 80)
    
    for strategy_name, file_path in STRATEGY_FILES.items():
        full_path = os.path.join(BASE_PATH, file_path)
        if os.path.exists(full_path):
            df = pd.read_csv(full_path)
            df = recalculate_metrics(df, strategy_name)
            strategies_data[strategy_name] = df
            print(f"  ✓ {strategy_name}: Loaded {len(df)} pairs")
        else:
            print(f"  ✗ {strategy_name}: File not found - {file_path}")
    
    return strategies_data


def load_strategy_equity_data(strategy_name, paths):
    """Load equity curve data for a strategy"""
    equity_curves = []
    pair_names = []
    
    for rel_path, pair_name in paths:
        file_path = os.path.join(BASE_PATH, rel_path)
        if not os.path.exists(file_path):
            continue
        
        if strategy_name == 'Reversal_Strategy':
            equity_df = load_reversal_strategy_equity_curve(file_path, pair_name)
        elif strategy_name in ['PairTradingEA', 'RSI_Correlation']:
            equity_df = load_pairtrading_equity_curve(file_path, pair_name)
        elif file_path.endswith('.xlsx'):
            equity_df = load_excel_equity_curve(file_path, pair_name)
        else:
            equity_df = load_csv_equity_curve(file_path, pair_name)
        
        if equity_df is not None and len(equity_df) > 0:
            equity_curves.append(equity_df)
            pair_names.append(pair_name)
    
    if not equity_curves:
        return None, None
    
    returns_list = []
    for equity_df in equity_curves:
        equity_df = equity_df[~equity_df.index.duplicated(keep='last')]
        daily_equity = equity_df.resample('D').last().ffill()
        ret = daily_equity.pct_change().fillna(0)
        returns_list.append(ret)
    
    returns = pd.concat(returns_list, axis=1, join='outer')
    returns = returns.fillna(0)
    
    return returns, pair_names


# ============================================================================
# SHEET CREATION FUNCTIONS - PORTFOLIO ANALYSIS
# ============================================================================

def create_sheet1_statistics(wb, strategies_data):
    """Create Sheet 1: Strategy Statistics with ALPHA column"""
    ws = wb.create_sheet("Strategy_Statistics")
    row = 1
    
    # Title
    ws.cell(row=row, column=1, value="COMPREHENSIVE STRATEGY STATISTICS (MT5 Sharpe Ratios)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value="Note: Initial Capital = Max Drawdown × 2 | ALPHA = Sharpe × √(Total Trades)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="C00000")
    row += 2
    
    for idx, (strategy_name, df) in enumerate(strategies_data.items()):
        display_name = get_strategy_display_name(strategy_name)
        ws.cell(row=row, column=1, value=f"Strategy: {display_name}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
        style_strategy_header(ws, row, 1, 18, STRATEGY_COLORS[idx % len(STRATEGY_COLORS)])
        row += 1
        
        # Added ALPHA column (18th column)
        headers = ['Currency_Pair', 'Sharpe_Ratio', 'Total_Trades', 'Win_Rate_%', 
                   'Total_Profit', 'Max_Drawdown', 'Initial_Capital', 'Final_Balance',
                   'Return_%', 'XIRR_%', 'Profit_Factor', 'Trading_Days', 
                   'Trading_Years', 'Start_Date', 'End_Date', 'Winning', 'Losing', 'ALPHA']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_idx, value=header)
        style_subheader(ws, row, 1, len(headers))
        row += 1
        
        start_data_row = row
        
        for _, pair_row in df.iterrows():
            ws.cell(row=row, column=1, value=pair_row['Currency_Pair'])
            ws.cell(row=row, column=2, value=round(pair_row['Sharpe_Ratio'], 2))
            ws.cell(row=row, column=3, value=int(pair_row['Total_Trades']))
            ws.cell(row=row, column=4, value=f"=ROUND(IF(C{row}>0, (P{row}/C{row})*100, 0), 2)")
            ws.cell(row=row, column=5, value=round(pair_row['Total_Profit'], 2))
            ws.cell(row=row, column=6, value=round(pair_row['Max_Drawdown'], 2))
            ws.cell(row=row, column=7, value=f"=ROUND(F{row}*2, 2)")
            ws.cell(row=row, column=8, value=f"=ROUND(G{row}+E{row}, 2)")
            ws.cell(row=row, column=9, value=f"=ROUND(IF(G{row}>0, (E{row}/G{row})*100, 0), 2)")
            ws.cell(row=row, column=10, value=f"=ROUND(IF(AND(G{row}>0, M{row}>0), ((1+I{row}/100)^(1/M{row})-1)*100, 0), 2)")
            ws.cell(row=row, column=11, value=round(pair_row['Profit_Factor'], 2))
            ws.cell(row=row, column=12, value=int(pair_row.get('Trading_Period_Days', 0)))
            ws.cell(row=row, column=13, value=f"=ROUND(L{row}/365, 2)")
            ws.cell(row=row, column=14, value=str(pair_row.get('Start_Date', 'N/A')))
            ws.cell(row=row, column=15, value=str(pair_row.get('End_Date', 'N/A')))
            ws.cell(row=row, column=16, value=int(pair_row['Winning_Trades']))
            ws.cell(row=row, column=17, value=int(pair_row['Losing_Trades']))
            # ALPHA = Sharpe × √(Total Trades)
            ws.cell(row=row, column=18, value=f"=ROUND(B{row}*SQRT(C{row}), 2)")
            row += 1
        
        end_data_row = row - 1
        
        # STRATEGY TOTAL row
        ws.cell(row=row, column=1, value="STRATEGY TOTAL")
        ws.cell(row=row, column=2, value=f"=ROUND(AVERAGE(B{start_data_row}:B{end_data_row}), 2)")
        ws.cell(row=row, column=3, value=f"=SUM(C{start_data_row}:C{end_data_row})")
        ws.cell(row=row, column=4, value=f"=ROUND(IF(C{row}>0, (P{row}/C{row})*100, 0), 2)")
        ws.cell(row=row, column=5, value=f"=ROUND(SUM(E{start_data_row}:E{end_data_row}), 2)")
        ws.cell(row=row, column=6, value=f"=ROUND(SUM(F{start_data_row}:F{end_data_row}), 2)")
        ws.cell(row=row, column=7, value=f"=ROUND(SUM(G{start_data_row}:G{end_data_row}), 2)")
        ws.cell(row=row, column=8, value=f"=ROUND(SUM(H{start_data_row}:H{end_data_row}), 2)")
        ws.cell(row=row, column=9, value=f"=ROUND(IF(G{row}>0, (E{row}/G{row})*100, 0), 2)")
        ws.cell(row=row, column=10, value=f"=ROUND(AVERAGE(J{start_data_row}:J{end_data_row}), 2)")
        ws.cell(row=row, column=11, value=f"=ROUND(AVERAGE(K{start_data_row}:K{end_data_row}), 2)")
        ws.cell(row=row, column=16, value=f"=SUM(P{start_data_row}:P{end_data_row})")
        ws.cell(row=row, column=17, value=f"=SUM(Q{start_data_row}:Q{end_data_row})")
        # Strategy ALPHA = Avg Sharpe × √(Total Trades)
        ws.cell(row=row, column=18, value=f"=ROUND(B{row}*SQRT(C{row}), 2)")
        style_result_row(ws, row, 1, 18, "FFF2CC")
        
        add_border(ws, start_data_row - 1, row, 1, 18)
        row += 3
    
    # Column widths
    for col in range(1, 19):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 28
    
    return ws


def create_sheet2_pair_allocation(wb, strategies_data):
    """Create Sheet 2: Pair Capital Distribution with proper borders and 2 decimals"""
    ws = wb.create_sheet("Pair_Capital_Distribution")
    row = 1
    
    ws.cell(row=row, column=1, value="PAIR CAPITAL DISTRIBUTION WITHIN STRATEGIES")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    for idx, (strategy_name, df) in enumerate(strategies_data.items()):
        if len(df) == 0:
            continue
        
        display_name = get_strategy_display_name(strategy_name)
        ws.cell(row=row, column=1, value=f"STRATEGY: {display_name} ({len(df)} pairs)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        style_strategy_header(ws, row, 1, 12, STRATEGY_COLORS[idx % len(STRATEGY_COLORS)])
        row += 1
        
        n_pairs = len(df)
        sharpe_ratios = df['Sharpe_Ratio'].values
        max_drawdowns = df['Max_Drawdown'].values
        returns = df['Correct_Return_Percent'].values
        
        weights = {
            'Equal_Weight': calculate_equal_weight(n_pairs),
            'Inverse_Volatility': calculate_inverse_volatility_weight(max_drawdowns),
            'Sharpe_Weighted': calculate_sharpe_weight(sharpe_ratios),
            'Risk_Parity': calculate_risk_parity_weight(sharpe_ratios, max_drawdowns),
            'Max_Sharpe': calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns)
        }
        
        headers = ['Currency_Pair', 'Equal_%', 'Inv_Vol_%', 'Sharpe_%', 
                   'Risk_Parity_%', 'Max_Sharpe_%', 'Sharpe_Ratio', 'Return_%', 'XIRR_%', 
                   'Max_DD', 'Initial_Cap', 'Profit']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_idx, value=header)
        style_subheader(ws, row, 1, len(headers))
        row += 1
        
        start_data_row = row
        
        for i, pair_name in enumerate(df['Currency_Pair'].values):
            ws.cell(row=row, column=1, value=pair_name)
            ws.cell(row=row, column=2, value=f"=ROUND(100/{n_pairs}, 2)")
            ws.cell(row=row, column=3, value=round(weights['Inverse_Volatility'][i], 2))
            ws.cell(row=row, column=4, value=round(weights['Sharpe_Weighted'][i], 2))
            ws.cell(row=row, column=5, value=round(weights['Risk_Parity'][i], 2))
            ws.cell(row=row, column=6, value=round(weights['Max_Sharpe'][i], 2))
            ws.cell(row=row, column=7, value=round(df.iloc[i]['Sharpe_Ratio'], 2))
            ws.cell(row=row, column=8, value=f"=ROUND(IF(K{row}>0, (L{row}/K{row})*100, 0), 2)")
            trading_years = df.iloc[i]['Trading_Years']
            ws.cell(row=row, column=9, value=f"=ROUND(IF(H{row}>0, ((1+H{row}/100)^(1/{trading_years:.4f})-1)*100, 0), 2)")
            ws.cell(row=row, column=10, value=round(df.iloc[i]['Max_Drawdown'], 2))
            ws.cell(row=row, column=11, value=f"=ROUND(J{row}*2, 2)")
            ws.cell(row=row, column=12, value=round(df.iloc[i]['Total_Profit'], 2))
            row += 1
        
        end_data_row = row - 1
        
        # TOTAL row
        ws.cell(row=row, column=1, value="TOTAL")
        for c in range(2, 7):
            ws.cell(row=row, column=c, value=f"=ROUND(SUM({get_column_letter(c)}{start_data_row}:{get_column_letter(c)}{end_data_row}), 2)")
        ws.cell(row=row, column=7, value=f"=ROUND(AVERAGE(G{start_data_row}:G{end_data_row}), 2)")
        ws.cell(row=row, column=10, value=f"=ROUND(SUM(J{start_data_row}:J{end_data_row}), 2)")
        ws.cell(row=row, column=11, value=f"=ROUND(SUM(K{start_data_row}:K{end_data_row}), 2)")
        ws.cell(row=row, column=12, value=f"=ROUND(SUM(L{start_data_row}:L{end_data_row}), 2)")
        style_result_row(ws, row, 1, 12, "D9E1F2")
        
        add_border(ws, start_data_row - 1, row, 1, 12)
        row += 3
    
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 28
    
    return ws


def create_sheet3_strategy_allocation(wb, strategies_data):
    """Create Sheet 3: Strategy Capital Distribution with portfolio metrics per allocation method"""
    ws = wb.create_sheet("Strategy_Capital_Distribution")
    row = 1
    
    ws.cell(row=row, column=1, value="STRATEGY CAPITAL DISTRIBUTION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    strategies = []
    for strategy_name, df in strategies_data.items():
        display_name = get_strategy_display_name(strategy_name)
        strategy_info = {
            'Strategy': display_name,
            'Num_Pairs': len(df),
            'Avg_Sharpe_Ratio': df['Sharpe_Ratio'].mean(),
            'Total_Profit': df['Total_Profit'].sum(),
            'Total_Initial_Capital': df['Correct_Initial_Balance'].sum(),
            'Max_Drawdown': df['Max_Drawdown'].sum(),
            'Correct_XIRR': df['Correct_XIRR'].mean(),
            'Trading_Years': df['Trading_Years'].mean(),
        }
        strategies.append(strategy_info)
    
    strategy_df = pd.DataFrame(strategies)
    n_strategies = len(strategy_df)
    
    sharpe_ratios = strategy_df['Avg_Sharpe_Ratio'].values
    max_drawdowns = strategy_df['Max_Drawdown'].values
    returns = (strategy_df['Total_Profit'] / strategy_df['Total_Initial_Capital'] * 100).values
    xirr_values = strategy_df['Correct_XIRR'].values
    total_capital = strategy_df['Total_Initial_Capital'].sum()
    total_profit = strategy_df['Total_Profit'].sum()
    
    weights = {
        'Equal_Weight': calculate_equal_weight(n_strategies),
        'Inverse_Volatility': calculate_inverse_volatility_weight(max_drawdowns),
        'Sharpe_Weighted': calculate_sharpe_weight(sharpe_ratios),
        'Risk_Parity': calculate_risk_parity_weight(sharpe_ratios, max_drawdowns),
        'Max_Sharpe': calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns)
    }
    
    ws.cell(row=row, column=1, value="STRATEGY WEIGHTS BY ALLOCATION METHOD")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    style_strategy_header(ws, row, 1, 12, "1F4E79")
    row += 1
    
    headers = ['Strategy', 'Pairs', 'Sharpe', 'Return_%', 'XIRR_%', 'Capital_Req',
               'Equal_%', 'Inv_Vol_%', 'Sharpe_%', 'Risk_Parity_%', 'Max_Sharpe_%', 'Profit']
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(headers))
    row += 1
    
    start_data_row = row
    
    for i, strat_row in strategy_df.iterrows():
        ws.cell(row=row, column=1, value=strat_row['Strategy'])
        ws.cell(row=row, column=2, value=int(strat_row['Num_Pairs']))
        ws.cell(row=row, column=3, value=round(strat_row['Avg_Sharpe_Ratio'], 2))
        ws.cell(row=row, column=4, value=f"=ROUND(IF(F{row}>0, (L{row}/F{row})*100, 0), 2)")
        trading_yrs = strat_row['Trading_Years']
        ws.cell(row=row, column=5, value=f"=ROUND(IF(D{row}>0, ((1+D{row}/100)^(1/{trading_yrs:.4f})-1)*100, 0), 2)")
        ws.cell(row=row, column=6, value=round(strat_row['Total_Initial_Capital'], 2))
        ws.cell(row=row, column=7, value=f"=ROUND(100/{n_strategies}, 2)")
        ws.cell(row=row, column=8, value=round(weights['Inverse_Volatility'][i], 2))
        ws.cell(row=row, column=9, value=round(weights['Sharpe_Weighted'][i], 2))
        ws.cell(row=row, column=10, value=round(weights['Risk_Parity'][i], 2))
        ws.cell(row=row, column=11, value=round(weights['Max_Sharpe'][i], 2))
        ws.cell(row=row, column=12, value=round(strat_row['Total_Profit'], 2))
        row += 1
    
    end_data_row = row - 1
    
    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=f"=SUM(B{start_data_row}:B{end_data_row})")
    ws.cell(row=row, column=6, value=f"=ROUND(SUM(F{start_data_row}:F{end_data_row}), 2)")
    for c in range(7, 12):
        ws.cell(row=row, column=c, value=f"=ROUND(SUM({get_column_letter(c)}{start_data_row}:{get_column_letter(c)}{end_data_row}), 2)")
    ws.cell(row=row, column=12, value=f"=ROUND(SUM(L{start_data_row}:L{end_data_row}), 2)")
    style_result_row(ws, row, 1, 12, "D9E1F2")
    total_row = row
    
    add_border(ws, start_data_row - 1, row, 1, 12)
    row += 3
    
    # =========================================================================
    # SECTION 2: FINAL PORTFOLIO METRICS BY ALLOCATION METHOD
    # =========================================================================
    ws.cell(row=row, column=1, value="FINAL PORTFOLIO METRICS BY ALLOCATION METHOD")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    style_strategy_header(ws, row, 1, 7, "70AD47")
    row += 1
    
    ws.cell(row=row, column=1, value="Use this table to compare what your final portfolio would look like with each allocation method")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    row += 1
    
    # Headers for portfolio metrics table
    perf_headers = ['Allocation_Method', 'Portfolio_Sharpe', 'Portfolio_XIRR_%', 
                    'Total_Capital_$', 'Expected_Profit_$', 'Overall_Return_%', 'Rating']
    for col_idx, header in enumerate(perf_headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(perf_headers))
    row += 1
    
    perf_start_row = row
    rho = STRATEGY_CORRELATION
    
    # Calculate portfolio metrics for each allocation method
    method_names = ['Equal_Weight', 'Inverse_Volatility', 'Sharpe_Weighted', 'Risk_Parity', 'Max_Sharpe']
    weight_cols = {'Equal_Weight': 'G', 'Inverse_Volatility': 'H', 'Sharpe_Weighted': 'I', 
                   'Risk_Parity': 'J', 'Max_Sharpe': 'K'}
    
    for method_name in method_names:
        w_col = weight_cols[method_name]
        
        ws.cell(row=row, column=1, value=method_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Portfolio Sharpe using weighted average formula with covariance
        # S_p = SUMPRODUCT(w, Sharpe*Vol) / SQRT((1-ρ)*SUMPRODUCT(w², Vol²) + ρ*(SUMPRODUCT(w, Vol))²)
        # Vol proxy = Capital/2 = Max_DD, so we use F column (Capital) / 2
        # But simpler: weighted average of Sharpe ratios adjusted for diversification
        # For Excel formula: use weighted average Sharpe
        numerator = f"SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, C{start_data_row}:C{end_data_row})"
        ws.cell(row=row, column=2, value=f"=ROUND({numerator}, 2)")
        
        # Portfolio XIRR = weighted average
        ws.cell(row=row, column=3, value=f"=ROUND(SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, E{start_data_row}:E{end_data_row}), 2)")
        
        # Total Capital = always the same (sum of all strategy capital requirements)
        ws.cell(row=row, column=4, value=f"=ROUND(F{total_row}, 2)")
        
        # Expected Profit = Total_Capital × Weighted_Average_Return%
        # Where Weighted_Average_Return = SUMPRODUCT(weight/100, Return_%)
        # Column D = Return_% for each strategy = (Profit/Capital)*100
        # Formula: Total_Capital × Weighted_Return / 100
        weighted_return_formula = f"SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, D{start_data_row}:D{end_data_row})"
        ws.cell(row=row, column=5, value=f"=ROUND(F{total_row} * {weighted_return_formula} / 100, 2)")
        
        # Overall Return = Weighted Average Return (same as what we used for Expected Profit)
        ws.cell(row=row, column=6, value=f"=ROUND({weighted_return_formula}, 2)")
        
        # Rating based on Portfolio Sharpe
        ws.cell(row=row, column=7, value=f'=IF(B{row}>=1.5,"★★★★★",IF(B{row}>=1,"★★★★",IF(B{row}>=0.5,"★★★",IF(B{row}>=0.25,"★★","★"))))')
        
        row += 1
    
    perf_end_row = row - 1
    add_border(ws, perf_start_row - 1, perf_end_row, 1, 7)
    row += 2
    
    # Best method recommendation
    ws.cell(row=row, column=1, value="RECOMMENDATION:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="C00000")
    ws.cell(row=row, column=2, value="Select the allocation method with highest Portfolio Sharpe for risk-adjusted returns")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    
    # Column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = 22
    
    return ws


def create_sheet4_final_portfolio(wb, strategies_data):
    """Create Sheet 4: Final Portfolio Analysis with correct Sharpe formula and borders"""
    ws = wb.create_sheet("Final_Portfolio_Analysis", 0)
    row = 1
    
    ws.cell(row=row, column=1, value="FINAL PORTFOLIO ANALYSIS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(bold=True, size=18, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # ALLOCATION CONFIGURATION section with border
    config_start_row = row
    ws.cell(row=row, column=1, value="ALLOCATION CONFIGURATION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    style_strategy_header(ws, row, 1, 3, "70AD47")
    row += 1
    
    ws.cell(row=row, column=1, value="Strategy Level Allocation:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=STRATEGY_ALLOCATION_METHOD)
    ws.cell(row=row, column=2).font = Font(size=11, color="0066CC")
    row += 1
    
    ws.cell(row=row, column=1, value="Pair Level Methods:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    
    for strategy, method in STRATEGY_PAIR_METHODS.items():
        display_name = get_strategy_display_name(strategy)
        ws.cell(row=row, column=1, value=f"  • {display_name}:")
        ws.cell(row=row, column=2, value=method)
        ws.cell(row=row, column=2).font = Font(color="0066CC")
        row += 1
    
    add_border(ws, config_start_row, row - 1, 1, 3)
    row += 2
    
    # Calculate strategy results
    strategy_results = []
    for strategy_name, df in strategies_data.items():
        if len(df) == 0:
            continue
        
        pair_method = STRATEGY_PAIR_METHODS.get(strategy_name, 'Equal_Weight')
        pair_weights = get_pair_weights(df, pair_method)
        
        total_capital = df['Correct_Initial_Balance'].sum()
        total_profit = df['Total_Profit'].sum()
        max_dd = df['Max_Drawdown'].sum()
        avg_trading_years = df['Trading_Years'].mean()
        
        # Calculate weighted average Sharpe (MT5 standard)
        # Portfolio Sharpe = weighted average of individual Sharpe ratios
        strategy_sharpe = np.sum((pair_weights / 100) * df['Sharpe_Ratio'].values)
        strategy_xirr = calculate_portfolio_xirr(pair_weights, df['Correct_XIRR'].values)
        
        display_name = get_strategy_display_name(strategy_name)
        strategy_results.append({
            'Strategy': display_name,
            'Pair_Method': pair_method,
            'Num_Pairs': len(df),
            'Strategy_Sharpe': strategy_sharpe,
            'Strategy_XIRR': strategy_xirr,
            'Total_Capital': total_capital,
            'Total_Profit': total_profit,
            'Max_Drawdown': max_dd,
            'Avg_Trading_Years': avg_trading_years
        })
    
    strategy_df = pd.DataFrame(strategy_results)
    n_strategies = len(strategy_df)
    
    # SECTION 1: Strategy-Level Performance
    ws.cell(row=row, column=1, value="SECTION 1: STRATEGY-LEVEL PERFORMANCE")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    style_strategy_header(ws, row, 1, 11, "4472C4")
    row += 1
    
    headers = ['Strategy', 'Pair_Method', 'Pairs', 'Sharpe', 'XIRR_%', 'Simple_Return_%', 'Return_%', 
               'Capital', 'Profit', 'Max_DD', 'Years']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(headers))
    row += 1
    
    start_data_row = row
    
    for idx, strat in strategy_df.iterrows():
        ws.cell(row=row, column=1, value=strat['Strategy'])
        ws.cell(row=row, column=2, value=strat['Pair_Method'])
        ws.cell(row=row, column=3, value=int(strat['Num_Pairs']))
        ws.cell(row=row, column=4, value=round(strat['Strategy_Sharpe'], 2))
        ws.cell(row=row, column=5, value=round(strat['Strategy_XIRR'], 2))
        # Simple Return % = (Profit / Capital / Years) * 100
        ws.cell(row=row, column=6, value=f"=ROUND(IF(H{row}>0, (I{row}/H{row}/K{row})*100, 0), 2)")
        # Return % = (Profit / Capital) * 100 (total return, not annualized)
        ws.cell(row=row, column=7, value=f"=ROUND(IF(H{row}>0, (I{row}/H{row})*100, 0), 2)")
        ws.cell(row=row, column=8, value=round(strat['Total_Capital'], 2))
        ws.cell(row=row, column=9, value=round(strat['Total_Profit'], 2))
        ws.cell(row=row, column=10, value=round(strat['Max_Drawdown'], 2))
        ws.cell(row=row, column=11, value=round(strat['Avg_Trading_Years'], 2))
        row += 1
    
    end_data_row = row - 1
    
    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=f"=SUM(C{start_data_row}:C{end_data_row})")
    ws.cell(row=row, column=8, value=f"=ROUND(SUM(H{start_data_row}:H{end_data_row}), 2)")
    ws.cell(row=row, column=9, value=f"=ROUND(SUM(I{start_data_row}:I{end_data_row}), 2)")
    ws.cell(row=row, column=10, value=f"=ROUND(SUM(J{start_data_row}:J{end_data_row}), 2)")
    style_result_row(ws, row, 1, 11, "FFF2CC")
    total_row_section1 = row
    
    add_border(ws, start_data_row - 1, row, 1, 11)
    row += 3
    
    # SECTION 2: Final Portfolio Allocation
    ws.cell(row=row, column=1, value="SECTION 2: FINAL PORTFOLIO ALLOCATION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_strategy_header(ws, row, 1, 8, "ED7D31")
    row += 1
    
    headers = ['Strategy', 'Weight_%', 'Allocated_Capital', 'Min_Required', 
               'Expected_Return_%', 'Expected_Profit', 'Sharpe', 'XIRR_%']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(headers))
    row += 1
    
    start_alloc_row = row
    
    for idx, strat in strategy_df.iterrows():
        data_row_ref = start_data_row + idx
        ws.cell(row=row, column=1, value=strat['Strategy'])
        ws.cell(row=row, column=2, value=f"=ROUND(100/{n_strategies}, 2)")
        ws.cell(row=row, column=3, value=f"=ROUND((B{row}/100)*$H${total_row_section1}, 2)")
        ws.cell(row=row, column=4, value=f"=ROUND(H{data_row_ref}, 2)")
        # Expected_Return_% uses Simple Return % (annualized) from Section 1, column 6
        ws.cell(row=row, column=5, value=f"=ROUND(F{data_row_ref}, 2)")
        ws.cell(row=row, column=6, value=f"=ROUND(C{row}*(E{row}/100), 2)")
        ws.cell(row=row, column=7, value=f"=ROUND(D{data_row_ref}, 2)")
        ws.cell(row=row, column=8, value=f"=ROUND(E{data_row_ref}, 2)")
        row += 1
    
    end_alloc_row = row - 1
    
    # TOTAL row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=f"=ROUND(SUM(B{start_alloc_row}:B{end_alloc_row}), 2)")
    ws.cell(row=row, column=3, value=f"=ROUND(SUM(C{start_alloc_row}:C{end_alloc_row}), 2)")
    ws.cell(row=row, column=4, value=f"=ROUND(SUM(D{start_alloc_row}:D{end_alloc_row}), 2)")
    ws.cell(row=row, column=6, value=f"=ROUND(SUM(F{start_alloc_row}:F{end_alloc_row}), 2)")
    style_result_row(ws, row, 1, 8, "FFF2CC")
    total_alloc_row = row
    
    add_border(ws, start_alloc_row - 1, row, 1, 8)
    row += 3
    
    # SECTION 3: Final Portfolio Performance
    ws.cell(row=row, column=1, value="SECTION 3: FINAL PORTFOLIO PERFORMANCE")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    style_strategy_header(ws, row, 1, 3, "70AD47")
    row += 1
    
    perf_start_row = row
    
    # Total Trading Pairs
    ws.cell(row=row, column=1, value="Total Trading Pairs")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=C{total_row_section1}")
    row += 1
    
    # Total Capital Required
    ws.cell(row=row, column=1, value="Total Capital Required (MDD×2)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=ROUND(H{total_row_section1}, 2)")
    ws.cell(row=row, column=3, value="$")
    row += 1
    
    # Total Maximum Drawdown
    ws.cell(row=row, column=1, value="Total Maximum Drawdown")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=ROUND(J{total_row_section1}, 2)")
    ws.cell(row=row, column=3, value="$")
    row += 1
    
    # Portfolio Sharpe Ratio - Using CORRECT weighted average formula
    # Portfolio Sharpe = Σ(weight_i × Sharpe_i) for MT5 standard
    ws.cell(row=row, column=1, value="Portfolio Sharpe Ratio")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    sharpe_formula = f"=ROUND(SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, G{start_alloc_row}:G{end_alloc_row}), 2)"
    ws.cell(row=row, column=2, value=sharpe_formula)
    ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="006600")
    ws.cell(row=row, column=3, value="★")
    ws.cell(row=row, column=3).font = Font(size=14, color="FFD700")
    row += 1
    
    # Portfolio XIRR
    ws.cell(row=row, column=1, value="Portfolio XIRR")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=ROUND(SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, H{start_alloc_row}:H{end_alloc_row}), 2)")
    ws.cell(row=row, column=3, value="%")
    row += 1
    
    # Portfolio Simple Return %
    ws.cell(row=row, column=1, value="Portfolio Simple Return %")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    # Weighted average of Simple Return % from Section 1
    ws.cell(row=row, column=2, value=f"=ROUND(SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, F{start_data_row}:F{end_data_row}), 2)")
    ws.cell(row=row, column=3, value="%")
    row += 1
    
    # Expected Total Profit
    ws.cell(row=row, column=1, value="Expected Total Profit")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=ROUND(I{total_row_section1}, 2)")
    ws.cell(row=row, column=3, value="$")
    row += 1
    
    # Overall Return on Capital
    ws.cell(row=row, column=1, value="Overall Return on Capital")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=ROUND(IF(H{total_row_section1}>0, (I{total_row_section1}/H{total_row_section1})*100, 0), 2)")
    ws.cell(row=row, column=3, value="%")
    
    add_border(ws, perf_start_row, row, 1, 3)
    row += 2
    
    # Formula explanation
    ws.cell(row=row, column=1, value="SHARPE RATIO FORMULA:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=10, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value="Portfolio Sharpe = Σ(Weight_i × Sharpe_i)")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    row += 1
    ws.cell(row=row, column=1, value="Where Sharpe_i = (Mean Profit / Std Dev) × √(Trades per Year)")
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    for col in range(4, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    return ws


# ============================================================================
# SHEET CREATION FUNCTIONS - CORRELATION ANALYSIS
# ============================================================================

def create_within_strategy_correlation_sheet(wb):
    """Create sheet showing correlation within each strategy"""
    ws = wb.create_sheet("Within_Strategy_Correlations")
    row = 1
    
    ws.cell(row=row, column=1, value="CORRELATION ANALYSIS: PAIRS WITHIN EACH STRATEGY")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    for idx, (strategy_name, paths) in enumerate(STRATEGY_EQUITY_PATHS.items()):
        print(f"  Loading {strategy_name}...")
        returns, pair_names = load_strategy_equity_data(strategy_name, paths)
        
        if returns is None or len(pair_names) < 2:
            continue
        
        corr_matrix = returns.corr()
        
        display_name = get_strategy_display_name(strategy_name)
        ws.cell(row=row, column=1, value=f"STRATEGY: {display_name}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(pair_names)+1)
        style_strategy_header(ws, row, 1, len(pair_names)+1, STRATEGY_COLORS[idx % len(STRATEGY_COLORS)])
        row += 1
        
        ws.cell(row=row, column=1, value="Pair")
        for col_idx, pair in enumerate(pair_names, 2):
            ws.cell(row=row, column=col_idx, value=pair)
        style_header(ws, row, 1, len(pair_names)+1)
        row += 1
        
        start_data_row = row
        
        for i, pair in enumerate(pair_names):
            ws.cell(row=row, column=1, value=pair)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for j, _ in enumerate(pair_names):
                corr_val = corr_matrix.iloc[i, j]
                cell = ws.cell(row=row, column=j+2, value=round(corr_val, 4))
                cell.alignment = Alignment(horizontal='center')
                if i == j:
                    cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            row += 1
        
        apply_correlation_color_scale(ws, start_data_row, row-1, 2, len(pair_names)+1)
        add_border(ws, start_data_row-1, row-1, 1, len(pair_names)+1)
        
        mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
        correlations = corr_matrix.where(mask).stack().values
        
        if len(correlations) > 0:
            row += 1
            ws.cell(row=row, column=1, value="Average Correlation:")
            ws.cell(row=row, column=2, value=round(np.mean(correlations), 4))
            row += 1
            ws.cell(row=row, column=1, value="Diversification Benefit:")
            ws.cell(row=row, column=2, value=f"{round((1 - np.mean(correlations)) * 100, 2)}%")
            ws.cell(row=row, column=1).font = Font(bold=True, color="006600")
        
        row += 3
    
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 25
    
    return ws


def create_between_strategy_correlation_sheet(wb):
    """Create sheet showing correlation between strategies"""
    ws = wb.create_sheet("Between_Strategy_Correlations")
    row = 1
    
    ws.cell(row=row, column=1, value="CORRELATION ANALYSIS: BETWEEN STRATEGIES")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    strategy_returns = {}
    strategy_names = []
    
    for strategy_name, paths in STRATEGY_EQUITY_PATHS.items():
        returns, pair_names = load_strategy_equity_data(strategy_name, paths)
        if returns is None:
            continue
        combined_return = returns.mean(axis=1)
        strategy_returns[strategy_name] = combined_return
        strategy_names.append(strategy_name)
    
    if len(strategy_names) < 2:
        ws.cell(row=row, column=1, value="Error: Insufficient strategy data")
        return ws
    
    strategy_df = pd.DataFrame(strategy_returns)
    corr_matrix = strategy_df.corr()
    
    ws.cell(row=row, column=1, value="STRATEGY CORRELATION MATRIX")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(strategy_names)+1)
    style_strategy_header(ws, row, 1, len(strategy_names)+1, "1F4E79")
    row += 1
    
    ws.cell(row=row, column=1, value="Strategy")
    for col_idx, strategy in enumerate(strategy_names, 2):
        display_name = get_strategy_display_name(strategy)
        ws.cell(row=row, column=col_idx, value=display_name)
    style_header(ws, row, 1, len(strategy_names)+1)
    row += 1
    
    start_data_row = row
    
    for i, strategy in enumerate(strategy_names):
        display_name = get_strategy_display_name(strategy)
        ws.cell(row=row, column=1, value=display_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for j, _ in enumerate(strategy_names):
            corr_val = corr_matrix.iloc[i, j]
            cell = ws.cell(row=row, column=j+2, value=round(corr_val, 4))
            cell.alignment = Alignment(horizontal='center')
            if i == j:
                cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        row += 1
    
    apply_correlation_color_scale(ws, start_data_row, row-1, 2, len(strategy_names)+1)
    add_border(ws, start_data_row-1, row-1, 1, len(strategy_names)+1)
    
    row += 2
    
    mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=1)
    correlations = corr_matrix.where(mask).stack().values
    avg_corr = np.mean(correlations)
    
    ws.cell(row=row, column=1, value="PORTFOLIO DIVERSIFICATION ANALYSIS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_strategy_header(ws, row, 1, 6, "70AD47")
    row += 1
    
    ws.cell(row=row, column=1, value="Average Strategy Correlation:")
    ws.cell(row=row, column=2, value=round(avg_corr, 4))
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    ws.cell(row=row, column=1, value="Diversification Benefit:")
    ws.cell(row=row, column=2, value=f"{round((1-avg_corr)*100, 2)}%")
    ws.cell(row=row, column=1).font = Font(bold=True, color="006600")
    ws.cell(row=row, column=2).font = Font(bold=True, color="006600", size=12)
    row += 2
    
    if avg_corr < 0.3:
        interpretation = "Excellent - Strategies are weakly correlated"
    elif avg_corr < 0.5:
        interpretation = "Good - Moderate correlation"
    elif avg_corr < 0.7:
        interpretation = "Fair - Strategies are somewhat correlated"
    else:
        interpretation = "Poor - Strategies are highly correlated"
    
    ws.cell(row=row, column=1, value="Interpretation:")
    ws.cell(row=row, column=2, value=interpretation)
    
    ws.column_dimensions['A'].width = 30
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return ws


def create_correlation_summary_sheet(wb):
    """Create executive summary sheet for correlation analysis"""
    ws = wb.create_sheet("Executive_Summary", 0)
    row = 1
    
    ws.cell(row=row, column=1, value="PORTFOLIO CORRELATION ANALYSIS - EXECUTIVE SUMMARY")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(bold=True, size=18, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 3
    
    ws.cell(row=row, column=1, value="OVERVIEW")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_strategy_header(ws, row, 1, 8, "1F4E79")
    row += 1
    
    overview = [
        "This analysis examines correlations at two levels:",
        "",
        "1. Within-Strategy: How correlated are pairs within each strategy?",
        "   → Low correlation = better diversification within strategy",
        "",
        "2. Between-Strategy: How correlated are different strategies?",
        "   → Low correlation = better diversification across strategies",
    ]
    
    for text in overview:
        ws.cell(row=row, column=1, value=text)
        row += 1
    
    row += 2
    
    ws.cell(row=row, column=1, value="RECOMMENDATIONS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    style_strategy_header(ws, row, 1, 8, "C00000")
    row += 1
    
    recommendations = [
        "✓ Review the 'Within_Strategy_Correlations' sheet for pair diversification",
        "✓ Review the 'Between_Strategy_Correlations' sheet for strategy diversification",
        "✓ Lower correlation values indicate better diversification",
        "✓ Aim for average correlations below 0.5 for good diversification",
    ]
    
    for rec in recommendations:
        ws.cell(row=row, column=1, value=rec)
        row += 1
    
    ws.column_dimensions['A'].width = 60
    
    return ws


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def create_portfolio_analysis_workbook(strategies_data):
    """Create the main Portfolio Analysis workbook"""
    print("\n" + "=" * 80)
    print("CREATING PORTFOLIO ANALYSIS SHEETS")
    print("=" * 80)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    print("  Creating Sheet 1: Strategy Statistics...")
    create_sheet1_statistics(wb, strategies_data)
    
    print("  Creating Sheet 2: Pair Capital Distribution...")
    create_sheet2_pair_allocation(wb, strategies_data)
    
    print("  Creating Sheet 3: Strategy Capital Distribution...")
    create_sheet3_strategy_allocation(wb, strategies_data)
    
    print("  Creating Sheet 4: Final Portfolio Analysis...")
    create_sheet4_final_portfolio(wb, strategies_data)
    
    output_path = os.path.join(BASE_PATH, 'Portfolio_Analysis_Sheets.xlsx')
    wb.save(output_path)
    print(f"\n  ✓ Saved: {output_path}")
    
    return output_path


def create_correlation_analysis_workbook():
    """Create the Correlation Analysis workbook"""
    print("\n" + "=" * 80)
    print("CREATING CORRELATION ANALYSIS SHEETS")
    print("=" * 80)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    print("  Creating Within-Strategy Correlations sheet...")
    create_within_strategy_correlation_sheet(wb)
    
    print("\n  Creating Between-Strategy Correlations sheet...")
    create_between_strategy_correlation_sheet(wb)
    
    print("  Creating Executive Summary...")
    create_correlation_summary_sheet(wb)
    
    output_path = os.path.join(BASE_PATH, 'Portfolio_Correlation_Analysis.xlsx')
    wb.save(output_path)
    print(f"\n  ✓ Saved: {output_path}")
    
    return output_path


def main():
    """Main entry point - runs all analyses"""
    print("\n" + "=" * 80)
    print("COMPREHENSIVE PORTFOLIO ANALYZER")
    print("=" * 80)
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Base Path: {BASE_PATH}")
    print("=" * 80)
    
    # Load all strategy data
    strategies_data = load_all_strategies()
    
    if not strategies_data:
        print("\nERROR: No strategy data loaded. Check file paths.")
        return
    
    print(f"\nLoaded {len(strategies_data)} strategies successfully.")
    
    # Create Portfolio Analysis workbook
    portfolio_path = create_portfolio_analysis_workbook(strategies_data)
    
    # Create Correlation Analysis workbook
    correlation_path = create_correlation_analysis_workbook()
    
    # Final summary
    print("\n" + "=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)
    print("\nOutput Files:")
    print(f"  1. {portfolio_path}")
    print("     - Final_Portfolio_Analysis (your selected allocations)")
    print("     - Strategy_Statistics (all strategies with MT5 Sharpe)")
    print("     - Pair_Capital_Distribution (5 allocation methods)")
    print("     - Strategy_Capital_Distribution (5 allocation methods)")
    print(f"\n  2. {correlation_path}")
    print("     - Executive_Summary (key insights)")
    print("     - Within_Strategy_Correlations (pair correlations)")
    print("     - Between_Strategy_Correlations (strategy correlations)")
    print("\n" + "=" * 80)
    
    # Print portfolio summary
    total_capital = sum(df['Correct_Initial_Balance'].sum() for df in strategies_data.values())
    total_profit = sum(df['Total_Profit'].sum() for df in strategies_data.values())
    total_pairs = sum(len(df) for df in strategies_data.values())
    
    print("\nPORTFOLIO SUMMARY:")
    print(f"  Total Strategies: {len(strategies_data)}")
    print(f"  Total Trading Pairs: {total_pairs}")
    print(f"  Total Capital Required: ${total_capital:,.2f}")
    print(f"  Total Expected Profit: ${total_profit:,.2f}")
    print(f"  Overall Return: {(total_profit/total_capital)*100:.2f}%")
    print("=" * 80)


if __name__ == "__main__":
    main()
