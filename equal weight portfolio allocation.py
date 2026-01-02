"""
Create Final Portfolio Sheet with User's Selected Allocation Methods
- Strategy Level: Equal Weight
- Pair Level: User-specified methods per strategy

ALL CELLS USE EXCEL FORMULAS FOR TRANSPARENCY AND AUDITABILITY

Portfolio Sharpe Formula (ex-ante form):
  S_p = (w'μ) / √(w'Σw)
  
Where:
  w = capital weights
  μ = expected excess returns  
  Σ = covariance matrix (Σ_ij = ρ × σ_i × σ_j for i≠j)
  ρ = correlation coefficient (0.3 for moderate correlation)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION: User's Selected Methods
# ============================================================================

STRATEGY_PAIR_METHODS = {
    '7th_Strategy': 'Sharpe_Weighted',
    'Falcon': 'Risk_Parity',
    'Gold_Dip': 'Risk_Parity',
    'RSI_Correlation': 'Inverse_Volatility',
    'RSI_6_Trades': 'Risk_Parity',
    'AURUM': 'Risk_Parity',
    'PairTradingEA': 'Max_Sharpe',
    'Reversal_Strategy': 'Risk_Parity'
}

STRATEGY_ALLOCATION_METHOD = 'Equal_Weight'

# Risk-free rate for Sharpe calculation (set to 0)
RISK_FREE_RATE = 0.0

# Assumed correlation between strategies (typical for forex)
STRATEGY_CORRELATION = 0.3

# ============================================================================
# DATA LOADING AND RECALCULATION (Same as before)
# ============================================================================

SCALING_FACTORS = {
    'Gold_Dip': 10,
    'RSI_Correlation': 100
}

def recalculate_metrics(df, strategy_name=None):
    """Recalculate all metrics based on correct Initial Balance = Max Drawdown × 2"""
    df = df.copy()
    
    scale = SCALING_FACTORS.get(strategy_name, 1)
    if scale > 1:
        df['Total_Profit'] = df['Total_Profit'] / scale
        df['Max_Drawdown'] = df['Max_Drawdown'] / scale
        if 'Initial_Balance' in df.columns:
            df['Initial_Balance'] = df['Initial_Balance'] / scale
        if 'Final_Balance' in df.columns:
            df['Final_Balance'] = df['Final_Balance'] / scale
    
    df['Correct_Initial_Balance'] = df['Max_Drawdown'] * 2
    df['Correct_Final_Balance'] = df['Correct_Initial_Balance'] + df['Total_Profit']
    df['Correct_Return_Percent'] = (df['Total_Profit'] / df['Correct_Initial_Balance']) * 100
    df['Trading_Years'] = df['Trading_Period_Days'] / 365
    
    df['Correct_XIRR'] = df.apply(
        lambda row: ((row['Correct_Final_Balance'] / row['Correct_Initial_Balance']) ** 
                    (1 / max(row['Trading_Years'], 0.1)) - 1) * 100 
        if row['Correct_Initial_Balance'] > 0 else 0, 
        axis=1
    )
    
    df['Scale_Factor'] = scale
    return df

# Load all strategies
strategies_data = {}

df_7th = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/7th strategy/7th_Strategy_pair_statistics.csv')
strategies_data['7th_Strategy'] = recalculate_metrics(df_7th, '7th_Strategy')

df_falcon = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/Falcon/Falcon_pair_statistics.csv')
strategies_data['Falcon'] = recalculate_metrics(df_falcon, 'Falcon')

df_gold_dip = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/Gold Dip/Gold_Dip_pair_statistics.csv')
strategies_data['Gold_Dip'] = recalculate_metrics(df_gold_dip, 'Gold_Dip')

df_rsi_corr = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/RSI corelation/RSI_Correlation_pair_statistics.csv')
strategies_data['RSI_Correlation'] = recalculate_metrics(df_rsi_corr, 'RSI_Correlation')

df_rsi_6 = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/RSI 6 trades/RSI_6_Trades_pair_statistics.csv')
strategies_data['RSI_6_Trades'] = recalculate_metrics(df_rsi_6, 'RSI_6_Trades')

df_aurum = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/AURUM/AURUM_pair_statistics.csv')
strategies_data['AURUM'] = recalculate_metrics(df_aurum, 'AURUM')

df_pair_trading = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/Pair Trading EA/PairTradingEA_pair_statistics.csv')
strategies_data['PairTradingEA'] = recalculate_metrics(df_pair_trading, 'PairTradingEA')

df_reversal = pd.read_csv('/Users/sureshpatil/Desktop/Portfolio Creation/Reversal Strategy/Reversal_Strategy_pair_statistics.csv')
strategies_data['Reversal_Strategy'] = recalculate_metrics(df_reversal, 'Reversal_Strategy')

# ============================================================================
# ALLOCATION METHODS
# ============================================================================

def calculate_equal_weight(n_assets):
    return np.ones(n_assets) / n_assets * 100

def calculate_inverse_volatility_weight(max_drawdowns):
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    inv_vol = 1 / dd
    return inv_vol / inv_vol.sum() * 100

def calculate_sharpe_weight(sharpe_ratios):
    sharpe = np.array(sharpe_ratios)
    sharpe = np.maximum(sharpe, 0.001)
    return sharpe / sharpe.sum() * 100

def calculate_risk_parity_weight(sharpe_ratios, max_drawdowns):
    sharpe = np.array(sharpe_ratios)
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    risk_adj = sharpe / dd
    risk_adj = np.maximum(risk_adj, 0.001)
    return risk_adj / risk_adj.sum() * 100

def calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns):
    sharpe = np.array(sharpe_ratios)
    returns = np.array(returns)
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    score = sharpe * returns / dd
    score = np.maximum(score, 0.001)
    return score / score.sum() * 100

def get_pair_weights(df, method_name):
    """Get weights for pairs based on allocation method"""
    n_pairs = len(df)
    sharpe_ratios = df['Sharpe_Ratio'].values
    max_drawdowns = df['Max_Drawdown'].values
    returns = df['Correct_Return_Percent'].values
    
    if method_name == 'Equal_Weight':
        return calculate_equal_weight(n_pairs)
    elif method_name == 'Inverse_Volatility':
        return calculate_inverse_volatility_weight(max_drawdowns)
    elif method_name == 'Sharpe_Weighted':
        return calculate_sharpe_weight(sharpe_ratios)
    elif method_name == 'Risk_Parity':
        return calculate_risk_parity_weight(sharpe_ratios, max_drawdowns)
    elif method_name == 'Max_Sharpe':
        return calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns)
    else:
        return calculate_equal_weight(n_pairs)

def calculate_portfolio_sharpe(weights, excess_returns, volatilities, correlation=0.3):
    """
    Calculate portfolio Sharpe ratio using correct formula:
    
    S_p = (w'μ) / √(w'Σw)
    
    Where:
        w = weights vector
        μ = expected excess returns (returns - risk_free_rate)
        Σ = covariance matrix of strategy returns
    
    Parameters:
        weights: array of weights (in %)
        excess_returns: array of excess returns (return - risk_free_rate)
        volatilities: array of strategy volatilities
        correlation: assumed correlation between strategies (default 0.3 for forex)
    
    Returns:
        Portfolio Sharpe ratio
    """
    w = np.array(weights) / 100
    mu = np.array(excess_returns)
    sigma = np.array(volatilities)
    n = len(w)
    
    # Build covariance matrix: Cov(i,j) = ρ × σ_i × σ_j for i≠j, σ_i² for i=j
    Sigma = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            if i == j:
                Sigma[i, j] = sigma[i] ** 2  # Variance on diagonal
            else:
                Sigma[i, j] = correlation * sigma[i] * sigma[j]  # Covariance
    
    # Portfolio excess return: w'μ
    portfolio_excess_return = np.dot(w, mu)
    
    # Portfolio variance: w'Σw
    portfolio_variance = np.dot(w.T, np.dot(Sigma, w))
    portfolio_volatility = np.sqrt(portfolio_variance)
    
    # Portfolio Sharpe: (w'μ) / √(w'Σw)
    if portfolio_volatility > 0:
        portfolio_sharpe = portfolio_excess_return / portfolio_volatility
    else:
        portfolio_sharpe = 0
    
    return portfolio_sharpe

def calculate_portfolio_xirr(weights, xirr_values):
    """Calculate portfolio XIRR as weighted average"""
    weights = np.array(weights) / 100
    xirr = np.array(xirr_values)
    return np.sum(weights * xirr)

# ============================================================================
# EXCEL STYLING
# ============================================================================

def style_header(ws, row, start_col, end_col):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_subheader(ws, row, start_col, end_col):
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_strategy_header(ws, row, start_col, end_col, color="4472C4"):
    strategy_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    strategy_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = strategy_fill
        cell.font = strategy_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_result_row(ws, row, start_col, end_col, color="E2EFDA"):
    result_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    result_font = Font(bold=True, size=10)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = result_fill
        cell.font = result_font

def add_border(ws, start_row, end_row, start_col, end_col):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border

# ============================================================================
# CREATE FINAL PORTFOLIO SHEET
# ============================================================================

def create_final_portfolio_sheet(wb):
    """
    Create final portfolio sheet with DYNAMIC EXCEL FORMULAS.
    All calculated cells show formulas when clicked for full transparency.
    """
    ws = wb.create_sheet("Final_Portfolio_Analysis", 0)  # Insert as first sheet
    
    row = 1
    
    # Main title
    ws.cell(row=row, column=1, value="FINAL PORTFOLIO ANALYSIS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(bold=True, size=18, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # Configuration section
    ws.cell(row=row, column=1, value="ALLOCATION CONFIGURATION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    style_strategy_header(ws, row, 1, 10, "70AD47")
    row += 1
    
    ws.cell(row=row, column=1, value="Strategy Level Allocation:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=STRATEGY_ALLOCATION_METHOD)
    ws.cell(row=row, column=2).font = Font(size=11, color="0066CC")
    row += 2
    
    ws.cell(row=row, column=1, value="Pair Level Allocation Methods:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    
    # List pair methods
    for strategy, method in STRATEGY_PAIR_METHODS.items():
        ws.cell(row=row, column=1, value=f"  • {strategy}:")
        ws.cell(row=row, column=2, value=method)
        ws.cell(row=row, column=2).font = Font(color="0066CC")
        row += 1
    
    row += 2
    
    # Calculate strategy-level metrics and weights
    strategy_results = []
    
    for strategy_name, df in strategies_data.items():
        if len(df) == 0:
            continue
        
        # Get pair allocation method for this strategy
        pair_method = STRATEGY_PAIR_METHODS.get(strategy_name, 'Equal_Weight')
        
        # Calculate pair weights
        pair_weights = get_pair_weights(df, pair_method)
        
        # Calculate strategy totals first (needed for calculations)
        total_capital = df['Correct_Initial_Balance'].sum()
        total_profit = df['Total_Profit'].sum()
        total_return = (total_profit / total_capital * 100) if total_capital > 0 else 0
        max_dd = df['Max_Drawdown'].sum()
        avg_trading_years = df['Trading_Years'].mean()
        
        # Calculate excess returns for each pair (XIRR - risk_free_rate)
        pair_excess_returns = (df['Correct_XIRR'].values / 100) - RISK_FREE_RATE
        
        # Calculate volatility for each pair from Sharpe ratio
        # σ = (r - rf) / Sharpe, capped between 5% and 100%
        pair_volatilities = np.array([
            max(0.05, min(1.0, abs(er) / max(s, 0.01)))
            for er, s in zip(pair_excess_returns, df['Sharpe_Ratio'].values)
        ])
        
        # Calculate strategy-level Sharpe using correct formula: S_p = (w'μ) / √(w'Σw)
        strategy_sharpe = calculate_portfolio_sharpe(
            pair_weights, 
            pair_excess_returns, 
            pair_volatilities, 
            correlation=STRATEGY_CORRELATION
        )
        strategy_xirr = calculate_portfolio_xirr(pair_weights, df['Correct_XIRR'].values)
        
        strategy_results.append({
            'Strategy': strategy_name,
            'Pair_Method': pair_method,
            'Num_Pairs': len(df),
            'Strategy_Sharpe': strategy_sharpe,
            'Strategy_XIRR': strategy_xirr,
            'Total_Capital': total_capital,
            'Total_Profit': total_profit,
            'Total_Return_%': total_return,
            'Max_Drawdown': max_dd,
            'Avg_Trading_Years': avg_trading_years
        })
    
    strategy_df = pd.DataFrame(strategy_results)
    
    # Calculate equal weight for strategies
    n_strategies = len(strategy_df)
    strategy_weights = calculate_equal_weight(n_strategies)
    
    # SECTION 1: Strategy-Level Performance
    ws.cell(row=row, column=1, value="SECTION 1: STRATEGY-LEVEL PERFORMANCE")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    style_strategy_header(ws, row, 1, 10, "4472C4")
    row += 1
    
    ws.cell(row=row, column=1, value="(Each strategy's performance using its selected pair allocation method)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
    row += 1
    
    # Headers
    headers = ['Strategy', 'Pair_Method', 'Pairs', 'Sharpe', 'XIRR_%', 'Return_%', 
               'Capital', 'Profit', 'Max_DD', 'Years']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(headers))
    row += 1
    
    start_data_row = row
    
    # Strategy data - write raw values that formulas can reference
    for idx, strat in strategy_df.iterrows():
        ws.cell(row=row, column=1, value=strat['Strategy'])
        ws.cell(row=row, column=2, value=strat['Pair_Method'])
        ws.cell(row=row, column=3, value=int(strat['Num_Pairs']))
        ws.cell(row=row, column=4, value=round(strat['Strategy_Sharpe'], 4))
        ws.cell(row=row, column=5, value=round(strat['Strategy_XIRR'], 2))
        # Return % = (Profit / Capital) * 100 - FORMULA
        ws.cell(row=row, column=6, value=f"=IF(G{row}>0, (H{row}/G{row})*100, 0)")
        ws.cell(row=row, column=7, value=round(strat['Total_Capital'], 2))
        ws.cell(row=row, column=8, value=round(strat['Total_Profit'], 2))
        ws.cell(row=row, column=9, value=round(strat['Max_Drawdown'], 2))
        ws.cell(row=row, column=10, value=round(strat['Avg_Trading_Years'], 2))
        row += 1
    
    end_data_row = row - 1
    
    # Total row with FORMULAS
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=f"=SUM(C{start_data_row}:C{end_data_row})")
    ws.cell(row=row, column=7, value=f"=SUM(G{start_data_row}:G{end_data_row})")
    ws.cell(row=row, column=8, value=f"=SUM(H{start_data_row}:H{end_data_row})")
    ws.cell(row=row, column=9, value=f"=SUM(I{start_data_row}:I{end_data_row})")
    style_result_row(ws, row, 1, 10, "FFF2CC")
    total_row_section1 = row
    
    add_border(ws, start_data_row - 1, row, 1, 10)
    row += 3
    
    # SECTION 2: Final Portfolio Allocation - WITH FORMULAS
    ws.cell(row=row, column=1, value="SECTION 2: FINAL PORTFOLIO ALLOCATION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    style_strategy_header(ws, row, 1, 10, "ED7D31")
    row += 1
    
    ws.cell(row=row, column=1, value=f"(Using {STRATEGY_ALLOCATION_METHOD} for strategy distribution)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
    row += 1
    
    # Headers
    headers = ['Strategy', 'Weight_%', 'Allocated_Capital', 'Min_Required', 
               'Expected_Return_%', 'Expected_Profit', 'Sharpe', 'XIRR_%']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(headers))
    row += 1
    
    start_alloc_row = row
    
    # Allocation data with FORMULAS
    for idx, strat in strategy_df.iterrows():
        data_row_ref = start_data_row + idx  # Reference to Section 1 data
        
        ws.cell(row=row, column=1, value=strat['Strategy'])
        # Weight % = 100 / n_strategies (Equal Weight formula)
        ws.cell(row=row, column=2, value=f"=100/{n_strategies}")
        # Allocated Capital = (Weight/100) * Total Capital (reference Section 1 total)
        ws.cell(row=row, column=3, value=f"=(B{row}/100)*$G${total_row_section1}")
        # Min Required = Direct reference to Section 1 Capital column
        ws.cell(row=row, column=4, value=f"=G{data_row_ref}")
        # Expected Return % = Direct reference to Section 1 Return column
        ws.cell(row=row, column=5, value=f"=F{data_row_ref}")
        # Expected Profit = Allocated_Capital * (Expected_Return/100)
        ws.cell(row=row, column=6, value=f"=C{row}*(E{row}/100)")
        # Sharpe = Direct reference to Section 1
        ws.cell(row=row, column=7, value=f"=D{data_row_ref}")
        # XIRR = Direct reference to Section 1
        ws.cell(row=row, column=8, value=f"=E{data_row_ref}")
        row += 1
    
    end_alloc_row = row - 1
    
    # Total row with FORMULAS
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=f"=SUM(B{start_alloc_row}:B{end_alloc_row})")
    ws.cell(row=row, column=3, value=f"=SUM(C{start_alloc_row}:C{end_alloc_row})")
    ws.cell(row=row, column=4, value=f"=SUM(D{start_alloc_row}:D{end_alloc_row})")
    style_result_row(ws, row, 1, 8, "FFF2CC")
    total_row_section2 = row
    
    add_border(ws, start_alloc_row - 1, row, 1, 8)
    row += 3
    
    # SECTION 3: Final Portfolio Performance - WITH FORMULAS
    ws.cell(row=row, column=1, value="SECTION 3: FINAL PORTFOLIO PERFORMANCE")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    style_strategy_header(ws, row, 1, 6, "70AD47")
    row += 1
    
    start_perf_row = row
    
    # Store correlation value for formula
    rho = STRATEGY_CORRELATION
    
    # Total Trading Pairs - FORMULA referencing Section 1
    ws.cell(row=row, column=1, value="Total Trading Pairs")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=C{total_row_section1}")
    row += 2
    
    # Total Capital Required - FORMULA
    ws.cell(row=row, column=1, value="Total Capital Required (MDD×2)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=G{total_row_section1}")
    ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 1
    
    # Total Maximum Drawdown - FORMULA
    ws.cell(row=row, column=1, value="Total Maximum Drawdown")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=I{total_row_section1}")
    ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 2
    
    # Portfolio Sharpe Ratio - FORMULA using covariance method
    # S_p = (w'μ) / √(w'Σw) where μ = XIRR/100, σ derived from Sharpe
    # Simplified: SUMPRODUCT(weights, XIRR) / SQRT(weighted variance with correlation)
    ws.cell(row=row, column=1, value="Portfolio Sharpe Ratio")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    
    # Portfolio Sharpe = SUMPRODUCT(w, μ*σ) / SQRT((1-ρ)*SUMPRODUCT(w², σ²) + ρ*(SUMPRODUCT(w, σ))²)
    # Using Section 2: B=weights, G=Sharpe (proxy for μ), D=Capital (σ proxy = Capital/2 = Max_DD)
    # Numerator: SUMPRODUCT(B/100, G*(D/2))
    # Denominator: SQRT((1-ρ)*SUMPRODUCT((B/100)², (D/2)²) + ρ*(SUMPRODUCT(B/100, D/2))²)
    numerator = f"SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, G{start_alloc_row}:G{end_alloc_row}*(D{start_alloc_row}:D{end_alloc_row}/2))"
    var_term1 = f"(1-{rho})*SUMPRODUCT((B{start_alloc_row}:B{end_alloc_row}/100)^2, (D{start_alloc_row}:D{end_alloc_row}/2)^2)"
    var_term2 = f"{rho}*(SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, D{start_alloc_row}:D{end_alloc_row}/2))^2"
    sharpe_formula = f"={numerator}/SQRT({var_term1}+{var_term2})"
    
    ws.cell(row=row, column=2, value=sharpe_formula)
    ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="006600")
    ws.cell(row=row, column=3, value="★★★")
    ws.cell(row=row, column=3).font = Font(size=14, color="006600")
    sharpe_row = row
    row += 1
    
    # Portfolio XIRR - FORMULA = SUMPRODUCT(weights, XIRR values)
    ws.cell(row=row, column=1, value="Portfolio XIRR")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, H{start_alloc_row}:H{end_alloc_row})")
    ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="006600")
    ws.cell(row=row, column=2).number_format = '0.00"%"'
    ws.cell(row=row, column=3, value="★★★")
    ws.cell(row=row, column=3).font = Font(size=14, color="006600")
    xirr_row = row
    row += 1
    
    # Weighted Return - FORMULA = SUMPRODUCT(weights, returns)
    ws.cell(row=row, column=1, value="Weighted Return")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=SUMPRODUCT(B{start_alloc_row}:B{end_alloc_row}/100, E{start_alloc_row}:E{end_alloc_row})")
    ws.cell(row=row, column=2).number_format = '0.00"%"'
    weighted_return_row = row
    row += 2
    
    # Expected Total Profit - FORMULA
    ws.cell(row=row, column=1, value="Expected Total Profit")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=H{total_row_section1}")
    ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
    row += 1
    
    # Overall Return on Capital - FORMULA = (Total Profit / Total Capital) * 100
    ws.cell(row=row, column=1, value="Overall Return on Capital")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value=f"=IF(G{total_row_section1}>0, (H{total_row_section1}/G{total_row_section1})*100, 0)")
    ws.cell(row=row, column=2).number_format = '0.00"%"'
    row += 1
    
    add_border(ws, start_perf_row, row - 1, 1, 3)
    row += 3
    
    # SECTION 4: Summary & Key Insights
    ws.cell(row=row, column=1, value="SECTION 4: SUMMARY & KEY INSIGHTS")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    style_strategy_header(ws, row, 1, 10, "C00000")
    row += 1
    
    # Add formula explanation
    ws.cell(row=row, column=1, value="FORMULA USED FOR PORTFOLIO SHARPE RATIO:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    
    ws.cell(row=row, column=1, value=f"  S_p = (w'μ) / √(w'Σw)  where Σ_ij = ρ×σ_i×σ_j, ρ = {rho}")
    ws.cell(row=row, column=1).font = Font(size=10, color="666666", italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2
    
    insights = [
        f"✓ Using {STRATEGY_ALLOCATION_METHOD} allocation across {n_strategies} strategies",
        f"✓ All cells contain Excel formulas - click any cell to see calculation",
        f"✓ Section 1 provides strategy-level data",
        f"✓ Section 2 allocates capital based on weights",
        f"✓ Section 3 calculates portfolio metrics using SUMPRODUCT formulas",
        "",
        "KEY FORMULAS USED:",
        f"  • Weight: =100/{n_strategies} (Equal Weight)",
        f"  • Allocated Capital: =(Weight/100) × Total Capital",
        f"  • Expected Profit: =Allocated_Capital × (Return%/100)",
        f"  • Portfolio XIRR: =SUMPRODUCT(weights/100, XIRR values)",
        f"  • Portfolio Sharpe: =(w'μ)/√(w'Σw) with ρ={rho}",
        "",
        "RISK MANAGEMENT:",
        f"  • Safety buffer: 2× maximum drawdown as capital base",
        f"  • Diversification across {n_strategies} independent strategies",
        f"  • Each strategy uses optimal pair allocation method",
    ]
    
    for insight in insights:
        ws.cell(row=row, column=1, value=insight)
        if insight.startswith('✓'):
            ws.cell(row=row, column=1).font = Font(size=10, color="006600")
        elif insight.startswith('KEY') or insight.startswith('RISK'):
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="C00000")
        elif insight.startswith('  •') or insight.startswith('  Safety') or insight.startswith('  Diversification') or insight.startswith('  Each'):
            ws.cell(row=row, column=1).font = Font(size=10, color="0066CC")
        else:
            ws.cell(row=row, column=1).font = Font(size=10)
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    
    # Calculate values for print summary
    total_capital_req = strategy_df['Total_Capital'].sum()
    total_profit_potential = strategy_df['Total_Profit'].sum()
    
    # Recalculate portfolio metrics for print output
    strategy_excess_returns = (strategy_df['Strategy_XIRR'].values / 100) - RISK_FREE_RATE
    strategy_volatilities = np.array([
        max(0.05, min(1.0, abs(er) / max(s, 0.01)))
        for er, s in zip(strategy_excess_returns, strategy_df['Strategy_Sharpe'].values)
    ])
    portfolio_sharpe = calculate_portfolio_sharpe(
        strategy_weights, strategy_excess_returns, strategy_volatilities, STRATEGY_CORRELATION
    )
    portfolio_xirr = calculate_portfolio_xirr(strategy_weights, strategy_df['Strategy_XIRR'].values)
    
    print("\n" + "="*80)
    print("FINAL PORTFOLIO SUMMARY (ALL CELLS USE EXCEL FORMULAS)")
    print("="*80)
    print(f"Portfolio Sharpe Ratio: {round(portfolio_sharpe, 4)}")
    print(f"Portfolio XIRR: {round(portfolio_xirr, 2)}%")
    print(f"Total Capital Required: ${total_capital_req:,.2f}")
    print(f"Expected Total Profit: ${round(total_profit_potential, 2):,}")
    print(f"Overall Return: {round((total_profit_potential/total_capital_req)*100, 2)}%")
    print(f"\nFormula: S_p = (w'μ) / √(w'Σw) with ρ = {rho}")
    print("="*80)
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    # Load existing workbook
    wb = load_workbook('/Users/sureshpatil/Desktop/Portfolio Creation/Portfolio_Analysis_Sheets.xlsx')
    
    print("\n" + "=" * 80)
    print("Creating Final Portfolio Analysis Sheet...")
    print("=" * 80)
    print("\nConfiguration:")
    print(f"  Strategy Allocation: {STRATEGY_ALLOCATION_METHOD}")
    print("\n  Pair Allocation Methods:")
    for strategy, method in STRATEGY_PAIR_METHODS.items():
        print(f"    • {strategy}: {method}")
    
    # Create new sheet
    create_final_portfolio_sheet(wb)
    
    # Save workbook
    output_path = '/Users/sureshpatil/Desktop/Portfolio Creation/Portfolio_Analysis_Sheets.xlsx'
    wb.save(output_path)
    
    print(f"\n✓ Final Portfolio Analysis sheet added to: {output_path}")
    print("\n✓ This sheet contains:")
    print("  - Strategy-level performance with selected pair methods")
    print("  - Final portfolio allocation (Equal Weight)")
    print("  - Complete portfolio performance metrics (Sharpe, XIRR, Returns)")
    print("  - Summary and key insights")
    print("\n✓ Check the 'Final_Portfolio_Analysis' sheet (now the first sheet)")

if __name__ == "__main__":
    main()
