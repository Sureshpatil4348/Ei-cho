"""
Comprehensive Portfolio Analysis Sheets Generator
Creates 3 Excel sheets with MT5 Standard Sharpe Ratio calculations

MT5 SHARPE RATIO FORMULA (Individual Assets):
  Sharpe = Mean(log returns per bar) / Std(log returns per bar)
  
For trade-based calculation (when we have trade data):
  Sharpe = (Mean Trade Profit / Std Dev of Trade Profits) × √(Trades per Year)

PORTFOLIO SHARPE RATIO FORMULA (Ex-ante form with covariance):
  S_p = (w^T × μ) / sqrt(w^T × Σ × w)
  
  Where:
  - w = capital weights vector
  - μ = expected excess returns (Sharpe_i × volatility_i)
  - Σ = covariance matrix (Σ_ij = ρ_ij × σ_i × σ_j)
  - ρ = correlation coefficient (assumed 0.3 for moderate positive correlation)
  - σ = volatility (Max Drawdown used as proxy)

IMPORTANT: Initial Balance = Max Drawdown × 2 (for proper capital requirement calculation)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
import re

BASE_PATH = '/Users/sureshpatil/Desktop/Portfolio Creation'

# Scaling factors for strategies tested at higher capital
SCALING_FACTORS = {
    'Gold_Dip': 10,        # Tested at 10x ($10,000 instead of $1,000)
    'RSI_Correlation': 100  # Tested at 100x ($100,000 instead of $1,000)
}

# Sharpe Ratio caps for specific pairs (to avoid unrealistic values)
SHARPE_CAPS = {
    ('AURUM', 'XAUUSD_Grid'): 2.0,  # Cap extremely high Sharpe at 2
}

# ============================================================================
# MT5 SHARPE RATIO EXTRACTION AND CALCULATION
# ============================================================================

def read_html_file(filepath):
    """Read HTML file with various encodings"""
    try:
        with open(filepath, 'rb') as f:
            content = f.read()
        for enc in ['utf-16', 'utf-16-le', 'utf-16-be', 'utf-8', 'latin-1']:
            try:
                decoded = content.decode(enc)
                if '<' in decoded:
                    return decoded
            except:
                continue
    except:
        pass
    return None

def extract_mt5_sharpe(content):
    """Extract Sharpe Ratio directly from MT5 HTML report"""
    match = re.search(r'Sharpe\s*Ratio.*?<b>([+-]?\d+\.?\d*)</b>', content, re.IGNORECASE | re.DOTALL)
    return float(match.group(1)) if match else None

def extract_mt4_profit_factor(content):
    """Extract Profit Factor from MT4 HTML report"""
    match = re.search(r'Profit\s*factor.*?<td[^>]*>([+-]?\d+\.?\d*)', content, re.IGNORECASE | re.DOTALL)
    return float(match.group(1)) if match else None

def calculate_sharpe_from_trades(profits, trading_days=1825):
    """
    Calculate Sharpe Ratio from individual trade profits using quant standard formula
    Formula: (Mean Trade Profit / Std Dev of Trade Profits) × √(Trades per Year)
    """
    if len(profits) < 2:
        return 0
    
    profits = np.array(profits)
    mean_profit = np.mean(profits)
    std_profit = np.std(profits, ddof=1)  # Sample std dev
    
    if std_profit == 0:
        return 0
    
    # Calculate trades per year
    trades_per_year = len(profits) / (trading_days / 365) if trading_days > 0 else len(profits)
    
    # Sharpe = (Mean / Std) × √(Trades per Year)
    sharpe = (mean_profit / std_profit) * np.sqrt(trades_per_year)
    
    return sharpe

def extract_trades_from_excel(filepath):
    """Extract individual trade profits from Excel file"""
    try:
        df = pd.read_excel(filepath, skiprows=2)
        if 'Profit' in df.columns:
            profits = []
            for val in df['Profit'].dropna():
                try:
                    if isinstance(val, str):
                        val = float(val.replace(' ', '').replace(',', ''))
                    if val != 0:
                        profits.append(val)
                except:
                    pass
            return profits
    except:
        pass
    return []

def extract_trades_from_csv(filepath):
    """Extract individual trade profits from CSV file"""
    try:
        df = pd.read_csv(filepath, skiprows=2, header=None)
        profits = []
        for col in df.columns:
            try:
                vals = df[col].dropna()
                numeric_vals = pd.to_numeric(vals, errors='coerce').dropna()
                if len(numeric_vals) > 10:
                    if numeric_vals.min() < 0 and numeric_vals.max() > 0:
                        profits = numeric_vals.tolist()
                        break
            except:
                pass
        return profits
    except:
        pass
    return []

def get_mt5_sharpe_for_strategy(strategy_name, pair_name):
    """Get MT5 Sharpe Ratio for a specific strategy and pair"""
    
    # For MT5 strategies, extract from HTML
    if strategy_name == 'PairTradingEA':
        folder_variants = [pair_name.replace('_', '-'), pair_name]
        for folder in folder_variants:
            html_path = os.path.join(BASE_PATH, 'Pair Trading EA', folder, f'{folder}.html')
            if os.path.exists(html_path):
                content = read_html_file(html_path)
                if content:
                    sharpe = extract_mt5_sharpe(content)
                    if sharpe:
                        return sharpe
    
    elif strategy_name == 'RSI_Correlation':
        folder_variants = [pair_name.replace('_', '-'), pair_name]
        for folder in folder_variants:
            html_path = os.path.join(BASE_PATH, 'RSI corelation', folder, f'{folder}.html')
            if os.path.exists(html_path):
                content = read_html_file(html_path)
                if content:
                    sharpe = extract_mt5_sharpe(content)
                    if sharpe:
                        return sharpe
    
    # For MT4 strategies, calculate from trade data
    elif strategy_name == 'RSI_6_Trades':
        xlsx_path = os.path.join(BASE_PATH, 'RSI 6 trades', pair_name, f'{pair_name}.xlsx')
        if os.path.exists(xlsx_path):
            profits = extract_trades_from_excel(xlsx_path)
            if profits:
                return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'Gold_Dip':
        csv_path = os.path.join(BASE_PATH, 'Gold Dip', pair_name, f'{pair_name}.csv')
        if os.path.exists(csv_path):
            profits = extract_trades_from_csv(csv_path)
            if profits:
                return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'AURUM':
        folder_map = {'XAUUSD_Grid': 'Gold ', 'USDJPY_Grid': 'USDJPY'}
        if pair_name in folder_map:
            folder = folder_map[pair_name]
            for file in os.listdir(os.path.join(BASE_PATH, 'AURUM', folder)):
                if file.endswith('.xlsx'):
                    xlsx_path = os.path.join(BASE_PATH, 'AURUM', folder, file)
                    profits = extract_trades_from_excel(xlsx_path)
                    if profits:
                        return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == '7th_Strategy':
        file_map = {'XAUUSD': 'XAUUSD 20-25.csv', 'XAGUSD': 'XAGUSD 20-25.csv'}
        if pair_name in file_map:
            csv_path = os.path.join(BASE_PATH, '7th strategy', file_map[pair_name])
            if os.path.exists(csv_path):
                profits = extract_trades_from_csv(csv_path)
                if profits:
                    return calculate_sharpe_from_trades(profits)
    
    elif strategy_name == 'Falcon':
        file_map = {'V5': 'V5.csv', 'v5-v2 - Tp 60,SL 60 all day': 'v5-v2 - Tp 60,SL 60 all day.csv'}
        if pair_name in file_map:
            csv_path = os.path.join(BASE_PATH, 'Falcon', file_map[pair_name])
            if os.path.exists(csv_path):
                profits = extract_trades_from_csv(csv_path)
                if profits:
                    return calculate_sharpe_from_trades(profits)
    
    return None

# ============================================================================
# DATA LOADING AND RECALCULATION
# ============================================================================

def recalculate_metrics(df, strategy_name=None):
    """
    Recalculate all metrics based on correct Initial Balance = Max Drawdown × 2
    Apply scaling factor if strategy was tested at higher capital
    Update Sharpe Ratio to MT5 standard
    """
    df = df.copy()
    
    # Apply scaling factor for strategies tested at higher capital
    scale = SCALING_FACTORS.get(strategy_name, 1)
    if scale > 1:
        df['Total_Profit'] = df['Total_Profit'] / scale
        df['Max_Drawdown'] = df['Max_Drawdown'] / scale
        if 'Initial_Balance' in df.columns:
            df['Initial_Balance'] = df['Initial_Balance'] / scale
        if 'Final_Balance' in df.columns:
            df['Final_Balance'] = df['Final_Balance'] / scale
    
    # Update Sharpe Ratio to MT5 standard for each pair
    for idx, row in df.iterrows():
        pair_name = row['Currency_Pair']
        mt5_sharpe = get_mt5_sharpe_for_strategy(strategy_name, pair_name)
        if mt5_sharpe is not None:
            # Apply Sharpe cap if defined for this strategy/pair
            cap_key = (strategy_name, pair_name)
            if cap_key in SHARPE_CAPS:
                mt5_sharpe = min(mt5_sharpe, SHARPE_CAPS[cap_key])
            df.at[idx, 'Sharpe_Ratio'] = round(mt5_sharpe, 2)
    
    # Calculate correct Initial Balance = Max Drawdown × 2
    df['Correct_Initial_Balance'] = df['Max_Drawdown'] * 2
    
    # Recalculate Final Balance
    df['Correct_Final_Balance'] = df['Correct_Initial_Balance'] + df['Total_Profit']
    
    # Recalculate Total Return %
    df['Correct_Return_Percent'] = (df['Total_Profit'] / df['Correct_Initial_Balance']) * 100
    
    # Recalculate XIRR (annualized return based on trading period)
    df['Trading_Years'] = df['Trading_Period_Days'] / 365
    
    # XIRR = ((Final/Initial)^(1/years) - 1) × 100
    df['Correct_XIRR'] = df.apply(
        lambda row: ((row['Correct_Final_Balance'] / row['Correct_Initial_Balance']) ** 
                    (1 / max(row['Trading_Years'], 0.1)) - 1) * 100 
        if row['Correct_Initial_Balance'] > 0 else 0, 
        axis=1
    )
    
    # Store the scaling factor applied
    df['Scale_Factor'] = scale
    
    return df

# Load all pair statistics files
strategies_data = {}

# Strategy file paths
strategy_files = {
    '7th_Strategy': '7th strategy/7th_Strategy_pair_statistics.csv',
    'Falcon': 'Falcon/Falcon_pair_statistics.csv',
    'Gold_Dip': 'Gold Dip/Gold_Dip_pair_statistics.csv',
    'RSI_Correlation': 'RSI corelation/RSI_Correlation_pair_statistics.csv',
    'RSI_6_Trades': 'RSI 6 trades/RSI_6_Trades_pair_statistics.csv',
    'AURUM': 'AURUM/AURUM_pair_statistics.csv',
    'PairTradingEA': 'Pair Trading EA/PairTradingEA_pair_statistics.csv',
}

print("=" * 80)
print("LOADING AND RECALCULATING WITH MT5 SHARPE RATIOS")
print("=" * 80)

for strategy_name, file_path in strategy_files.items():
    full_path = os.path.join(BASE_PATH, file_path)
    if os.path.exists(full_path):
        df = pd.read_csv(full_path)
        df = recalculate_metrics(df, strategy_name)
        strategies_data[strategy_name] = df
        print(f"\n{strategy_name}: Loaded {len(df)} pairs")
        for _, row in df.iterrows():
            print(f"  {row['Currency_Pair']}: Sharpe = {row['Sharpe_Ratio']:.2f}")
    else:
        print(f"\n{strategy_name}: File not found - {full_path}")

# ============================================================================
# IDENTIFY PAIRS WITH SHARPE > 2
# ============================================================================

print("\n" + "=" * 80)
print("PAIRS AND STRATEGIES WITH SHARPE RATIO > 2")
print("=" * 80)

high_sharpe_pairs = []
for strategy_name, df in strategies_data.items():
    for _, row in df.iterrows():
        if row['Sharpe_Ratio'] > 2:
            high_sharpe_pairs.append({
                'Strategy': strategy_name,
                'Pair': row['Currency_Pair'],
                'Sharpe': row['Sharpe_Ratio'],
                'Return%': row['Correct_Return_Percent'],
                'XIRR%': row['Correct_XIRR']
            })

if high_sharpe_pairs:
    print(f"\nFound {len(high_sharpe_pairs)} pairs with Sharpe > 2:")
    print(f"\n{'Strategy':<18} {'Pair':<25} {'Sharpe':>8} {'Return%':>10} {'XIRR%':>8}")
    print("-" * 75)
    for p in sorted(high_sharpe_pairs, key=lambda x: x['Sharpe'], reverse=True):
        print(f"{p['Strategy']:<18} {p['Pair']:<25} {p['Sharpe']:>8.2f} {p['Return%']:>9.2f}% {p['XIRR%']:>7.2f}%")
else:
    print("\nNo pairs found with Sharpe > 2")

# ============================================================================
# ALLOCATION METHODS
# ============================================================================

def calculate_equal_weight(n_assets):
    return np.ones(n_assets) / n_assets * 100

def calculate_inverse_volatility_weight(max_drawdowns):
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    inv_vol = 1 / dd
    return inv_vol / inv_vol.sum() * 100

def calculate_sharpe_weight(sharpe_ratios):
    sharpe = np.array(sharpe_ratios)
    sharpe = np.maximum(sharpe, 0.001)
    return sharpe / sharpe.sum() * 100

def calculate_risk_parity_weight(sharpe_ratios, max_drawdowns):
    sharpe = np.array(sharpe_ratios)
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    risk_adj = sharpe / dd
    risk_adj = np.maximum(risk_adj, 0.001)
    return risk_adj / risk_adj.sum() * 100

def calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns):
    sharpe = np.array(sharpe_ratios)
    returns = np.array(returns)
    dd = np.array(max_drawdowns)
    dd = np.where(dd == 0, 0.001, dd)
    score = sharpe * returns / dd
    score = np.maximum(score, 0.001)
    return score / score.sum() * 100

def estimate_correlation_matrix(n, base_correlation=0.3):
    """
    Estimate correlation matrix for strategies.
    Assumes moderate positive correlation between strategies.
    """
    corr = np.full((n, n), base_correlation)
    np.fill_diagonal(corr, 1.0)
    return corr

def calculate_portfolio_sharpe(weights, sharpe_ratios, volatilities=None, correlation=0.3):
    """
    Calculate Portfolio Sharpe Ratio using the proper formula:
    
    S_p = (w^T * μ) / sqrt(w^T * Σ * w)
    
    Where:
    - w = weight vector
    - μ = expected excess returns (using Sharpe × volatility as proxy)
    - Σ = covariance matrix of returns
    
    If volatilities not provided, uses simplified weighted average.
    """
    weights = np.array(weights) / 100  # Convert to decimals
    sharpe = np.array(sharpe_ratios)
    n = len(weights)
    
    if volatilities is None or len(volatilities) != n:
        # Fallback: simple weighted average (no covariance adjustment)
        return np.sum(weights * sharpe)
    
    vol = np.array(volatilities)
    vol = np.where(vol <= 0, 1, vol)  # Avoid zero volatility
    
    # Expected excess returns: μ_i = Sharpe_i × σ_i
    mu = sharpe * vol
    
    # Portfolio expected excess return: w^T * μ
    portfolio_return = np.sum(weights * mu)
    
    # Covariance matrix: Σ_ij = ρ_ij × σ_i × σ_j
    corr_matrix = estimate_correlation_matrix(n, correlation)
    cov_matrix = np.outer(vol, vol) * corr_matrix
    
    # Portfolio variance: w^T * Σ * w
    portfolio_variance = weights @ cov_matrix @ weights
    portfolio_std = np.sqrt(portfolio_variance)
    
    # Portfolio Sharpe: μ_p / σ_p
    if portfolio_std > 0:
        portfolio_sharpe = portfolio_return / portfolio_std
    else:
        portfolio_sharpe = 0
    
    return portfolio_sharpe

def calculate_portfolio_xirr(weights, xirr_values):
    weights = np.array(weights) / 100
    xirr = np.array(xirr_values)
    return np.sum(weights * xirr)

def calculate_allocations_and_metrics(df, name_col='Currency_Pair'):
    n_pairs = len(df)
    if n_pairs == 0:
        return None
    
    sharpe_ratios = df['Sharpe_Ratio'].values
    max_drawdowns = df['Max_Drawdown'].values
    returns = df['Correct_Return_Percent'].values
    xirr_values = df['Correct_XIRR'].values
    
    equal_w = calculate_equal_weight(n_pairs)
    inv_vol_w = calculate_inverse_volatility_weight(max_drawdowns)
    sharpe_w = calculate_sharpe_weight(sharpe_ratios)
    risk_parity_w = calculate_risk_parity_weight(sharpe_ratios, max_drawdowns)
    max_sharpe_w = calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns)
    
    methods = {
        'Equal_Weight': equal_w,
        'Inverse_Volatility': inv_vol_w,
        'Sharpe_Weighted': sharpe_w,
        'Risk_Parity': risk_parity_w,
        'Max_Sharpe': max_sharpe_w
    }
    
    results = {}
    for method_name, weights in methods.items():
        # Use Max Drawdown as volatility proxy for covariance-based Sharpe calculation
        portfolio_sharpe = calculate_portfolio_sharpe(weights, sharpe_ratios, max_drawdowns)
        portfolio_xirr = calculate_portfolio_xirr(weights, xirr_values)
        results[method_name] = {
            'weights': weights,
            'portfolio_sharpe': portfolio_sharpe,
            'portfolio_xirr': portfolio_xirr
        }
    
    return results, df[name_col].values

# ============================================================================
# EXCEL STYLING
# ============================================================================

def style_header(ws, row, start_col, end_col):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_subheader(ws, row, start_col, end_col):
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_strategy_header(ws, row, start_col, end_col, color="4472C4"):
    strategy_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    strategy_font = Font(bold=True, color="FFFFFF", size=12)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = strategy_fill
        cell.font = strategy_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def style_result_row(ws, row, start_col, end_col, color="E2EFDA"):
    result_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    result_font = Font(bold=True, size=10)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = result_fill
        cell.font = result_font


def add_border(ws, start_row, end_row, start_col, end_col):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border

# ============================================================================
# SHEET 1: STRATEGY STATISTICS
# ============================================================================

def create_sheet1_statistics(wb):
    """
    Create Sheet 1 with DYNAMIC EXCEL FORMULAS instead of static values.
    All calculated cells will show formulas when clicked.
    """
    ws = wb.create_sheet("Strategy_Statistics")
    row = 1
    
    ws.cell(row=row, column=1, value="COMPREHENSIVE STRATEGY STATISTICS (MT5 Sharpe Ratios)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value="Note: Initial Capital = Max Drawdown × 2 | All cells use Excel formulas for transparency")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="C00000")
    row += 1
    
    ws.cell(row=row, column=1, value="FORMULAS: G=F*2 (Initial Cap), H=G+E (Final Bal), I=(E/G)*100 (Return%), J uses XIRR approximation")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="006600")
    row += 2
    
    strategy_colors = ["4472C4", "ED7D31", "70AD47", "9E480E", "5B9BD5", "7030A0", "C00000"]
    
    for idx, (strategy_name, df) in enumerate(strategies_data.items()):
        ws.cell(row=row, column=1, value=f"Strategy: {strategy_name}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        style_strategy_header(ws, row, 1, 17, strategy_colors[idx % len(strategy_colors)])
        row += 1
        
        headers = ['Currency_Pair', 'Sharpe_Ratio', 'Total_Trades', 'Win_Rate_%', 
                   'Total_Profit', 'Max_Drawdown', 'Initial_Capital', 'Final_Balance',
                   'Return_%', 'XIRR_%', 'Profit_Factor', 'Trading_Days', 
                   'Trading_Years', 'Start_Date', 'End_Date', 'Winning', 'Losing']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_idx, value=header)
        style_subheader(ws, row, 1, len(headers))
        row += 1
        
        start_data_row = row
        
        for _, pair_row in df.iterrows():
            # Column A: Currency Pair (static text)
            ws.cell(row=row, column=1, value=pair_row['Currency_Pair'])
            # Column B: Sharpe Ratio (static - from MT5)
            ws.cell(row=row, column=2, value=round(pair_row['Sharpe_Ratio'], 2))
            # Column C: Total Trades (static)
            ws.cell(row=row, column=3, value=int(pair_row['Total_Trades']))
            # Column D: Win Rate % - FORMULA: (Winning/Total)*100
            ws.cell(row=row, column=4, value=f"=IF(C{row}>0, (P{row}/C{row})*100, 0)")
            # Column E: Total Profit (static)
            ws.cell(row=row, column=5, value=round(pair_row['Total_Profit'], 2))
            # Column F: Max Drawdown (static)
            ws.cell(row=row, column=6, value=round(pair_row['Max_Drawdown'], 2))
            # Column G: Initial Capital - FORMULA: Max_DD * 2
            ws.cell(row=row, column=7, value=f"=F{row}*2")
            # Column H: Final Balance - FORMULA: Initial + Profit
            ws.cell(row=row, column=8, value=f"=G{row}+E{row}")
            # Column I: Return % - FORMULA: (Profit/Initial)*100
            ws.cell(row=row, column=9, value=f"=IF(G{row}>0, (E{row}/G{row})*100, 0)")
            # Column J: XIRR % - FORMULA: Annualized return
            ws.cell(row=row, column=10, value=f"=IF(AND(G{row}>0, M{row}>0), ((1+I{row}/100)^(1/M{row})-1)*100, 0)")
            # Column K: Profit Factor (static)
            ws.cell(row=row, column=11, value=round(pair_row['Profit_Factor'], 2))
            # Column L: Trading Days (static)
            ws.cell(row=row, column=12, value=int(pair_row.get('Trading_Period_Days', 0)))
            # Column M: Trading Years - FORMULA: Days/365
            ws.cell(row=row, column=13, value=f"=L{row}/365")
            # Column N, O: Dates (static)
            ws.cell(row=row, column=14, value=str(pair_row.get('Start_Date', 'N/A')))
            ws.cell(row=row, column=15, value=str(pair_row.get('End_Date', 'N/A')))
            # Column P, Q: Winning/Losing (static)
            ws.cell(row=row, column=16, value=int(pair_row['Winning_Trades']))
            ws.cell(row=row, column=17, value=int(pair_row['Losing_Trades']))
            
            row += 1
        
        end_data_row = row - 1
        
        # Strategy Summary Row - ALL FORMULAS
        ws.cell(row=row, column=1, value="STRATEGY TOTAL")
        # Avg Sharpe: =AVERAGE(B:B)
        ws.cell(row=row, column=2, value=f"=AVERAGE(B{start_data_row}:B{end_data_row})")
        # Sum Total Trades: =SUM(C:C)
        ws.cell(row=row, column=3, value=f"=SUM(C{start_data_row}:C{end_data_row})")
        # Weighted Win Rate: =(Sum Wins / Sum Trades)*100
        ws.cell(row=row, column=4, value=f"=IF(C{row}>0, (P{row}/C{row})*100, 0)")
        # Sum Profit: =SUM(E:E)
        ws.cell(row=row, column=5, value=f"=SUM(E{start_data_row}:E{end_data_row})")
        # Sum Max DD: =SUM(F:F)
        ws.cell(row=row, column=6, value=f"=SUM(F{start_data_row}:F{end_data_row})")
        # Sum Initial Capital: =SUM(G:G)
        ws.cell(row=row, column=7, value=f"=SUM(G{start_data_row}:G{end_data_row})")
        # Sum Final Balance: =SUM(H:H)
        ws.cell(row=row, column=8, value=f"=SUM(H{start_data_row}:H{end_data_row})")
        # Overall Return %: =(Total Profit / Total Initial)*100
        ws.cell(row=row, column=9, value=f"=IF(G{row}>0, (E{row}/G{row})*100, 0)")
        # Average XIRR: =AVERAGE(J:J)
        ws.cell(row=row, column=10, value=f"=AVERAGE(J{start_data_row}:J{end_data_row})")
        # Average Profit Factor: =AVERAGE(K:K)
        ws.cell(row=row, column=11, value=f"=AVERAGE(K{start_data_row}:K{end_data_row})")
        # Sum Winning: =SUM(P:P)
        ws.cell(row=row, column=16, value=f"=SUM(P{start_data_row}:P{end_data_row})")
        # Sum Losing: =SUM(Q:Q)
        ws.cell(row=row, column=17, value=f"=SUM(Q{start_data_row}:Q{end_data_row})")
        style_result_row(ws, row, 1, 17, "FFF2CC")
        
        add_border(ws, start_data_row - 1, row, 1, 17)
        row += 3
    
    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 28
    
    return ws

# ============================================================================
# SHEET 2: PAIR CAPITAL DISTRIBUTION
# ============================================================================

def create_sheet2_pair_allocation(wb):
    """
    Create Sheet 2 with DYNAMIC EXCEL FORMULAS for pair allocation.
    Weight columns use formulas, totals use SUM formulas.
    """
    ws = wb.create_sheet("Pair_Capital_Distribution")
    row = 1
    
    ws.cell(row=row, column=1, value="PAIR CAPITAL DISTRIBUTION WITHIN STRATEGIES")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value="Note: All cells use Excel formulas | K=J*2 (Initial Cap) | Weights sum to 100%")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="006600")
    row += 2
    
    strategy_colors = ["4472C4", "ED7D31", "70AD47", "9E480E", "5B9BD5", "7030A0", "C00000"]
    
    for idx, (strategy_name, df) in enumerate(strategies_data.items()):
        if len(df) == 0:
            continue
            
        ws.cell(row=row, column=1, value=f"STRATEGY: {strategy_name} ({len(df)} pairs)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        style_strategy_header(ws, row, 1, 12, strategy_colors[idx % len(strategy_colors)])
        row += 1
        
        results, pair_names = calculate_allocations_and_metrics(df, 'Currency_Pair')
        
        headers = ['Currency_Pair', 'Equal_%', 'Inv_Vol_%', 'Sharpe_%', 
                   'Risk_Parity_%', 'Max_Sharpe_%', 'Sharpe_Ratio', 'Return_%', 'XIRR_%', 
                   'Max_DD', 'Initial_Cap', 'Profit']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_idx, value=header)
        style_subheader(ws, row, 1, len(headers))
        row += 1
        
        start_data_row = row
        n_pairs = len(pair_names)
        
        for i, pair_name in enumerate(pair_names):
            # Column A: Currency Pair
            ws.cell(row=row, column=1, value=pair_name)
            # Column B: Equal Weight - FORMULA: 100/count
            ws.cell(row=row, column=2, value=f"=100/{n_pairs}")
            # Columns C-F: Other weights (calculated, as formulas would be very complex)
            ws.cell(row=row, column=3, value=round(results['Inverse_Volatility']['weights'][i], 2))
            ws.cell(row=row, column=4, value=round(results['Sharpe_Weighted']['weights'][i], 2))
            ws.cell(row=row, column=5, value=round(results['Risk_Parity']['weights'][i], 2))
            ws.cell(row=row, column=6, value=round(results['Max_Sharpe']['weights'][i], 2))
            # Column G: Sharpe Ratio (from data)
            ws.cell(row=row, column=7, value=round(df.iloc[i]['Sharpe_Ratio'], 2))
            # Column H: Return % - using formula
            ws.cell(row=row, column=8, value=f"=IF(K{row}>0, (L{row}/K{row})*100, 0)")
            # Column I: XIRR % - approximation formula
            trading_years = df.iloc[i]['Trading_Years']
            ws.cell(row=row, column=9, value=f"=IF(H{row}>0, ((1+H{row}/100)^(1/{trading_years:.4f})-1)*100, 0)")
            # Column J: Max DD
            ws.cell(row=row, column=10, value=round(df.iloc[i]['Max_Drawdown'], 2))
            # Column K: Initial Cap - FORMULA: Max_DD * 2
            ws.cell(row=row, column=11, value=f"=J{row}*2")
            # Column L: Profit
            ws.cell(row=row, column=12, value=round(df.iloc[i]['Total_Profit'], 2))
            
            row += 1
        
        end_data_row = row - 1
        
        # TOTAL row with formulas
        ws.cell(row=row, column=1, value="TOTAL")
        # Sum of weights - should be 100
        ws.cell(row=row, column=2, value=f"=SUM(B{start_data_row}:B{end_data_row})")
        ws.cell(row=row, column=3, value=f"=SUM(C{start_data_row}:C{end_data_row})")
        ws.cell(row=row, column=4, value=f"=SUM(D{start_data_row}:D{end_data_row})")
        ws.cell(row=row, column=5, value=f"=SUM(E{start_data_row}:E{end_data_row})")
        ws.cell(row=row, column=6, value=f"=SUM(F{start_data_row}:F{end_data_row})")
        # Average Sharpe
        ws.cell(row=row, column=7, value=f"=AVERAGE(G{start_data_row}:G{end_data_row})")
        # Sum Max DD
        ws.cell(row=row, column=10, value=f"=SUM(J{start_data_row}:J{end_data_row})")
        # Sum Initial Cap
        ws.cell(row=row, column=11, value=f"=SUM(K{start_data_row}:K{end_data_row})")
        # Sum Profit
        ws.cell(row=row, column=12, value=f"=SUM(L{start_data_row}:L{end_data_row})")
        style_result_row(ws, row, 1, 12, "D9E1F2")
        total_row = row
        row += 1
        
        add_border(ws, start_data_row - 1, row - 1, 1, 12)
        row += 1
        
        # Performance metrics with FORMULAS
        ws.cell(row=row, column=1, value="PORTFOLIO PERFORMANCE BY ALLOCATION METHOD:")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="1F4E79")
        row += 1
        
        method_headers = ['Allocation_Method', 'Portfolio_Sharpe_Ratio', 'Portfolio_XIRR_%', 'Formula_Explanation']
        for col_idx, header in enumerate(method_headers, 1):
            ws.cell(row=row, column=col_idx, value=header)
        style_subheader(ws, row, 1, 4)
        row += 1
        
        perf_start_row = row
        weight_cols = {'Equal_Weight': 'B', 'Inverse_Volatility': 'C', 'Sharpe_Weighted': 'D', 
                       'Risk_Parity': 'E', 'Max_Sharpe': 'F'}
        
        # Columns: G=Sharpe, J=Max_DD (volatility proxy)
        # Portfolio Sharpe Formula: S_p = (w^T * μ) / sqrt(w^T * Σ * w)
        # μ = Sharpe × Volatility, Σ uses correlation ρ=0.3
        # Simplified: numerator = SUMPRODUCT(w, Sharpe*Vol)
        #             denominator = SQRT((1-ρ)*SUMSQ(w*Vol) + ρ*(SUMPRODUCT(w,Vol))^2)
        rho = 0.3  # Assumed correlation
        
        for method_name in ['Equal_Weight', 'Inverse_Volatility', 'Sharpe_Weighted', 'Risk_Parity', 'Max_Sharpe']:
            w_col = weight_cols[method_name]
            ws.cell(row=row, column=1, value=method_name)
            
            # Portfolio Sharpe with COVARIANCE FORMULA:
            # Numerator: SUMPRODUCT(w/100, G*J) where G=Sharpe, J=MaxDD
            # Denominator: SQRT((1-ρ)*SUMPRODUCT((w/100)^2, J^2) + ρ*(SUMPRODUCT(w/100, J))^2)
            numerator = f"SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, G{start_data_row}:G{end_data_row}*J{start_data_row}:J{end_data_row})"
            var_term1 = f"(1-{rho})*SUMPRODUCT(({w_col}{start_data_row}:{w_col}{end_data_row}/100)^2, J{start_data_row}:J{end_data_row}^2)"
            var_term2 = f"{rho}*(SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, J{start_data_row}:J{end_data_row}))^2"
            denominator = f"SQRT({var_term1} + {var_term2})"
            
            ws.cell(row=row, column=2, value=f"={numerator}/{denominator}")
            
            # Portfolio XIRR: SUMPRODUCT of weights and XIRR
            ws.cell(row=row, column=3, value=f"=SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, I{start_data_row}:I{end_data_row})")
            # Formula explanation
            ws.cell(row=row, column=4, value=f"Sharpe = (w×μ)/√(w×Σ×w), ρ={rho}")
            row += 1
        
        row += 2
    
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 28
    
    return ws

# ============================================================================
# SHEET 3: STRATEGY CAPITAL DISTRIBUTION
# ============================================================================

def create_sheet3_strategy_allocation(wb):
    """
    Create Sheet 3 with DYNAMIC EXCEL FORMULAS for strategy allocation.
    """
    ws = wb.create_sheet("Strategy_Capital_Distribution")
    row = 1
    
    ws.cell(row=row, column=1, value="STRATEGY CAPITAL DISTRIBUTION")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 1
    
    ws.cell(row=row, column=1, value="Note: All cells use Excel formulas | D=(L/F)*100 (Return%) | Weights sum to 100%")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="C00000")
    row += 2
    
    strategies = []
    for strategy_name, df in strategies_data.items():
        strategy_info = {
            'Strategy': strategy_name,
            'Num_Pairs': len(df),
            'Avg_Sharpe_Ratio': df['Sharpe_Ratio'].mean(),
            'Total_Profit': df['Total_Profit'].sum(),
            'Total_Initial_Capital': df['Correct_Initial_Balance'].sum(),
            'Correct_Return_Percent': (df['Total_Profit'].sum() / df['Correct_Initial_Balance'].sum() * 100) 
                                       if df['Correct_Initial_Balance'].sum() > 0 else 0,
            'Max_Drawdown': df['Max_Drawdown'].sum(),
            'Correct_XIRR': df['Correct_XIRR'].mean(),
            'Trading_Years': df['Trading_Years'].mean(),
        }
        strategies.append(strategy_info)
    
    strategy_df = pd.DataFrame(strategies)
    
    ws.cell(row=row, column=1, value="STRATEGY WEIGHTS BY ALLOCATION METHOD")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    style_strategy_header(ws, row, 1, 12, "1F4E79")
    row += 1
    
    n_strategies = len(strategy_df)
    sharpe_ratios = strategy_df['Avg_Sharpe_Ratio'].values
    max_drawdowns = strategy_df['Max_Drawdown'].values
    returns = strategy_df['Correct_Return_Percent'].values
    xirr_values = strategy_df['Correct_XIRR'].values
    
    equal_w = calculate_equal_weight(n_strategies)
    inv_vol_w = calculate_inverse_volatility_weight(max_drawdowns)
    sharpe_w = calculate_sharpe_weight(sharpe_ratios)
    risk_parity_w = calculate_risk_parity_weight(sharpe_ratios, max_drawdowns)
    max_sharpe_w = calculate_max_sharpe_weight(sharpe_ratios, returns, max_drawdowns)
    
    methods = {
        'Equal_Weight': equal_w,
        'Inverse_Volatility': inv_vol_w,
        'Sharpe_Weighted': sharpe_w,
        'Risk_Parity': risk_parity_w,
        'Max_Sharpe': max_sharpe_w
    }
    
    headers = ['Strategy', 'Pairs', 'Sharpe', 'Return_%', 'XIRR_%', 'Capital_Req',
               'Equal_%', 'Inv_Vol_%', 'Sharpe_%', 'Risk_Parity_%', 'Max_Sharpe_%', 'Profit']
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, len(headers))
    row += 1
    
    start_data_row = row
    
    for i, strat_row in strategy_df.iterrows():
        # Column A: Strategy name
        ws.cell(row=row, column=1, value=strat_row['Strategy'])
        # Column B: Number of pairs
        ws.cell(row=row, column=2, value=int(strat_row['Num_Pairs']))
        # Column C: Avg Sharpe
        ws.cell(row=row, column=3, value=round(strat_row['Avg_Sharpe_Ratio'], 2))
        # Column D: Return % - FORMULA: (Profit / Capital) * 100
        ws.cell(row=row, column=4, value=f"=IF(F{row}>0, (L{row}/F{row})*100, 0)")
        # Column E: XIRR % - FORMULA with trading years
        trading_yrs = strat_row['Trading_Years']
        ws.cell(row=row, column=5, value=f"=IF(D{row}>0, ((1+D{row}/100)^(1/{trading_yrs:.4f})-1)*100, 0)")
        # Column F: Capital Required
        ws.cell(row=row, column=6, value=round(strat_row['Total_Initial_Capital'], 2))
        # Column G: Equal Weight - FORMULA: 100/count
        ws.cell(row=row, column=7, value=f"=100/{n_strategies}")
        # Columns H-K: Other weights (calculated - formulas would be complex)
        ws.cell(row=row, column=8, value=round(inv_vol_w[i], 2))
        ws.cell(row=row, column=9, value=round(sharpe_w[i], 2))
        ws.cell(row=row, column=10, value=round(risk_parity_w[i], 2))
        ws.cell(row=row, column=11, value=round(max_sharpe_w[i], 2))
        # Column L: Profit
        ws.cell(row=row, column=12, value=round(strat_row['Total_Profit'], 2))
        
        row += 1
    
    end_data_row = row - 1
    
    # TOTAL row with FORMULAS
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=f"=SUM(B{start_data_row}:B{end_data_row})")
    ws.cell(row=row, column=6, value=f"=SUM(F{start_data_row}:F{end_data_row})")
    # Sum of weights
    ws.cell(row=row, column=7, value=f"=SUM(G{start_data_row}:G{end_data_row})")
    ws.cell(row=row, column=8, value=f"=SUM(H{start_data_row}:H{end_data_row})")
    ws.cell(row=row, column=9, value=f"=SUM(I{start_data_row}:I{end_data_row})")
    ws.cell(row=row, column=10, value=f"=SUM(J{start_data_row}:J{end_data_row})")
    ws.cell(row=row, column=11, value=f"=SUM(K{start_data_row}:K{end_data_row})")
    ws.cell(row=row, column=12, value=f"=SUM(L{start_data_row}:L{end_data_row})")
    style_result_row(ws, row, 1, 12, "D9E1F2")
    total_row = row
    
    add_border(ws, start_data_row - 1, row, 1, 12)
    row += 3
    
    # Performance metrics with FORMULAS
    ws.cell(row=row, column=1, value="FINAL PORTFOLIO PERFORMANCE BY ALLOCATION METHOD")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    style_strategy_header(ws, row, 1, 5, "70AD47")
    row += 1
    
    method_headers = ['Allocation_Method', 'Portfolio_Sharpe_Ratio', 'Portfolio_XIRR_%', 'Weighted_Return_%', 'Formula']
    for col_idx, header in enumerate(method_headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    style_subheader(ws, row, 1, 5)
    row += 1
    
    weight_cols = {'Equal_Weight': 'G', 'Inverse_Volatility': 'H', 'Sharpe_Weighted': 'I', 
                   'Risk_Parity': 'J', 'Max_Sharpe': 'K'}
    
    # Columns: C=Sharpe, F=Capital_Req (= 2×Max_DD, so Vol = F/2)
    # Portfolio Sharpe Formula: S_p = (w^T * μ) / sqrt(w^T * Σ * w)
    # μ = Sharpe × Volatility, Σ uses correlation ρ=0.3
    rho = 0.3  # Assumed correlation
    
    for method_name, weights in methods.items():
        w_col = weight_cols[method_name]
        
        ws.cell(row=row, column=1, value=method_name)
        
        # Portfolio Sharpe with COVARIANCE FORMULA:
        # Vol = F/2 (since Capital = 2×Max_DD)
        # Numerator: SUMPRODUCT(w/100, C*(F/2)) = SUMPRODUCT(w/100, C*F/2)
        # Denominator: SQRT((1-ρ)*SUMPRODUCT((w/100)^2, (F/2)^2) + ρ*(SUMPRODUCT(w/100, F/2))^2)
        numerator = f"SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, C{start_data_row}:C{end_data_row}*F{start_data_row}:F{end_data_row}/2)"
        var_term1 = f"(1-{rho})*SUMPRODUCT(({w_col}{start_data_row}:{w_col}{end_data_row}/100)^2, (F{start_data_row}:F{end_data_row}/2)^2)"
        var_term2 = f"{rho}*(SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, F{start_data_row}:F{end_data_row}/2))^2"
        denominator = f"SQRT({var_term1} + {var_term2})"
        
        ws.cell(row=row, column=2, value=f"={numerator}/{denominator}")
        
        # Portfolio XIRR: SUMPRODUCT formula
        ws.cell(row=row, column=3, value=f"=SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, E{start_data_row}:E{end_data_row})")
        # Weighted Return: SUMPRODUCT formula
        ws.cell(row=row, column=4, value=f"=SUMPRODUCT({w_col}{start_data_row}:{w_col}{end_data_row}/100, D{start_data_row}:D{end_data_row})")
        # Formula explanation
        ws.cell(row=row, column=5, value=f"Sharpe=(w×μ)/√(w×Σ×w), ρ={rho}")
        row += 1
    
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 22
    
    return ws

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    print("\n" + "=" * 80)
    print("Creating Excel sheets...")
    print("=" * 80)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    print("\nCreating Sheet 1: Strategy Statistics...")
    create_sheet1_statistics(wb)
    
    print("Creating Sheet 2: Pair Capital Distribution...")
    create_sheet2_pair_allocation(wb)
    
    print("Creating Sheet 3: Strategy Capital Distribution...")
    create_sheet3_strategy_allocation(wb)
    
    output_path = os.path.join(BASE_PATH, 'Portfolio_Analysis_3Sheets.xlsx')
    wb.save(output_path)
    print(f"\n✓ Workbook saved to: {output_path}")

if __name__ == "__main__":
    main()
