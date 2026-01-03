"""
Portfolio Allocation Simulator - Dynamic Excel Version
Generates an Excel sheet with formulas that automatically recalculate
when starting balance is changed.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter


def parse_strategy_statistics(xl):
    """Parse the Strategy_Statistics sheet to extract pair-level data."""
    df = pd.read_excel(xl, sheet_name='Strategy_Statistics', header=None)
    
    strategies = {}
    current_strategy = None
    
    for idx, row in df.iterrows():
        first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        
        if first_cell.startswith('Strategy:'):
            current_strategy = first_cell.replace('Strategy:', '').strip()
            strategies[current_strategy] = []
        
        elif current_strategy and pd.notna(row.iloc[1]) and row.iloc[1] != 'Sharpe_Ratio':
            try:
                sharpe = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                total_trades = int(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
                total_profit = float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
                max_dd = float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
                profit_factor = float(row.iloc[10]) if pd.notna(row.iloc[10]) else 0
                trading_days = int(row.iloc[11]) if pd.notna(row.iloc[11]) else 0
                trading_years = trading_days / 365 if trading_days > 0 else 5
                
                if first_cell and first_cell != 'STRATEGY TOTAL':
                    strategies[current_strategy].append({
                        'pair': first_cell,
                        'sharpe': sharpe,
                        'total_trades': total_trades,
                        'total_profit': total_profit,
                        'max_dd': max_dd,
                        'profit_factor': profit_factor,
                        'trading_years': trading_years,
                        'initial_capital': max_dd * 2
                    })
            except (ValueError, TypeError):
                continue
    
    return strategies


def parse_pair_allocation(xl):
    """Parse the Pair_Capital_Distribution sheet to get allocation percentages."""
    df = pd.read_excel(xl, sheet_name='Pair_Capital_Distribution', header=None)
    
    allocations = {}
    current_strategy = None
    
    for idx, row in df.iterrows():
        first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        
        if first_cell.startswith('STRATEGY:'):
            strategy_name = first_cell.replace('STRATEGY:', '').strip()
            strategy_name = strategy_name.split('(')[0].strip()
            current_strategy = strategy_name
            allocations[current_strategy] = {}
        
        elif current_strategy and pd.notna(row.iloc[0]) and row.iloc[0] not in ['Currency_Pair', 'TOTAL', '']:
            try:
                pair = str(row.iloc[0])
                if pair and pair != 'nan':
                    allocations[current_strategy][pair] = {
                        'equal': float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0,
                        'inv_vol': float(row.iloc[2]) if pd.notna(row.iloc[2]) else 0,
                        'sharpe': float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0,
                        'risk_parity': float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0,
                        'max_sharpe': float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0,
                    }
            except (ValueError, TypeError):
                continue
    
    return allocations


def parse_strategy_allocation(xl):
    """Parse the Strategy_Capital_Distribution sheet."""
    df = pd.read_excel(xl, sheet_name='Strategy_Capital_Distribution', header=None)
    
    allocations = {}
    
    for idx, row in df.iterrows():
        first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        
        if first_cell and first_cell not in ['Strategy', 'TOTAL', 'nan', '', 'NaN']:
            if first_cell.startswith('STRATEGY') or first_cell.startswith('FINAL') or first_cell.startswith('Use'):
                continue
            try:
                pairs = int(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                sharpe = float(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
                capital_req = float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
                profit = float(row.iloc[11]) if pd.notna(row.iloc[11]) else 0
                
                if pairs > 0:
                    allocations[first_cell] = {
                        'pairs': pairs,
                        'sharpe': sharpe,
                        'capital_req': capital_req,
                        'profit': profit,
                    }
            except (ValueError, TypeError):
                continue
    
    return allocations


def get_pair_allocation_method(strategy_name):
    """Return the pair-level allocation method for each strategy."""
    methods = {
        '7th_Strategy': 'sharpe',
        'Falcon': 'risk_parity',
        'Gold_Dip': 'risk_parity',
        'RSI_Correlation': 'inv_vol',
        'RSI_6_Trades': 'risk_parity',
        'AURUM': 'risk_parity',
        'PairTradingEA': 'max_sharpe',
        'Reversal_Strategy': 'risk_parity'
    }
    return methods.get(strategy_name, 'equal')


def prepare_data(strategy_stats, pair_alloc, strategy_alloc):
    """Prepare all data for Excel output."""
    num_strategies = len(strategy_alloc)
    strategy_weight = 100 / num_strategies if num_strategies > 0 else 0
    
    results = []
    
    for strategy_name in strategy_alloc.keys():
        pairs_data = strategy_stats.get(strategy_name, [])
        pair_allocations = pair_alloc.get(strategy_name, {})
        alloc_method = get_pair_allocation_method(strategy_name)
        
        for pair_info in pairs_data:
            pair_name = pair_info['pair']
            pair_pct_data = pair_allocations.get(pair_name, {})
            pair_pct = pair_pct_data.get(alloc_method, 100 / len(pairs_data) if pairs_data else 100)
            
            if pair_pct == 0:
                pair_pct = 100 / len(pairs_data) if pairs_data else 100
            
            historical_profit = pair_info['total_profit']
            historical_capital = pair_info['initial_capital']
            trading_years = pair_info['trading_years']
            
            if historical_capital > 0 and trading_years > 0:
                annual_return_pct = (historical_profit / historical_capital / trading_years) * 100
            else:
                annual_return_pct = 0
            
            results.append({
                'strategy': strategy_name,
                'pair': pair_name,
                'strategy_weight': strategy_weight,
                'pair_weight': pair_pct,
                'alloc_method': alloc_method.replace('_', ' ').title(),
                'annual_return': annual_return_pct,
                'sharpe': pair_info['sharpe'],
                'max_dd': pair_info['max_dd'],
                'required_capital': pair_info['initial_capital'],
                'historical_profit': historical_profit,
                'trading_years': trading_years
            })
    
    return results, num_strategies


def create_allocation_method_sheet(wb, sheet_name, data, strategy_stats, pair_alloc, 
                                    allocation_method, balance_cell_ref):
    """Create a sheet for a specific allocation method showing profit simulation."""
    ws = wb.create_sheet(title=sheet_name)
    
    # Styles
    title_font = Font(bold=True, size=16, color="2F5496")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    money_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    method_display_names = {
        'equal': 'Equal Weight',
        'inv_vol': 'Inverse Volatility',
        'sharpe': 'Sharpe Weighted',
        'risk_parity': 'Risk Parity',
        'max_sharpe': 'Max Sharpe Optimization'
    }
    
    # ========== ROW 1: Title ==========
    row = 1
    ws.cell(row=row, column=1, value=f"PROFIT SIMULATION - {method_display_names.get(allocation_method, allocation_method).upper()}")
    ws.cell(row=row, column=1).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    
    # ========== ROW 3: Reference to Starting Balance ==========
    row = 3
    ws.cell(row=row, column=1, value="Starting Balance (from main sheet):")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=f"='Portfolio Allocation'!{balance_cell_ref}")
    ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="008000")
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    ws.cell(row=row, column=2).fill = formula_fill
    LOCAL_BALANCE = "B3"
    
    # ========== PORTFOLIO SUMMARY ==========
    row = 5
    ws.cell(row=row, column=1, value="PORTFOLIO SUMMARY")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row += 1
    TOTAL_PROFIT_ROW = row
    ws.cell(row=row, column=1, value="Total Expected Annual Profit:")
    ws.cell(row=row, column=1).font = subheader_font
    # Will be set later
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = money_fill
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    
    row += 1
    TOTAL_RETURN_ROW = row
    ws.cell(row=row, column=1, value="Portfolio Annual Return:")
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = money_fill
    ws.cell(row=row, column=2).number_format = '0.00%'
    
    # ========== STRATEGY LEVEL ALLOCATION ==========
    row = 10
    ws.cell(row=row, column=1, value="STRATEGY LEVEL ALLOCATION")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    
    row += 1
    strat_headers = ['Strategy', 'Pairs', 'Strategy Weight %', 'Allocated Capital', 
                     'Expected Annual Profit', 'Return %']
    for col, header in enumerate(strat_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    STRAT_DATA_START = row
    
    # Group data by strategy and calculate weights using this method
    strategy_data = {}
    for item in data:
        strat = item['strategy']
        if strat not in strategy_data:
            strategy_data[strat] = {
                'pairs': 0,
                'pair_items': []
            }
        strategy_data[strat]['pairs'] += 1
        strategy_data[strat]['pair_items'].append(item)
    
    num_strategies = len(strategy_data)
    strategy_weight = 1 / num_strategies if num_strategies > 0 else 0
    
    strategy_rows = {}
    for strat_name, strat_info in strategy_data.items():
        strategy_rows[strat_name] = row
        
        ws.cell(row=row, column=1, value=strat_name)
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=strat_info['pairs'])
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=3, value=strategy_weight)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).number_format = '0.00%'
        
        # Allocated Capital formula
        ws.cell(row=row, column=4, value=f"={LOCAL_BALANCE}*C{row}")
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).fill = formula_fill
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        
        # Expected Profit - will be set later via SUMIF
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).fill = formula_fill
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        
        # Return %
        ws.cell(row=row, column=6, value=f"=IF(D{row}>0,E{row}/D{row},0)")
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = formula_fill
        ws.cell(row=row, column=6).number_format = '0.00%'
        
        row += 1
    
    STRAT_DATA_END = row - 1
    
    # Strategy Total Row
    STRAT_TOTAL_ROW = row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=1).fill = money_fill
    
    ws.cell(row=row, column=2, value=f"=SUM(B{STRAT_DATA_START}:B{STRAT_DATA_END})")
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=2).fill = money_fill
    
    ws.cell(row=row, column=3, value=f"=SUM(C{STRAT_DATA_START}:C{STRAT_DATA_END})")
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=3).fill = money_fill
    ws.cell(row=row, column=3).number_format = '0.00%'
    
    ws.cell(row=row, column=4, value=f"=SUM(D{STRAT_DATA_START}:D{STRAT_DATA_END})")
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=4).border = border
    ws.cell(row=row, column=4).fill = money_fill
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    
    ws.cell(row=row, column=5, value=f"=SUM(E{STRAT_DATA_START}:E{STRAT_DATA_END})")
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=5).fill = money_fill
    ws.cell(row=row, column=5).number_format = '$#,##0.00'
    
    ws.cell(row=row, column=6, value=f"=IF(D{row}>0,E{row}/D{row},0)")
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).border = border
    ws.cell(row=row, column=6).fill = money_fill
    ws.cell(row=row, column=6).number_format = '0.00%'
    
    # ========== PAIR LEVEL ALLOCATION ==========
    row += 3
    ws.cell(row=row, column=1, value="PAIR LEVEL ALLOCATION")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    
    row += 1
    pair_headers = ['Strategy', 'Pair', 'Strategy Wt %', 'Pair Wt %', 'Combined Wt %',
                    'Allocated Capital', 'Annual Return %', 'Expected Profit', 
                    'Sharpe', 'Max DD']
    for col, header in enumerate(pair_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row += 1
    PAIR_DATA_START = row
    
    # Sort data by strategy then pair
    sorted_data = sorted(data, key=lambda x: (x['strategy'], x['pair']))
    
    current_strategy = None
    for item in sorted_data:
        if current_strategy != item['strategy']:
            if current_strategy is not None:
                row += 1  # Blank row between strategies
            current_strategy = item['strategy']
        
        # Get pair weight for this allocation method
        strat_pair_alloc = pair_alloc.get(item['strategy'], {})
        pair_pct_data = strat_pair_alloc.get(item['pair'], {})
        pair_weight = pair_pct_data.get(allocation_method, 0)
        
        # If no allocation for this method, use equal weight
        if pair_weight == 0:
            num_pairs = strategy_data[item['strategy']]['pairs']
            pair_weight = 100 / num_pairs if num_pairs > 0 else 100
        
        pair_weight_decimal = pair_weight / 100
        
        # Col A: Strategy
        ws.cell(row=row, column=1, value=item['strategy'])
        ws.cell(row=row, column=1).border = border
        
        # Col B: Pair
        ws.cell(row=row, column=2, value=item['pair'])
        ws.cell(row=row, column=2).border = border
        
        # Col C: Strategy Weight %
        ws.cell(row=row, column=3, value=strategy_weight)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).number_format = '0.00%'
        
        # Col D: Pair Weight %
        ws.cell(row=row, column=4, value=pair_weight_decimal)
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).number_format = '0.00%'
        
        # Col E: Combined Weight % (formula)
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).fill = formula_fill
        ws.cell(row=row, column=5).number_format = '0.00%'
        
        # Col F: Allocated Capital (formula)
        ws.cell(row=row, column=6, value=f"={LOCAL_BALANCE}*E{row}")
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = formula_fill
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        
        # Col G: Annual Return % (static from historical)
        ws.cell(row=row, column=7, value=item['annual_return'] / 100)
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).number_format = '0.00%'
        
        # Col H: Expected Profit (formula)
        ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
        ws.cell(row=row, column=8).border = border
        ws.cell(row=row, column=8).fill = formula_fill
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        
        # Col I: Sharpe
        ws.cell(row=row, column=9, value=item['sharpe'])
        ws.cell(row=row, column=9).border = border
        ws.cell(row=row, column=9).number_format = '0.00'
        
        # Col J: Max DD
        ws.cell(row=row, column=10, value=item['max_dd'])
        ws.cell(row=row, column=10).border = border
        ws.cell(row=row, column=10).number_format = '$#,##0.00'
        
        row += 1
    
    PAIR_DATA_END = row - 1
    
    # Pair Total Row
    row += 1
    PAIR_TOTAL_ROW = row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=1).fill = money_fill
    
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = money_fill
    
    # Sum Combined Weight
    ws.cell(row=row, column=5, value=f"=SUMIF(A{PAIR_DATA_START}:A{PAIR_DATA_END},\"<>\",E{PAIR_DATA_START}:E{PAIR_DATA_END})")
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=5).fill = money_fill
    ws.cell(row=row, column=5).number_format = '0.00%'
    
    # Sum Allocated Capital
    ws.cell(row=row, column=6, value=f"=SUMIF(A{PAIR_DATA_START}:A{PAIR_DATA_END},\"<>\",F{PAIR_DATA_START}:F{PAIR_DATA_END})")
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).border = border
    ws.cell(row=row, column=6).fill = money_fill
    ws.cell(row=row, column=6).number_format = '$#,##0.00'
    
    ws.cell(row=row, column=7).border = border
    ws.cell(row=row, column=7).fill = money_fill
    
    # Sum Expected Profit
    ws.cell(row=row, column=8, value=f"=SUMIF(A{PAIR_DATA_START}:A{PAIR_DATA_END},\"<>\",H{PAIR_DATA_START}:H{PAIR_DATA_END})")
    ws.cell(row=row, column=8).font = Font(bold=True)
    ws.cell(row=row, column=8).border = border
    ws.cell(row=row, column=8).fill = money_fill
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    
    for col in range(9, 11):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = money_fill
    
    # Now set strategy profit formulas using SUMIF
    for strat_name, strat_row in strategy_rows.items():
        formula = f'=SUMIF(A{PAIR_DATA_START}:A{PAIR_DATA_END},A{strat_row},H{PAIR_DATA_START}:H{PAIR_DATA_END})'
        ws.cell(row=strat_row, column=5, value=formula)
    
    # Set portfolio summary formulas
    ws.cell(row=TOTAL_PROFIT_ROW, column=2, value=f"=E{STRAT_TOTAL_ROW}")
    ws.cell(row=TOTAL_RETURN_ROW, column=2, value=f"=IF({LOCAL_BALANCE}>0,B{TOTAL_PROFIT_ROW}/{LOCAL_BALANCE},0)")
    
    # Column widths
    column_widths = {'A': 18, 'B': 14, 'C': 14, 'D': 12, 'E': 14, 
                     'F': 16, 'G': 14, 'H': 16, 'I': 10, 'J': 12}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    return ws


def create_dynamic_excel(data, num_strategies, starting_balance, output_file, 
                         strategy_stats=None, pair_alloc=None):
    """Create Excel with dynamic formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Allocation"
    
    # Styles
    title_font = Font(bold=True, size=18, color="2F5496")
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    money_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    formula_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # ========== ROW 1: Title ==========
    row = 1
    ws.cell(row=row, column=1, value="PORTFOLIO ALLOCATION SIMULATOR")
    ws.cell(row=row, column=1).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    
    # ========== ROW 3: Starting Balance Input ==========
    row = 3
    BALANCE_CELL = "B3"  # This is the key cell that user can change
    
    ws.cell(row=row, column=1, value="STARTING BALANCE:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=starting_balance)
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008000")
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thick_border
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    ws.cell(row=row, column=3, value="← CHANGE THIS VALUE")
    ws.cell(row=row, column=3).font = Font(italic=True, color="FF0000")
    
    # ========== ROW 5-10: Portfolio Summary ==========
    row = 5
    ws.cell(row=row, column=1, value="PORTFOLIO SUMMARY")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row += 1
    num_pairs = len(data)
    
    # Total Strategies (static)
    ws.cell(row=row, column=1, value="Total Strategies:")
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=2, value=num_strategies)
    ws.cell(row=row, column=2).font = Font(bold=True)
    
    row += 1
    # Total Trading Pairs (static)
    ws.cell(row=row, column=1, value="Total Trading Pairs:")
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=2, value=num_pairs)
    ws.cell(row=row, column=2).font = Font(bold=True)
    
    row += 1
    # Total Allocated Capital (formula)
    TOTAL_CAPITAL_ROW = row
    ws.cell(row=row, column=1, value="Total Allocated Capital:")
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=2, value=f"={BALANCE_CELL}")
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = formula_fill
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    
    row += 1
    # Expected Annual Profit (formula - will be updated after detail table)
    PROFIT_SUMMARY_ROW = row
    ws.cell(row=row, column=1, value="Expected Annual Profit:")
    ws.cell(row=row, column=1).font = subheader_font
    # Formula will be set after detail table is created
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = formula_fill
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    
    row += 1
    # Expected Annual Return (formula)
    RETURN_SUMMARY_ROW = row
    ws.cell(row=row, column=1, value="Expected Annual Return:")
    ws.cell(row=row, column=1).font = subheader_font
    # Formula will be set after detail table
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = formula_fill
    ws.cell(row=row, column=2).number_format = '0.00%'
    
    # ========== ROW 13+: Strategy Summary Table ==========
    row = 13
    ws.cell(row=row, column=1, value="STRATEGY ALLOCATION SUMMARY")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    
    row += 1
    strategy_headers = ['Strategy', 'Pairs', 'Weight %', 'Allocated Capital', 'Expected Profit', 'Return %']
    for col, header in enumerate(strategy_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    STRATEGY_HEADER_ROW = row
    row += 1
    STRATEGY_DATA_START = row
    
    # Group data by strategy
    strategy_data = {}
    for item in data:
        strat = item['strategy']
        if strat not in strategy_data:
            strategy_data[strat] = {
                'pairs': 0,
                'weight': item['strategy_weight'],
                'pair_items': []
            }
        strategy_data[strat]['pairs'] += 1
        strategy_data[strat]['pair_items'].append(item)
    
    strategy_rows = {}
    for strat_name, strat_info in strategy_data.items():
        strategy_rows[strat_name] = row
        
        # Strategy name
        ws.cell(row=row, column=1, value=strat_name)
        ws.cell(row=row, column=1).border = border
        
        # Pairs count
        ws.cell(row=row, column=2, value=strat_info['pairs'])
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        # Weight % (static - equal weight)
        weight_pct = strat_info['weight'] / 100
        ws.cell(row=row, column=3, value=weight_pct)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).number_format = '0.00%'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        # Allocated Capital (formula)
        ws.cell(row=row, column=4, value=f"={BALANCE_CELL}*C{row}")
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).fill = formula_fill
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        
        # Expected Profit (will be SUMIF from detail table - placeholder for now)
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=5).fill = formula_fill
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        
        # Return % (formula)
        ws.cell(row=row, column=6, value=f"=IF(D{row}>0,E{row}/D{row},0)")
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = formula_fill
        ws.cell(row=row, column=6).number_format = '0.00%'
        
        row += 1
    
    STRATEGY_DATA_END = row - 1
    
    # Total row for strategy summary
    STRATEGY_TOTAL_ROW = row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=1).fill = money_fill
    
    ws.cell(row=row, column=2, value=f"=SUM(B{STRATEGY_DATA_START}:B{STRATEGY_DATA_END})")
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=2).fill = money_fill
    
    ws.cell(row=row, column=3, value=f"=SUM(C{STRATEGY_DATA_START}:C{STRATEGY_DATA_END})")
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=3).fill = money_fill
    ws.cell(row=row, column=3).number_format = '0.00%'
    
    ws.cell(row=row, column=4, value=f"=SUM(D{STRATEGY_DATA_START}:D{STRATEGY_DATA_END})")
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=4).border = border
    ws.cell(row=row, column=4).fill = money_fill
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    
    ws.cell(row=row, column=5, value=f"=SUM(E{STRATEGY_DATA_START}:E{STRATEGY_DATA_END})")
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=5).fill = money_fill
    ws.cell(row=row, column=5).number_format = '$#,##0.00'
    
    ws.cell(row=row, column=6, value=f"=IF(D{row}>0,E{row}/D{row},0)")
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).border = border
    ws.cell(row=row, column=6).fill = money_fill
    ws.cell(row=row, column=6).number_format = '0.00%'
    
    # ========== DETAILED PAIR ALLOCATION ==========
    row += 3
    DETAIL_HEADER_ROW = row
    ws.cell(row=row, column=1, value="DETAILED PAIR ALLOCATION")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    
    row += 1
    detail_headers = [
        'Strategy', 'Pair', 'Strategy Weight %', 'Pair Weight %', 'Allocation Method',
        'Allocated Capital', 'Annual Return %', 'Expected Profit', 
        'Sharpe Ratio', 'Max Drawdown', 'Min Required Capital'
    ]
    
    for col, header in enumerate(detail_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    DETAIL_COL_HEADERS_ROW = row
    row += 1
    DETAIL_DATA_START = row
    
    # Sort data by strategy then by pair weight descending
    sorted_data = sorted(data, key=lambda x: (x['strategy'], -x['pair_weight']))
    
    current_strategy = None
    pair_rows_by_strategy = {}
    
    for item in sorted_data:
        # Add blank row between strategies
        if current_strategy != item['strategy']:
            if current_strategy is not None:
                row += 1
            current_strategy = item['strategy']
            if current_strategy not in pair_rows_by_strategy:
                pair_rows_by_strategy[current_strategy] = []
        
        pair_rows_by_strategy[current_strategy].append(row)
        
        # Col A: Strategy
        ws.cell(row=row, column=1, value=item['strategy'])
        ws.cell(row=row, column=1).border = border
        
        # Col B: Pair
        ws.cell(row=row, column=2, value=item['pair'])
        ws.cell(row=row, column=2).border = border
        
        # Col C: Strategy Weight % (static)
        ws.cell(row=row, column=3, value=item['strategy_weight'] / 100)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).number_format = '0.00%'
        
        # Col D: Pair Weight % (static)
        ws.cell(row=row, column=4, value=item['pair_weight'] / 100)
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).number_format = '0.00%'
        
        # Col E: Allocation Method
        ws.cell(row=row, column=5, value=item['alloc_method'])
        ws.cell(row=row, column=5).border = border
        
        # Col F: Allocated Capital (FORMULA)
        # = Starting Balance * Strategy Weight * Pair Weight
        formula = f"={BALANCE_CELL}*C{row}*D{row}"
        ws.cell(row=row, column=6, value=formula)
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=6).fill = formula_fill
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        
        # Col G: Annual Return % (static - based on historical data)
        ws.cell(row=row, column=7, value=item['annual_return'] / 100)
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=7).number_format = '0.00%'
        
        # Col H: Expected Profit (FORMULA)
        # = Allocated Capital * Annual Return %
        ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
        ws.cell(row=row, column=8).border = border
        ws.cell(row=row, column=8).fill = formula_fill
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        
        # Col I: Sharpe Ratio (static)
        ws.cell(row=row, column=9, value=item['sharpe'])
        ws.cell(row=row, column=9).border = border
        ws.cell(row=row, column=9).number_format = '0.00'
        
        # Col J: Max Drawdown (static)
        ws.cell(row=row, column=10, value=item['max_dd'])
        ws.cell(row=row, column=10).border = border
        ws.cell(row=row, column=10).number_format = '$#,##0.00'
        
        # Col K: Required Capital (static)
        ws.cell(row=row, column=11, value=item['required_capital'])
        ws.cell(row=row, column=11).border = border
        ws.cell(row=row, column=11).number_format = '$#,##0.00'
        
        row += 1
    
    DETAIL_DATA_END = row - 1
    
    # Add TOTAL row for detail section
    row += 1
    DETAIL_TOTAL_ROW = row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=1).fill = money_fill
    
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = money_fill
    
    # Sum of Allocated Capital
    ws.cell(row=row, column=6, value=f"=SUMIF(A{DETAIL_DATA_START}:A{DETAIL_DATA_END},\"<>\",F{DETAIL_DATA_START}:F{DETAIL_DATA_END})")
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).border = border
    ws.cell(row=row, column=6).fill = money_fill
    ws.cell(row=row, column=6).number_format = '$#,##0.00'
    
    ws.cell(row=row, column=7).border = border
    ws.cell(row=row, column=7).fill = money_fill
    
    # Sum of Expected Profit
    ws.cell(row=row, column=8, value=f"=SUMIF(A{DETAIL_DATA_START}:A{DETAIL_DATA_END},\"<>\",H{DETAIL_DATA_START}:H{DETAIL_DATA_END})")
    ws.cell(row=row, column=8).font = Font(bold=True)
    ws.cell(row=row, column=8).border = border
    ws.cell(row=row, column=8).fill = money_fill
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    
    for col in range(9, 12):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = money_fill
    
    # Now go back and set the strategy expected profit formulas using SUMIF
    for strat_name, strat_row in strategy_rows.items():
        # Expected Profit = Sum of all profits for this strategy in detail table
        formula = f'=SUMIF(A{DETAIL_DATA_START}:A{DETAIL_DATA_END},A{strat_row},H{DETAIL_DATA_START}:H{DETAIL_DATA_END})'
        ws.cell(row=strat_row, column=5, value=formula)
    
    # Set summary formulas
    ws.cell(row=PROFIT_SUMMARY_ROW, column=2, value=f"=E{STRATEGY_TOTAL_ROW}")
    ws.cell(row=RETURN_SUMMARY_ROW, column=2, value=f"=IF(B{TOTAL_CAPITAL_ROW}>0,B{PROFIT_SUMMARY_ROW}/B{TOTAL_CAPITAL_ROW},0)")
    
    # ========== Column Widths ==========
    column_widths = {
        'A': 22, 'B': 28, 'C': 16, 'D': 14, 'E': 18,
        'F': 18, 'G': 16, 'H': 18, 'I': 12, 'J': 14, 'K': 16
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ========== CREATE 5 ALLOCATION METHOD SHEETS ==========
    allocation_methods = [
        ('equal', 'Equal Weight'),
        ('inv_vol', 'Inv Volatility'),
        ('sharpe', 'Sharpe Weighted'),
        ('risk_parity', 'Risk Parity'),
        ('max_sharpe', 'Max Sharpe')
    ]
    
    if strategy_stats and pair_alloc:
        print("\n📊 Creating allocation method simulation sheets...")
        for method_key, sheet_name in allocation_methods:
            create_allocation_method_sheet(
                wb=wb,
                sheet_name=sheet_name,
                data=data,
                strategy_stats=strategy_stats,
                pair_alloc=pair_alloc,
                allocation_method=method_key,
                balance_cell_ref=BALANCE_CELL
            )
            print(f"   ✓ Created sheet: {sheet_name}")
    
    # Save
    wb.save(output_file)
    print(f"\n✅ Dynamic Excel file saved: {output_file}")
    print(f"   → Change cell B3 to update all allocations automatically!")
    print(f"   → 5 allocation method sheets show profit simulations for each method!")


def get_starting_balance():
    """Prompt user for starting balance."""
    while True:
        try:
            balance = float(input("\n💰 Enter your starting balance ($): "))
            if balance <= 0:
                print("Please enter a positive amount.")
                continue
            return balance
        except ValueError:
            print("Invalid input. Please enter a numeric value.")


def main():
    print("=" * 60)
    print("   PORTFOLIO ALLOCATION SIMULATOR (Dynamic Formulas)")
    print("=" * 60)
    
    excel_file = 'Portfolio_Analysis_Sheets.xlsx'
    print(f"\n📊 Loading data from: {excel_file}")
    
    try:
        xl = pd.ExcelFile(excel_file)
    except FileNotFoundError:
        print(f"❌ Error: Could not find {excel_file}")
        return
    
    print("📈 Parsing strategy statistics...")
    strategy_stats = parse_strategy_statistics(xl)
    
    print("📊 Parsing pair allocations...")
    pair_alloc = parse_pair_allocation(xl)
    
    print("💼 Parsing strategy allocations...")
    strategy_alloc = parse_strategy_allocation(xl)
    
    print(f"\n✅ Loaded {len(strategy_alloc)} strategies with {sum(len(v) for v in strategy_stats.values())} pairs")
    
    # Get starting balance
    starting_balance = get_starting_balance()
    
    # Prepare data
    print("\n🔢 Preparing allocation data...")
    data, num_strategies = prepare_data(strategy_stats, pair_alloc, strategy_alloc)
    
    # Create dynamic Excel
    output_file = 'Portfolio_Allocation_Dynamic.xlsx'
    create_dynamic_excel(data, num_strategies, starting_balance, output_file,
                         strategy_stats=strategy_stats, pair_alloc=pair_alloc)
    
    # Display summary
    total_pairs = len(data)
    avg_return = sum(d['annual_return'] for d in data) / len(data) if data else 0
    
    print("\n" + "=" * 60)
    print("   SUMMARY")
    print("=" * 60)
    print(f"💰 Starting Balance: ${starting_balance:,.2f}")
    print(f"📊 Strategies: {num_strategies}")
    print(f"🔗 Trading Pairs: {total_pairs}")
    print(f"\n📝 The Excel file contains dynamic formulas.")
    print(f"   Just change cell B3 to recalculate everything!")
    print(f"\n📊 Sheets Created:")
    print(f"   1. Portfolio Allocation - Main allocation view")
    print(f"   2. Equal Weight - Equal weight simulation")
    print(f"   3. Inv Volatility - Inverse volatility simulation")
    print(f"   4. Sharpe Weighted - Sharpe ratio weighted simulation")
    print(f"   5. Risk Parity - Risk parity simulation")
    print(f"   6. Max Sharpe - Max Sharpe optimization simulation")
    print("=" * 60)


if __name__ == "__main__":
    main()
